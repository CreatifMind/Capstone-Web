from __future__ import annotations

import os
import re
import json
import math
import tempfile
import traceback
import threading
import time
import random
import shutil
import subprocess
import hashlib
import hmac
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from io import BytesIO
from mimetypes import guess_type
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse
from uuid import UUID, uuid4, uuid5
from typing import Any, Callable, TypeVar
from zoneinfo import ZoneInfo

import httpx
from dotenv import load_dotenv
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, RedirectResponse, Response
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from PIL import Image, ImageOps, UnidentifiedImageError
from pydantic import BaseModel
from supabase import create_client
from ultralytics import YOLO

BACKEND_ROOT = Path(__file__).resolve().parent
APP_ROOT = BACKEND_ROOT.parent

load_dotenv(BACKEND_ROOT / ".env")


DRIVE_ID_PATTERN = re.compile(r"^[A-Za-z0-9_-]+$")


def normalize_drive_folder_id(value: str | None) -> str | None:
    if value is None:
        return None
    folder_id = value.strip()
    if not folder_id:
        return None

    parsed = urlparse(folder_id)
    if parsed.scheme or parsed.netloc:
        if parsed.scheme not in {"http", "https"}:
            raise ValueError("Google Drive folder id must be a raw id or HTTPS URL")
        candidates: list[str] = []
        parts = [part for part in parsed.path.split("/") if part]
        for index, part in enumerate(parts[:-1]):
            if part == "folders":
                candidates.append(parts[index + 1])
        for key in ("id", "folderId"):
            candidates.extend(parse_qs(parsed.query).get(key, []))
        unique = list(dict.fromkeys(item.strip() for item in candidates if item.strip()))
        if len(unique) != 1:
            raise ValueError("Google Drive folder URL must contain exactly one folder id")
        folder_id = unique[0]

    if not DRIVE_ID_PATTERN.fullmatch(folder_id):
        raise ValueError("Google Drive folder id contains invalid characters")
    return folder_id


def configured_drive_folder_id() -> str | None:
    return normalize_drive_folder_id(os.getenv("GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID") or os.getenv("GOOGLE_DRIVE_FOLDER_ID"))

_model_path = Path(os.getenv("MODEL_PATH", str(BACKEND_ROOT / "models" / "best.pt"))).expanduser()
if _model_path.is_absolute():
    MODEL_PATH = _model_path
else:
    MODEL_PATH = next((path for path in (BACKEND_ROOT / _model_path, APP_ROOT / _model_path) if path.exists()), BACKEND_ROOT / _model_path)

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID = configured_drive_folder_id()
GOOGLE_OAUTH_REDIRECT_URI = os.getenv("GOOGLE_OAUTH_REDIRECT_URI", "http://localhost:8000/api/google/callback")


def cors_origins() -> list[str]:
    environment = os.getenv("ENVIRONMENT", "development").strip().lower()
    if environment == "production":
        origin = os.getenv("FRONTEND_ORIGIN", "https://purityloop-ai.vercel.app").strip().rstrip("/")
        if (
            not origin
            or origin == "*"
            or origin.startswith("http://localhost")
            or origin.startswith("http://127.0.0.1")
            or not origin.startswith("https://")
        ):
            raise RuntimeError("Production FRONTEND_ORIGIN must be a single HTTPS origin.")
        return [origin]

    origins = [
        "http://localhost:3000",
        "https://purityloop-ai.vercel.app",
    ]
    for env_name in ("FRONTEND_ORIGIN", "CORS_ORIGINS", "ALLOWED_ORIGINS"):
        origins.extend(item.strip().rstrip("/") for item in os.getenv(env_name, "").split(",") if item.strip())
    return list(dict.fromkeys(origins))


ALLOWED_ORIGINS = cors_origins()

app = FastAPI(title="PurityLoop AI Backend")
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

model = None
model_device_effective = None


def _new_supabase_client():
    return create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY) if SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY else None


# Request handlers may keep using this client. Workers always create their own executor.
supabase = _new_supabase_client()
SCAN_RESULTS_TABLE = "scan_results"
DETECTED_MATERIALS_TABLE = "detected_materials"
REVIEW_DECISIONS_TABLE = "scan_review_decisions"
JOBS_TABLE = "processing_jobs"
UPLOAD_SESSIONS_TABLE = "upload_sessions"
PROCESSED_DRIVE_FILES_TABLE = "processed_drive_files"
FALSE_POSITIVE_REPORTS_TABLE = "false_positive_reports"
PREVIEW_BUCKET = os.getenv("SUPABASE_STORAGE_BUCKET", "mock_uploaded_images")
VIDEO_WORK_ROOT = Path(os.getenv("VIDEO_WORK_ROOT", "/tmp/purityloop"))
SERVICE_MODE = os.getenv("SERVICE_MODE", "api").strip().lower() or "api"
PROCESSING_BACKEND = os.getenv("PROCESSING_BACKEND") or ("cloud-tasks" if os.getenv("ENVIRONMENT", "").lower() == "production" else "local-thread")
MODEL_DEVICE = os.getenv("MODEL_DEVICE", "cpu").strip() or "cpu"
CLOUD_TASKS_PROJECT_ID = os.getenv("CLOUD_TASKS_PROJECT_ID")
CLOUD_TASKS_LOCATION = os.getenv("CLOUD_TASKS_LOCATION")
CLOUD_TASKS_QUEUE = os.getenv("CLOUD_TASKS_QUEUE")
CLOUD_TASKS_WORKER_URL = os.getenv("CLOUD_TASKS_WORKER_URL")
CLOUD_TASKS_OIDC_AUDIENCE = os.getenv("CLOUD_TASKS_OIDC_AUDIENCE")
CLOUD_TASKS_CALLER_SERVICE_ACCOUNT = os.getenv("CLOUD_TASKS_CALLER_SERVICE_ACCOUNT")
CLOUD_TASKS_DISPATCH_DEADLINE_SECONDS = int(os.getenv("CLOUD_TASKS_DISPATCH_DEADLINE_SECONDS", "900") or 900)
JOB_LEASE_SECONDS = int(os.getenv("JOB_LEASE_SECONDS", "1200") or 1200)
DEFAULT_VIDEO_FPS = float(os.getenv("DEFAULT_VIDEO_FPS", "30") or 30)
UPLOAD_CHUNK_SIZE_BYTES = 8 * 1024 * 1024
MAX_VIDEO_UPLOAD_BYTES = int(os.getenv("MAX_VIDEO_UPLOAD_BYTES", "0") or 0)
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.file"]
# The upload folder is configured server-side, not selected with Google Picker.
# OAuth therefore needs access to that existing folder and its idempotency search.
OAUTH_DRIVE_SCOPES = [
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/drive.file",
]
MODEL_CANDIDATE_THRESHOLD = 0.10
DECISION_CONFIDENCE_THRESHOLD = 0.32
CONFIRMATION_THRESHOLD = DECISION_CONFIDENCE_THRESHOLD
PREVIEW_BOX_CONFIDENCE_THRESHOLD = 0.25
PREVIEW_BOX_NMS_IOU_THRESHOLD = 0.50
PREVIEW_EDGE_STRIP_ASPECT_RATIO_MIN = 0.15
PREVIEW_EDGE_STRIP_ASPECT_RATIO_MAX = 6.0
PREVIEW_EDGE_TOLERANCE_RATIO = 0.01
PREVIEW_EDGE_TOLERANCE_MIN_PIXELS = 2.0
ANALYTICS_PAGE_SIZE = 500
ANALYTICS_CHILD_PAGE_SIZE = 500
SCAN_HISTORY_DEFAULT_LIMIT = 10
SCAN_HISTORY_MAX_LIMIT = 100
SCAN_HISTORY_EXPORT_BATCH_SIZE = 500
SCAN_HISTORY_EXPORT_CHILD_BATCH_SIZE = 200
SCAN_HISTORY_THUMBNAIL_SIZE = (80, 80)
SCAN_HISTORY_THUMBNAIL_WORKERS = 8
SCAN_HISTORY_IMAGE_TIMEOUT = httpx.Timeout(5.0, connect=3.0)
SCAN_HISTORY_IMAGE_MAX_BYTES = 2 * 1024 * 1024
MALAYSIA_TIMEZONE = ZoneInfo("Asia/Kuala_Lumpur")
ANALYTICS_MALAYSIA_TZ = MALAYSIA_TIMEZONE
ANALYTICS_MATERIAL_ESTIMATES = {
    "general trash": {"label": "General Trash", "average_weight_kg": 0.100, "price_per_kg_rm": 0.00, "material_class": "contaminant"},
    "food organic": {"label": "Food Organic", "average_weight_kg": 0.080, "price_per_kg_rm": 0.00, "material_class": "contaminant"},
    "metal": {"label": "Metal", "average_weight_kg": 0.020, "price_per_kg_rm": 1.20, "material_class": "recyclable"},
    "plastic": {"label": "Plastic", "average_weight_kg": 0.032, "price_per_kg_rm": 0.50, "material_class": "recyclable"},
    "glass": {"label": "Glass", "average_weight_kg": 0.300, "price_per_kg_rm": 0.10, "material_class": "recyclable"},
    "textile": {"label": "Textile", "average_weight_kg": 0.150, "price_per_kg_rm": 0.00, "material_class": "contaminant"},
    "paper": {"label": "Paper", "average_weight_kg": 0.005, "price_per_kg_rm": 0.30, "material_class": "recyclable"},
    "battery": {"label": "Battery", "average_weight_kg": 0.023, "price_per_kg_rm": 3.50, "material_class": "contaminant"},
    "cardboard": {"label": "Cardboard", "average_weight_kg": 0.125, "price_per_kg_rm": 0.25, "material_class": "recyclable"},
}
BROWSER_CONFIDENCE_THRESHOLD = MODEL_CANDIDATE_THRESHOLD
BROWSER_NMS_IOU_THRESHOLD = 0.70
BROWSER_MODEL_NAME = "best.onnx"
BROWSER_MODEL_PATH = APP_ROOT / "public" / "models" / "purityloop" / BROWSER_MODEL_NAME
BROWSER_MODEL_VERSION = "v3_ffremask_9cls"
BROWSER_INFERENCE_ENGINE = "browser-onnx"
BACKEND_BUILD_VERSION = "browser-confidence-object-metrics-fix-20260805"
BROWSER_MODEL_CLASSES = (
    "plastic", "paper", "cardboard", "metal", "glass", "textile", "food_organic", "battery", "general_trash",
)
CLASS_THRESHOLDS = {
    0: 0.12,  # plastic
    1: 0.20,  # paper
    2: 0.20,  # cardboard
    3: 0.18,  # metal
    4: 0.20,  # glass
    5: 0.25,  # textile
    6: 0.15,  # food_organic
    7: 0.25,  # battery
    8: 0.10,  # general_trash
}
YOLO_CALIBRATION_CANDIDATE_CONFIDENCE = 0.05
GENERAL_TRASH_CATEGORY = "general_trash"
BROWSER_CONFIDENCE_DETAIL = "Detection confidence must be between 0 and 1."
BROWSER_DECISION_CONFIDENCE_THRESHOLD = DECISION_CONFIDENCE_THRESHOLD
BROWSER_CONFIDENCE_CONTRACT_DETAIL = f"Browser confidence threshold must be {BROWSER_DECISION_CONFIDENCE_THRESHOLD:.2f}."
BROWSER_NMS_CONTRACT_DETAIL = f"Browser NMS IoU threshold must be {BROWSER_NMS_IOU_THRESHOLD:.2f}."
MAX_IMAGE_BYTES = 10 * 1024 * 1024
SUPABASE_RETRY_ATTEMPTS = 6
SUPABASE_TRANSIENT_ERRORS = (
    httpx.RemoteProtocolError,
    httpx.ConnectError,
    httpx.ReadError,
    httpx.WriteError,
    httpx.TimeoutException,
)
T = TypeVar("T")


class SupabaseTemporarilyUnavailable(RuntimeError):
    """A transient transport failure after bounded reconnect attempts."""


class SupabaseExecutor:
    """Retry one Supabase operation and replace a broken PostgREST client."""

    def __init__(self, client: Any | None = None, client_factory: Callable[[], Any] = _new_supabase_client,
                 attempts: int = SUPABASE_RETRY_ATTEMPTS, sleeper: Callable[[float], None] = time.sleep,
                 random_value: Callable[[], float] = random.random):
        self.client = client or client_factory()
        self.client_factory = client_factory
        self.attempts = attempts
        self.sleeper = sleeper
        self.random_value = random_value

    def execute(self, operation: Callable[[Any], T], recover: Callable[[Any], T | None] | None = None) -> T:
        if not self.client:
            raise RuntimeError("Supabase backend env is not configured")
        last_error: Exception | None = None
        for attempt in range(self.attempts):
            try:
                return operation(self.client)
            except SUPABASE_TRANSIENT_ERRORS as exc:
                last_error = exc
                # A protocol disconnect can leave the synchronous HTTP/2 connection unusable.
                self.client = self.client_factory()
                if recover and self.client:
                    try:
                        recovered = recover(self.client)
                        if recovered is not None:
                            return recovered
                    except SUPABASE_TRANSIENT_ERRORS:
                        self.client = self.client_factory()
                if attempt + 1 < self.attempts:
                    self.sleeper(min(4.0, 0.25 * (2 ** attempt)) + (self.random_value() * 0.2))
        raise SupabaseTemporarilyUnavailable("Supabase connection temporarily unavailable") from last_error
CATEGORY_CLASS_MAP = {
    "general_trash": "contaminant",
    "food_organics": "contaminant",
    "textile": "contaminant",
    "battery": "contaminant",
    "metal": "recyclable",
    "plastic": "recyclable",
    "glass": "recyclable",
    "paper": "recyclable",
    "cardboard": "recyclable",
}
CATEGORY_ROUTES = {
    "general_trash": "General-Waste Disposal",
    "food_organics": "Organic Waste / Compost",
    "textile": "Textile Recovery / Contaminant Route",
    "battery": "Battery / E-Waste Collection",
    "metal": "Metal Sorting Bin",
    "plastic": "Plastic Sorting Bin",
    "glass": "Glass Sorting Bin",
    "paper": "Paper Sorting Bin",
    "cardboard": "Cardboard Sorting Bin",
}
VIDEO_TRACK_MIN_FRAMES = max(1, int(os.getenv("VIDEO_TRACK_MIN_FRAMES", "3")))
VIDEO_TRACK_SHORT_CONFIDENCE = float(os.getenv("VIDEO_TRACK_SHORT_CONFIDENCE", "0.92"))
VIDEO_TRACK_LOST_BUFFER = max(1, int(os.getenv("VIDEO_TRACK_LOST_BUFFER", "15")))
VIDEO_TRACK_RECOVERY_IOU = float(os.getenv("VIDEO_TRACK_RECOVERY_IOU", "0.35"))
VIDEO_TRACK_RECOVERY_CENTER_DISTANCE = float(os.getenv("VIDEO_TRACK_RECOVERY_CENTER_DISTANCE", "0.18"))
VIDEO_LOGICAL_MERGE_MAX_GAP = max(1, int(os.getenv("VIDEO_LOGICAL_MERGE_MAX_GAP", "90")))
VIDEO_LOGICAL_MERGE_CENTER_DISTANCE = float(os.getenv("VIDEO_LOGICAL_MERGE_CENTER_DISTANCE", "0.22"))
VIDEO_LOGICAL_MERGE_SIZE_RATIO = float(os.getenv("VIDEO_LOGICAL_MERGE_SIZE_RATIO", "0.65"))
VIDEO_DUPLICATE_MAX_GAP = max(1, int(os.getenv("VIDEO_DUPLICATE_MAX_GAP", str(VIDEO_LOGICAL_MERGE_MAX_GAP))))
VIDEO_DUPLICATE_MAX_OVERLAP_FRAMES = max(0, int(os.getenv("VIDEO_DUPLICATE_MAX_OVERLAP_FRAMES", "5")))
VIDEO_DUPLICATE_STRONG_OVERLAP_MAX_FRAMES = max(VIDEO_DUPLICATE_MAX_OVERLAP_FRAMES, int(os.getenv("VIDEO_DUPLICATE_STRONG_OVERLAP_MAX_FRAMES", "10")))
VIDEO_DUPLICATE_MEANINGFUL_OVERLAP_FRAMES = max(1, int(os.getenv("VIDEO_DUPLICATE_MEANINGFUL_OVERLAP_FRAMES", "12")))
VIDEO_DUPLICATE_CENTER_DISTANCE = float(os.getenv("VIDEO_DUPLICATE_CENTER_DISTANCE", "0.18"))
VIDEO_DUPLICATE_STRONG_CENTER_DISTANCE = float(os.getenv("VIDEO_DUPLICATE_STRONG_CENTER_DISTANCE", "0.12"))
VIDEO_DUPLICATE_IOU = float(os.getenv("VIDEO_DUPLICATE_IOU", "0.08"))
VIDEO_DUPLICATE_OVERLAP_IOU = float(os.getenv("VIDEO_DUPLICATE_OVERLAP_IOU", "0.30"))
VIDEO_DUPLICATE_STRONG_OVERLAP_IOU = float(os.getenv("VIDEO_DUPLICATE_STRONG_OVERLAP_IOU", "0.90"))
VIDEO_DUPLICATE_SIZE_RATIO = float(os.getenv("VIDEO_DUPLICATE_SIZE_RATIO", "0.65"))
VIDEO_DUPLICATE_STRONG_SIZE_RATIO = float(os.getenv("VIDEO_DUPLICATE_STRONG_SIZE_RATIO", "0.70"))
VIDEO_DUPLICATE_APPEARANCE_SIMILARITY = float(os.getenv("VIDEO_DUPLICATE_APPEARANCE_SIMILARITY", "0.82"))
VIDEO_DUPLICATE_MIN_APPEARANCE_SIMILARITY = float(os.getenv("VIDEO_DUPLICATE_MIN_APPEARANCE_SIMILARITY", "0.45"))
VIDEO_DUPLICATE_STABLE_TRACK_FRAMES = max(1, int(os.getenv("VIDEO_DUPLICATE_STABLE_TRACK_FRAMES", "18")))
VIDEO_DUPLICATE_FULL_FRAME_AREA = float(os.getenv("VIDEO_DUPLICATE_FULL_FRAME_AREA", "0.70"))
VIDEO_DUPLICATE_TRUNCATED_EDGE_MARGIN = float(os.getenv("VIDEO_DUPLICATE_TRUNCATED_EDGE_MARGIN", "0.015"))
VIDEO_CAMERA_MOTION_MIN_RESPONSE = float(os.getenv("VIDEO_CAMERA_MOTION_MIN_RESPONSE", "0.03"))
VIDEO_DUPLICATE_MIN_OBSERVATION_CONFIDENCE = float(os.getenv("VIDEO_DUPLICATE_MIN_OBSERVATION_CONFIDENCE", "0.25"))
VIDEO_DUPLICATE_MIN_RELIABLE_OBSERVATIONS = max(1, int(os.getenv("VIDEO_DUPLICATE_MIN_RELIABLE_OBSERVATIONS", "3")))
VIDEO_DUPLICATE_HANDOVER_MEDIAN_IOU = float(os.getenv("VIDEO_DUPLICATE_HANDOVER_MEDIAN_IOU", "0.80"))
VIDEO_DUPLICATE_WEAK_FRAGMENT_MAX_GAP = max(1, int(os.getenv("VIDEO_DUPLICATE_WEAK_FRAGMENT_MAX_GAP", "18")))
VIDEO_DUPLICATE_APPEARANCE_FRAGMENT_MAX_GAP = max(1, int(os.getenv("VIDEO_DUPLICATE_APPEARANCE_FRAGMENT_MAX_GAP", "30")))
VIDEO_TRACKER_CONFIG = os.getenv("VIDEO_TRACKER_CONFIG", "config/bytetrack_purityloop.yaml")
VIDEO_TRACK_DEBUG_LOGS = os.getenv("VIDEO_TRACK_DEBUG_LOGS", "true").lower() != "false"
VIDEO_PHYSICAL_RECONCILIATION_DEFAULTS = {
    "VIDEO_DUPLICATE_MAX_GAP": "Maximum frame gap considered for automatic fragment association. Higher reduces false splits but increases pair checks.",
    "VIDEO_DUPLICATE_MAX_OVERLAP_FRAMES": "Brief overlap allowed for likely old/new ByteTrack ID handovers.",
    "VIDEO_DUPLICATE_STRONG_OVERLAP_MAX_FRAMES": "Upper bound for short overlap when boxes strongly agree.",
    "VIDEO_DUPLICATE_MEANINGFUL_OVERLAP_FRAMES": "Duration at which simultaneous low-IoU boxes become a hard separation blocker.",
    "VIDEO_DUPLICATE_CENTER_DISTANCE": "Maximum raw or stabilized endpoint distance for non-overlapping fragments.",
    "VIDEO_DUPLICATE_STRONG_CENTER_DISTANCE": "Tighter distance for overlap handovers.",
    "VIDEO_DUPLICATE_OVERLAP_IOU": "Minimum overlap-frame IoU used to reject separate simultaneous boxes.",
    "VIDEO_DUPLICATE_STRONG_OVERLAP_IOU": "IoU threshold for a brief ID handover.",
    "VIDEO_DUPLICATE_SIZE_RATIO": "Minimum compatible width/height ratio; prevents merging differently sized objects.",
    "VIDEO_DUPLICATE_APPEARANCE_SIMILARITY": "Object-only crop similarity that supports a merge.",
    "VIDEO_DUPLICATE_MIN_APPEARANCE_SIMILARITY": "Strong object-only appearance mismatch blocker.",
    "VIDEO_DUPLICATE_FULL_FRAME_AREA": "Representative boxes above this normalized area are low-quality and cannot be positive appearance evidence.",
    "VIDEO_DUPLICATE_MIN_OBSERVATION_CONFIDENCE": "Minimum confidence for an observation to support a hard separation decision.",
    "VIDEO_DUPLICATE_MIN_RELIABLE_OBSERVATIONS": "Minimum valid observations required before a track can prove sustained separation.",
    "VIDEO_DUPLICATE_HANDOVER_MEDIAN_IOU": "Median reliable overlap IoU that supports a brief ByteTrack ID handover.",
    "VIDEO_DUPLICATE_WEAK_FRAGMENT_MAX_GAP": "Maximum gap for category-consistent broad or clipped fragment bridges.",
    "VIDEO_DUPLICATE_APPEARANCE_FRAGMENT_MAX_GAP": "Maximum gap for clipped fragments supported by object-only appearance.",
    "VIDEO_CAMERA_MOTION_MIN_RESPONSE": "Minimum phase-correlation response for camera-motion compensation; invalid motion is ignored.",
}


def _coerce_float(value: Any, fallback: float = 0.0) -> float:
    try:
        number = float(value)
        return number if math.isfinite(number) else fallback
    except (TypeError, ValueError):
        return fallback


def _bbox_iou(first: list[float] | None, second: list[float] | None) -> float:
    if not first or not second or len(first) < 4 or len(second) < 4:
        return 0.0
    ax1, ay1, ax2, ay2 = first[:4]
    bx1, by1, bx2, by2 = second[:4]
    ix1, iy1 = max(ax1, bx1), max(ay1, by1)
    ix2, iy2 = min(ax2, bx2), min(ay2, by2)
    inter = max(0.0, ix2 - ix1) * max(0.0, iy2 - iy1)
    area_a = max(0.0, ax2 - ax1) * max(0.0, ay2 - ay1)
    area_b = max(0.0, bx2 - bx1) * max(0.0, by2 - by1)
    union = area_a + area_b - inter
    return inter / union if union > 0 else 0.0


def _bbox_center(box: list[float] | None) -> tuple[float, float]:
    if not box or len(box) < 4:
        return (0.0, 0.0)
    return ((box[0] + box[2]) / 2, (box[1] + box[3]) / 2)


def _parse_counting_line(options: dict | None) -> dict | None:
    raw = (options or {}).get("counting_line") or os.getenv("VIDEO_COUNTING_LINE")
    if not raw:
        return None
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except json.JSONDecodeError:
            axis, _, position = raw.partition("=")
            raw = {"axis": axis, "position": position}
    if not isinstance(raw, dict):
        return None
    axis = str(raw.get("axis") or "x").lower()
    if axis not in {"x", "y"}:
        axis = "x"
    position = _coerce_float(raw.get("position"), 0.5)
    if position > 1:
        position = position / 100
    direction = str(raw.get("direction") or "any").lower()
    if direction not in {"positive", "negative", "any"}:
        direction = "any"
    return {"axis": axis, "position": max(0.0, min(1.0, position)), "direction": direction}


def _video_debug(event: str, **fields) -> None:
    if not VIDEO_TRACK_DEBUG_LOGS:
        return
    safe = {"event": event, **fields}
    print(f"[video-track] {json.dumps(safe, default=str, sort_keys=True)}")


def _bbox_size(box: list[float] | None) -> tuple[float, float, float]:
    if not box or len(box) < 4:
        return (0.0, 0.0, 0.0)
    width = max(0.0, box[2] - box[0])
    height = max(0.0, box[3] - box[1])
    aspect = width / height if height > 0 else 0.0
    return width, height, aspect


def _bits_to_hex(bits: list[int]) -> str:
    value = 0
    for bit in bits:
        value = (value << 1) | (1 if bit else 0)
    return f"{value:0{(len(bits) + 3) // 4}x}" if bits else ""


def _hex_similarity(first: str | None, second: str | None) -> float | None:
    if not first or not second:
        return None
    try:
        first_bits = bin(int(first, 16))[2:].zfill(len(first) * 4)
        second_bits = bin(int(second, 16))[2:].zfill(len(second) * 4)
    except ValueError:
        return None
    length = min(len(first_bits), len(second_bits))
    if length <= 0:
        return None
    distance = sum(1 for index in range(length) if first_bits[index] != second_bits[index])
    return 1.0 - (distance / length)


def _histogram_similarity(first: list[float] | None, second: list[float] | None) -> float | None:
    if not first or not second:
        return None
    length = min(len(first), len(second))
    dot = sum(_coerce_float(first[index]) * _coerce_float(second[index]) for index in range(length))
    first_norm = math.sqrt(sum(_coerce_float(value) ** 2 for value in first[:length]))
    second_norm = math.sqrt(sum(_coerce_float(value) ** 2 for value in second[:length]))
    return dot / (first_norm * second_norm) if first_norm and second_norm else None


def appearance_fingerprint_from_bytes(image_bytes: bytes | None, *, bbox: list[float] | None = None, padding_ratio: float = 0.0, strip_top_ratio: float = 0.0, max_bytes: int = 2 * 1024 * 1024) -> dict | None:
    if not image_bytes or len(image_bytes) > max_bytes:
        return None
    try:
        image = Image.open(BytesIO(image_bytes)).convert("RGB")
    except Exception:
        return None
    width, height = image.size
    used_object_crop = False
    if bbox and len(bbox) >= 4:
        x1, y1, x2, y2 = [_coerce_float(value) for value in bbox[:4]]
        if max(x1, y1, x2, y2) <= 1:
            x1, x2 = x1 * width, x2 * width
            y1, y2 = y1 * height, y2 * height
        pad_x = max(0.0, x2 - x1) * max(0.0, padding_ratio)
        pad_y = max(0.0, y2 - y1) * max(0.0, padding_ratio)
        left, top = max(0, int(x1 - pad_x)), max(0, int(y1 - pad_y))
        right, bottom = min(width, int(x2 + pad_x)), min(height, int(y2 + pad_y))
        if right > left and bottom > top:
            image = image.crop((left, top, right, bottom))
            width, height = image.size
            used_object_crop = True
    if strip_top_ratio > 0 and height > 12:
        image = image.crop((0, min(height - 1, int(height * strip_top_ratio)), width, height))
    image = ImageOps.autocontrast(image.resize((64, 64)))
    gray = image.convert("L")
    small = list(gray.resize((8, 8)).getdata())
    mean = sum(small) / len(small)
    average_hash = _bits_to_hex([1 if value >= mean else 0 for value in small])
    dhash_pixels = list(gray.resize((9, 8)).getdata())
    edge_bits = []
    for row in range(8):
        offset = row * 9
        edge_bits.extend(1 if dhash_pixels[offset + col] > dhash_pixels[offset + col + 1] else 0 for col in range(8))
    histogram: list[float] = []
    for channel in image.split():
        bins = [0] * 8
        for value in channel.getdata():
            bins[min(7, int(value) // 32)] += 1
        total = sum(bins) or 1
        histogram.extend(round(item / total, 6) for item in bins)
    return {
        "average_hash": average_hash,
        "edge_hash": _bits_to_hex(edge_bits),
        "color_histogram": histogram,
        "source": "object_crop" if used_object_crop else "crop_bytes",
    }


def appearance_similarity(first: dict | None, second: dict | None) -> dict:
    if not first or not second:
        return {"status": "appearance unavailable", "score": None}
    average_score = _hex_similarity(first.get("average_hash"), second.get("average_hash"))
    edge_score = _hex_similarity(first.get("edge_hash"), second.get("edge_hash"))
    color_score = _histogram_similarity(first.get("color_histogram"), second.get("color_histogram"))
    scores = [value for value in (average_score, edge_score, color_score) if value is not None]
    if not scores:
        return {"status": "hash missing", "score": None}
    fallback = sum(scores) / len(scores)
    score = (0.35 * (average_score if average_score is not None else fallback)) + (0.35 * (edge_score if edge_score is not None else fallback)) + (0.30 * (color_score if color_score is not None else fallback))
    return {
        "status": "appearance compared successfully",
        "score": round(score, 4),
        "average_hash_similarity": round(average_score, 4) if average_score is not None else None,
        "edge_similarity": round(edge_score, 4) if edge_score is not None else None,
        "color_histogram_similarity": round(color_score, 4) if color_score is not None else None,
    }


def _median(values: list[float]) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    middle = len(ordered) // 2
    return ordered[middle] if len(ordered) % 2 else (ordered[middle - 1] + ordered[middle]) / 2


def _fingerprint_quality(item: dict) -> float:
    source = str(item.get("source") or "").lower()
    if "full_frame" in source or "context" in source:
        return 0.0
    return max(0.0, min(1.0, _coerce_float(item.get("quality_score"), 1.0)))


def _appearance_evidence(first: dict, second: dict) -> dict:
    first_items = (first.get("track_debug") or {}).get("appearance_fingerprints") or []
    second_items = (second.get("track_debug") or {}).get("appearance_fingerprints") or []
    comparisons = []
    for first_item in first_items:
        for second_item in second_items:
            pair_quality = min(_fingerprint_quality(first_item), _fingerprint_quality(second_item))
            if pair_quality < 0.5:
                continue
            current = appearance_similarity(first_item, second_item)
            if current.get("score") is None:
                continue
            comparisons.append({
                **current,
                "quality": round(pair_quality, 4),
                "frames": [first_item.get("frame"), second_item.get("frame")],
            })
    if not comparisons:
        return {
            "status": "appearance unavailable",
            "score": None,
            "best_valid_similarity": None,
            "strongest_median_similarity": None,
            "valid_pair_count": 0,
            "agreeing_pair_count": 0,
            "conflicting_pair_count": 0,
        }
    comparisons.sort(key=lambda item: (_coerce_float(item.get("score")), _coerce_float(item.get("quality"))), reverse=True)
    scores = [_coerce_float(item.get("score")) for item in comparisons]
    strongest = scores[: min(3, len(scores))]
    best = comparisons[0]
    return {
        **best,
        "status": "appearance compared successfully",
        "score": round(_median(strongest) or 0.0, 4),
        "best_valid_similarity": round(scores[0], 4),
        "strongest_median_similarity": round(_median(strongest) or 0.0, 4),
        "valid_pair_count": len(comparisons),
        "agreeing_pair_count": sum(score >= VIDEO_DUPLICATE_APPEARANCE_SIMILARITY for score in scores),
        "conflicting_pair_count": sum(score < VIDEO_DUPLICATE_MIN_APPEARANCE_SIMILARITY for score in scores),
    }


def _best_appearance_similarity(first: dict, second: dict) -> dict:
    return _appearance_evidence(first, second)


def _bounded_representative_fingerprints(items: list[dict], limit: int = 5) -> list[dict]:
    by_frame = {
        int(_coerce_float(item.get("frame"))): item
        for item in items
        if isinstance(item, dict) and item.get("frame") is not None
    }
    ordered = [by_frame[frame] for frame in sorted(by_frame)]
    if len(ordered) <= limit:
        return ordered
    candidates = [
        ordered[0],
        ordered[len(ordered) // 2],
        ordered[-1],
        max(ordered, key=lambda item: (_coerce_float(item.get("confidence")), -int(_coerce_float(item.get("frame"))))),
        max(ordered, key=lambda item: (_fingerprint_quality(item), _coerce_float(item.get("confidence")), -int(_coerce_float(item.get("frame"))))),
    ]
    selected = {int(_coerce_float(item.get("frame"))): item for item in candidates}
    if len(selected) < limit:
        for item in sorted(ordered, key=lambda value: (_fingerprint_quality(value), _coerce_float(value.get("confidence"))), reverse=True):
            selected.setdefault(int(_coerce_float(item.get("frame"))), item)
            if len(selected) >= limit:
                break
    return [selected[frame] for frame in sorted(selected)][:limit]


def _bbox_quality(box: list[float] | None) -> dict:
    if not box or len(box) < 4:
        return {"score": 0.0, "valid": False, "reason": "bbox missing"}
    x1, y1, x2, y2 = [_coerce_float(value) for value in box[:4]]
    width = max(0.0, x2 - x1)
    height = max(0.0, y2 - y1)
    area = width * height
    touches_edge = (
        x1 <= VIDEO_DUPLICATE_TRUNCATED_EDGE_MARGIN
        or y1 <= VIDEO_DUPLICATE_TRUNCATED_EDGE_MARGIN
        or x2 >= 1.0 - VIDEO_DUPLICATE_TRUNCATED_EDGE_MARGIN
        or y2 >= 1.0 - VIDEO_DUPLICATE_TRUNCATED_EDGE_MARGIN
    )
    if width <= 0 or height <= 0:
        return {"score": 0.0, "valid": False, "reason": "bbox invalid"}
    if area >= VIDEO_DUPLICATE_FULL_FRAME_AREA:
        return {"score": 0.05, "valid": False, "reason": "bbox near full frame", "area": round(area, 4)}
    score = 1.0
    if touches_edge:
        score -= 0.35
    if area < 0.005:
        score -= 0.25
    return {
        "score": round(max(0.0, score), 4),
        "valid": score >= 0.5,
        "reason": "valid" if score >= 0.5 else "low quality bbox",
        "area": round(area, 4),
        "touches_edge": touches_edge,
    }


class VideoCameraMotionState:
    """Small deterministic global-translation estimator for future MP4 reconciliation."""

    def __init__(self, width: int, height: int):
        self.width = max(1, int(width))
        self.height = max(1, int(height))
        self.previous_gray = None
        self.dx = 0.0
        self.dy = 0.0
        self.valid_frames = 0
        self.failed_frames = 0
        self.latest_status = "not_started"

    def observe(self, frame: Any, frame_index: int) -> dict:
        try:
            import cv2
            import numpy as np
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            small = cv2.resize(gray, (232, 416)).astype(np.float32)
            if self.previous_gray is None:
                self.previous_gray = small
                self.latest_status = "reference"
                self.valid_frames += 1
                return self.metadata(frame_index, response=None)
            (step_x, step_y), response = cv2.phaseCorrelate(small, self.previous_gray)
            self.previous_gray = small
            if response < VIDEO_CAMERA_MOTION_MIN_RESPONSE:
                self.failed_frames += 1
                self.latest_status = "weak_phase_correlation"
                return self.metadata(frame_index, response=response)
            self.dx += float(step_x) * (self.width / 232)
            self.dy += float(step_y) * (self.height / 416)
            self.valid_frames += 1
            self.latest_status = "ok"
            return self.metadata(frame_index, response=response)
        except Exception:
            self.failed_frames += 1
            self.latest_status = "motion_estimate_failed"
            return self.metadata(frame_index, response=None)

    def metadata(self, frame_index: int, *, response: float | None) -> dict:
        return {
            "frame": frame_index,
            "status": self.latest_status,
            "valid": self.latest_status in {"reference", "ok"},
            "dx": round(self.dx, 4),
            "dy": round(self.dy, 4),
            "response": round(response, 4) if response is not None else None,
            "valid_frames": self.valid_frames,
            "failed_frames": self.failed_frames,
        }

    def scene_center_for_bbox(self, box: list[float] | None) -> dict | None:
        if self.latest_status not in {"reference", "ok"} or not box:
            return None
        cx, cy = _bbox_center(box)
        return {
            "x": round((cx * self.width + self.dx) / self.width, 6),
            "y": round((cy * self.height + self.dy) / self.height, 6),
        }

    def summary(self) -> dict:
        return {
            "status": "ok" if self.valid_frames > 1 else self.latest_status,
            "valid_frames": self.valid_frames,
            "failed_frames": self.failed_frames,
            "dx": round(self.dx, 4),
            "dy": round(self.dy, 4),
            "min_response": VIDEO_CAMERA_MOTION_MIN_RESPONSE,
        }


def attach_camera_motion_to_detections(detections: list[dict], motion: VideoCameraMotionState, motion_meta: dict) -> list[dict]:
    output = []
    for detection in detections:
        enriched = dict(detection)
        scene_center = motion.scene_center_for_bbox(enriched.get("bbox"))
        if scene_center:
            enriched["scene_center"] = scene_center
            enriched["camera_motion_status"] = motion_meta.get("status")
            enriched["camera_motion_response"] = motion_meta.get("response")
            enriched["camera_motion_reliable"] = bool(motion_meta.get("valid"))
        output.append(enriched)
    return output


@dataclass
class VideoTrackState:
    key: str
    raw_track_ids: set[str] = field(default_factory=set)
    first_frame: int = 0
    last_frame: int = 0
    first_timestamp: float = 0.0
    last_timestamp: float = 0.0
    class_votes: dict[str, float] = field(default_factory=dict)
    class_names: dict[str, str] = field(default_factory=dict)
    confidences: list[float] = field(default_factory=list)
    observations: list[dict] = field(default_factory=list)
    path: list[dict] = field(default_factory=list)
    best_observation: dict = field(default_factory=dict)
    best_frame_bytes: bytes | None = None
    best_frame_dimensions: dict | None = None
    best_crop_bytes: bytes | None = None
    appearance_fingerprints: list[dict] = field(default_factory=list)
    counted: bool = False
    last_center: tuple[float, float] | None = None
    scene_path: list[dict] = field(default_factory=list)


class VideoTrackAggregator:
    """Collect sequential ByteTrack observations into one row per physical object."""

    def __init__(
        self,
        upload_id: str,
        *,
        min_frames: int = VIDEO_TRACK_MIN_FRAMES,
        short_track_confidence: float = VIDEO_TRACK_SHORT_CONFIDENCE,
        lost_buffer: int = VIDEO_TRACK_LOST_BUFFER,
        recovery_iou: float = VIDEO_TRACK_RECOVERY_IOU,
        recovery_center_distance: float = VIDEO_TRACK_RECOVERY_CENTER_DISTANCE,
        counting_line: dict | None = None,
    ):
        self.upload_id = str(upload_id)
        self.min_frames = min_frames
        self.short_track_confidence = short_track_confidence
        self.lost_buffer = lost_buffer
        self.recovery_iou = recovery_iou
        self.recovery_center_distance = recovery_center_distance
        self.counting_line = counting_line
        self.active: dict[str, VideoTrackState] = {}
        self.raw_to_key: dict[str, str] = {}
        self.finalized_track_ids: set[str] = set()
        self.finalized: list[dict] = []
        self._next_synthetic = 1

    def observe(self, frame_index: int, timestamp: float, detections: list[dict]) -> list[dict]:
        for detection in detections:
            self._observe_one(frame_index, timestamp, detection)
        return self.flush_stale(frame_index)

    def flush_stale(self, frame_index: int, *, force: bool = False) -> list[dict]:
        flushed = []
        for key, state in list(self.active.items()):
            if force or frame_index - state.last_frame > self.lost_buffer:
                self.active.pop(key, None)
                for raw_id in state.raw_track_ids:
                    if self.raw_to_key.get(raw_id) == key:
                        self.raw_to_key.pop(raw_id, None)
                material = self._finalize(state)
                if material:
                    self.finalized_track_ids.update(state.raw_track_ids)
                    self.finalized.append(material)
                    flushed.append(material)
                    _video_debug(
                        "track_finalized",
                        scan_id=self.upload_id,
                        logical_object_id=material.get("stable_object_id"),
                        source_track_ids=material.get("source_track_ids"),
                        first_frame=material.get("track_first_frame"),
                        last_frame=material.get("track_last_frame"),
                        observations=material.get("track_frame_count"),
                        final_class=material.get("category"),
                    )
        return flushed

    def finish(self, frame_index: int = 0) -> list[dict]:
        return self.flush_stale(frame_index, force=True)

    def _observe_one(self, frame_index: int, timestamp: float, detection: dict) -> None:
        track_id = detection.get("track_id")
        raw_id = str(track_id) if track_id is not None and str(track_id) != "" else ""
        if raw_id and raw_id in self.finalized_track_ids:
            _video_debug("track_observation_skipped_finalized", scan_id=self.upload_id, frame=frame_index, raw_track_id=raw_id)
            return
        category = material_category(detection.get("category") or detection.get("material_name"))
        box = [round(_coerce_float(value), 6) for value in detection.get("bbox") or []][:4]
        key = self._resolve_key(raw_id, category, box, frame_index)
        state = self.active.get(key)
        confidence = max(0.0, min(1.0, _coerce_float(detection.get("confidence"))))
        if not state:
            state = VideoTrackState(
                key=key,
                first_frame=frame_index,
                last_frame=frame_index,
                first_timestamp=timestamp,
                last_timestamp=timestamp,
            )
            self.active[key] = state
        if raw_id:
            state.raw_track_ids.add(raw_id)
            self.raw_to_key[raw_id] = key
        state.last_frame = frame_index
        state.last_timestamp = timestamp
        state.class_votes[category] = state.class_votes.get(category, 0.0) + confidence
        state.class_names.setdefault(category, str(detection.get("material_name") or category))
        state.confidences.append(confidence)
        center = _bbox_center(box)
        state.path.append({"frame": frame_index, "timestamp": round(timestamp, 3), "x": round(center[0], 4), "y": round(center[1], 4)})
        scene_center = detection.get("scene_center") if isinstance(detection.get("scene_center"), dict) else None
        if scene_center and scene_center.get("x") is not None and scene_center.get("y") is not None:
            state.scene_path.append({
                "frame": frame_index,
                "timestamp": round(timestamp, 3),
                "x": round(_coerce_float(scene_center.get("x")), 6),
                "y": round(_coerce_float(scene_center.get("y")), 6),
                "motion_status": detection.get("camera_motion_status"),
                "motion_response": detection.get("camera_motion_response"),
            })
        self._update_counting_line(state, center)
        observation = {
            "frame": frame_index,
            "timestamp": round(timestamp, 3),
            "track_id": raw_id or key,
            "category": category,
            "confidence": round(confidence, 4),
            "bbox": box,
            "bbox_percent": detection.get("bbox_percent"),
        }
        if scene_center:
            observation["scene_center"] = {
                "x": round(_coerce_float(scene_center.get("x")), 6),
                "y": round(_coerce_float(scene_center.get("y")), 6),
            }
            observation["camera_motion_status"] = detection.get("camera_motion_status")
            observation["camera_motion_response"] = detection.get("camera_motion_response")
            observation["camera_motion_reliable"] = bool(detection.get("camera_motion_reliable"))
        _video_debug(
            "track_observation",
            scan_id=self.upload_id,
            frame=frame_index,
            raw_track_id=raw_id or key,
            class_name=category,
            confidence=round(confidence, 4),
            bbox=box,
        )
        state.observations.append(observation)
        fingerprint = appearance_fingerprint_from_bytes(detection.get("crop_bytes"), strip_top_ratio=0.0)
        if fingerprint:
            bbox_quality = _bbox_quality(box)
            fingerprint = {
                **fingerprint,
                "frame": frame_index,
                "confidence": round(confidence, 4),
                "quality_score": bbox_quality.get("score"),
                "bbox_quality": bbox_quality.get("reason"),
            }
            state.appearance_fingerprints.append(fingerprint)
            state.appearance_fingerprints = _bounded_representative_fingerprints(state.appearance_fingerprints)
        if confidence >= _coerce_float(state.best_observation.get("confidence"), -1):
            state.best_observation = {
                **observation,
                "mask": detection.get("mask"),
                "best_box": detection.get("best_box") or detection.get("bbox_percent") or box,
            }
            state.best_frame_bytes = detection.get("frame_bytes") or state.best_frame_bytes
            if detection.get("frame_width") and detection.get("frame_height"):
                state.best_frame_dimensions = {
                    "width": int(detection["frame_width"]),
                    "height": int(detection["frame_height"]),
                }
            state.best_crop_bytes = detection.get("crop_bytes") or state.best_crop_bytes
        state.last_center = center

    def _resolve_key(self, raw_id: str, category: str, box: list[float], frame_index: int) -> str:
        if raw_id and raw_id in self.raw_to_key and self.raw_to_key[raw_id] in self.active:
            return self.raw_to_key[raw_id]
        recovered = self._recover_key(category, box, frame_index)
        if recovered:
            return recovered
        if raw_id:
            return raw_id
        key = f"synthetic-{self._next_synthetic}"
        self._next_synthetic += 1
        return key

    def _recover_key(self, category: str, box: list[float], frame_index: int) -> str | None:
        best_key = None
        best_score = 0.0
        cx, cy = _bbox_center(box)
        for key, state in self.active.items():
            if frame_index - state.last_frame > self.lost_buffer:
                continue
            previous_category = max(state.class_votes, key=state.class_votes.get, default="")
            if previous_category != category:
                continue
            previous_box = state.best_observation.get("bbox") or []
            iou = _bbox_iou(previous_box, box)
            px, py = state.last_center or _bbox_center(previous_box)
            distance = math.hypot(cx - px, cy - py)
            score = max(iou, 1.0 - distance)
            if (iou >= self.recovery_iou or distance <= self.recovery_center_distance) and score > best_score:
                best_score = score
                best_key = key
        return best_key

    def _update_counting_line(self, state: VideoTrackState, center: tuple[float, float]) -> None:
        if not self.counting_line:
            return
        previous = state.last_center
        if previous is None or state.counted:
            return
        index = 0 if self.counting_line["axis"] == "x" else 1
        position = self.counting_line["position"]
        before = previous[index] - position
        after = center[index] - position
        crossed = before == 0 or after == 0 or before * after < 0
        if not crossed:
            return
        direction = self.counting_line["direction"]
        if direction == "positive" and after <= before:
            return
        if direction == "negative" and after >= before:
            return
        state.counted = True

    def _finalize(self, state: VideoTrackState) -> dict | None:
        frame_count = len(state.observations)
        max_confidence = max(state.confidences or [0.0])
        if frame_count < self.min_frames and max_confidence < self.short_track_confidence:
            return None
        final_category = max(state.class_votes, key=state.class_votes.get, default="unknown")
        avg_confidence = sum(state.confidences) / frame_count if frame_count else 0.0
        recyclable_status, contaminant_status = material_status(final_category)
        hazard_status = "hazard" if CATEGORY_CLASS_MAP.get(final_category) == "contaminant" else "clear"
        best_box = state.best_observation.get("bbox_percent") or {}
        if not isinstance(best_box, dict):
            best_box = {}
        best_norm_box = state.best_observation.get("bbox") or []
        start_center = state.path[0] if state.path else {"x": 0, "y": 0}
        end_center = state.path[-1] if state.path else {"x": 0, "y": 0}
        start_scene_center = state.scene_path[0] if state.scene_path else None
        end_scene_center = state.scene_path[-1] if state.scene_path else None
        widths, heights, aspects = [], [], []
        for observation in state.observations:
            width, height, aspect = _bbox_size(observation.get("bbox"))
            widths.append(width)
            heights.append(height)
            aspects.append(aspect)
        source_track_ids = sorted(state.raw_track_ids)
        stable_suffix = re.sub(r"[^A-Za-z0-9_.-]+", "-", state.key).strip("-") or "object"
        material = {
            "stable_object_id": f"{self.upload_id}-track-{stable_suffix}",
            "object_uid": f"{self.upload_id}-track-{stable_suffix}",
            "source_track_ids": source_track_ids,
            "track_id": ",".join(sorted(state.raw_track_ids)) or stable_suffix,
            "material_name": state.class_names.get(final_category, final_category),
            "category": final_category,
            "confidence": round(max_confidence, 4),
            "track_avg_confidence": round(avg_confidence, 4),
            "track_max_confidence": round(max_confidence, 4),
            "track_first_frame": state.first_frame,
            "track_last_frame": state.last_frame,
            "track_first_timestamp": round(state.first_timestamp, 3),
            "track_last_timestamp": round(state.last_timestamp, 3),
            "track_duration_seconds": round(max(0.0, state.last_timestamp - state.first_timestamp), 3),
            "track_frame_count": frame_count,
            "track_hazard_status": hazard_status,
            "track_counted": True if not self.counting_line else state.counted,
            "track_start_center": {"x": start_center.get("x", 0), "y": start_center.get("y", 0)},
            "track_end_center": {"x": end_center.get("x", 0), "y": end_center.get("y", 0)},
            **({"track_start_scene_center": {"x": start_scene_center.get("x"), "y": start_scene_center.get("y")}} if start_scene_center else {}),
            **({"track_end_scene_center": {"x": end_scene_center.get("x"), "y": end_scene_center.get("y")}} if end_scene_center else {}),
            "track_avg_width": round(sum(widths) / len(widths), 6) if widths else 0,
            "track_avg_height": round(sum(heights) / len(heights), 6) if heights else 0,
            "track_avg_aspect_ratio": round(sum(aspects) / len(aspects), 6) if aspects else 0,
            "recyclable_status": recyclable_status,
            "contaminant_status": contaminant_status,
            **evaluate_material(final_category, max_confidence),
            "bbox_x": round(_coerce_float(best_box.get("x")), 2),
            "bbox_y": round(_coerce_float(best_box.get("y")), 2),
            "bbox_width": round(_coerce_float(best_box.get("width")), 2),
            "bbox_height": round(_coerce_float(best_box.get("height")), 2),
            "best_box": state.best_observation.get("best_box"),
            "best_bbox_norm": best_norm_box,
            "segmentation_mask": state.best_observation.get("mask"),
            "track_path": state.path,
            "track_debug": {
                "frame_observations": state.observations,
                "accepted_track_fragments": _track_frame_fragments(state.observations),
                "class_votes": {key: round(value, 4) for key, value in state.class_votes.items()},
                "raw_track_ids": sorted(state.raw_track_ids),
                "appearance_fingerprints": state.appearance_fingerprints,
                "stabilized_track_path": state.scene_path,
                "representative_frame_dimensions": state.best_frame_dimensions,
                "representative_bbox_format": "normalized_original_frame_xyxy",
            },
            "_best_crop_bytes": state.best_frame_bytes or state.best_crop_bytes,
        }
        return material


def canonical_category_key(value: str | None) -> str:
    key = " ".join(str(value or "").strip().lower().replace("_", " ").replace("-", " ").split())
    if "food" in key or "organic" in key:
        return "food_organics"
    if "general" in key or "trash" in key or "waste" in key:
        return "general_trash"
    if "textile" in key or "fabric" in key or "cloth" in key:
        return "textile"
    if "battery" in key:
        return "battery"
    if "cardboard" in key or "box" in key:
        return "cardboard"
    if "glass" in key or "jar" in key:
        return "glass"
    if "paper" in key:
        return "paper"
    if "metal" in key or "aluminum" in key or "aluminium" in key or "can" in key:
        return "metal"
    if "plastic" in key or "bottle" in key or "pet" in key or "film" in key:
        return "plastic"
    return "unknown"


def filter_scans_by_final_category(scans: list[dict], materials: list[dict], decisions: list[dict], category: str) -> list[dict]:
    latest = {}
    for decision in sorted(decisions, key=lambda item: str(item.get("created_at", ""))):
        latest[str(decision.get("detected_material_id", ""))] = decision
    materials_by_scan = {}
    for material in materials:
        materials_by_scan.setdefault(str(material.get("scan_result_id", "")), []).append(material)

    def final_category(scan: dict, material: dict) -> str:
        decision = latest.get(str(material.get("id", "")), {})
        return canonical_category_key(decision.get("chosen_category") or scan.get("verified_category") or material.get("category") or material.get("material_name"))

    return [
        scan for scan in scans
        if any(final_category(scan, material) == category for material in materials_by_scan.get(str(scan.get("id", "")), []))
    ]


def filter_materials_by_final_category(materials: list[dict], scans: list[dict], decisions: list[dict], category: str) -> list[dict]:
    latest = latest_decisions_by_material(decisions)
    scan_by_id = {str(scan.get("id")): scan for scan in scans if scan.get("id")}
    return [
        material for material in materials
        if canonical_category_key(
            (latest.get(str(material.get("id") or "")) or {}).get("chosen_category")
            or (scan_by_id.get(str(material.get("scan_result_id") or "")) or {}).get("verified_category")
            or material.get("category")
            or material.get("material_name")
        ) == category
    ]


@dataclass(frozen=True)
class Principal:
    kind: str
    id: str
    scopes: frozenset[str]
    email: str | None = None
    role: str = "authenticated"
    claims: dict[str, Any] = field(default_factory=dict)
    profile_id: str | None = None


AUTHENTICATED_SCOPES = frozenset({"scan:read", "scan:write", "job:read", "review:write"})
bearer_scheme = HTTPBearer(auto_error=False)


def _auth_401(detail: str = "Authentication required.") -> HTTPException:
    return HTTPException(status_code=401, detail=detail, headers={"WWW-Authenticate": "Bearer"})


def _configured_secret_tokens() -> list[str]:
    return [value for value in (SUPABASE_SERVICE_ROLE_KEY, os.getenv("SUPABASE_ANON_KEY")) if value]


def _looks_like_configured_secret(token: str) -> bool:
    return any(hmac.compare_digest(token, secret) for secret in _configured_secret_tokens())


def is_production() -> bool:
    return os.getenv("ENVIRONMENT", "development").strip().lower() == "production"


def _get_attr(value: Any, name: str, default: Any = None) -> Any:
    if isinstance(value, dict):
        return value.get(name, default)
    return getattr(value, name, default)


def _extract_user(auth_response: Any) -> Any:
    return _get_attr(auth_response, "user") or _get_attr(_get_attr(auth_response, "data"), "user")


def _lookup_user_profile_id(client: Any, auth_user_id: str) -> str | None:
    try:
        response = client.table("user_profiles").select("id").eq("auth_user_id", auth_user_id).maybe_single().execute()
    except Exception:
        return None
    profile_id = _get_attr(_get_attr(response, "data"), "id")
    return str(profile_id) if profile_id else None


def scan_user_id(principal: Principal | None, database: SupabaseExecutor | None = None) -> str | None:
    if not principal or principal.kind != "user":
        return None
    if principal.profile_id:
        return principal.profile_id
    if not database or not database.client:
        return None
    return _lookup_user_profile_id(database.client, principal.id)


def verify_supabase_token(token: str) -> Principal:
    token = str(token or "").strip()
    if not token or _looks_like_configured_secret(token):
        raise _auth_401("Invalid authentication token.")

    client = _new_supabase_client()
    if not client:
        raise HTTPException(status_code=500, detail="Authentication is not configured.")

    try:
        auth_response = client.auth.get_user(token)
    except Exception:
        raise _auth_401("Invalid authentication token.") from None

    user = _extract_user(auth_response)
    user_id = str(_get_attr(user, "id", "") or "")
    email = _get_attr(user, "email")
    is_anonymous = bool(_get_attr(user, "is_anonymous", False))
    if not user_id or is_anonymous:
        raise _auth_401("Invalid authentication token.")

    claims = {"sub": user_id, "email": email, "role": "authenticated"}
    profile_id = _lookup_user_profile_id(client, user_id)
    return Principal("user", user_id, AUTHENTICATED_SCOPES, email=email, role="authenticated", claims=claims, profile_id=profile_id)


def require_principal(credentials: HTTPAuthorizationCredentials | None = Depends(bearer_scheme)) -> Principal:
    if not credentials or credentials.scheme.lower() != "bearer":
        raise _auth_401()
    return verify_supabase_token(credentials.credentials)


def _api_key_digest(value: str) -> str:
    return hashlib.sha256(str(value or "").encode("utf-8")).hexdigest()


def require_scope(scope: str):
    def dependency(principal: Principal = Depends(require_principal)) -> Principal:
        if scope not in principal.scopes:
            raise HTTPException(status_code=403, detail=f"Missing API scope: {scope}")
        return principal
    return dependency


def scoped_query(query, principal: Principal):
    # Capstone data is shared; current schema has no per-user ownership filter.
    return query


def _sanitize_device_name(value: Any) -> str | None:
    if value is None:
        return None
    return re.sub(r"[^A-Za-z0-9 ._:+-]+", "", str(value)).strip()[:120] or None


def _cuda_device_index(device: str) -> int:
    parts = device.split(":", 1)
    if parts[0] != "cuda" or len(parts) > 2:
        raise RuntimeError(f"Unsupported MODEL_DEVICE: {device}")
    if len(parts) == 1 or parts[1] == "":
        return 0
    try:
        index = int(parts[1])
    except ValueError as exc:
        raise RuntimeError(f"Invalid CUDA MODEL_DEVICE index: {device}") from exc
    if index < 0:
        raise RuntimeError(f"Invalid CUDA MODEL_DEVICE index: {device}")
    return index


def _torch_cuda_diagnostics() -> tuple[Any | None, dict]:
    try:
        import torch
    except Exception as exc:
        return None, {
            "model_device_requested": MODEL_DEVICE,
            "cuda_available": False,
            "cuda_device_count": 0,
            "cuda_device_name": None,
            "torch_import_error": type(exc).__name__,
        }

    cuda = getattr(torch, "cuda", None)
    cuda_available = bool(cuda and cuda.is_available())
    cuda_device_count = int(cuda.device_count()) if cuda and cuda_available else 0
    device_name = None
    if cuda_available and cuda_device_count > 0:
        try:
            device_name = _sanitize_device_name(cuda.get_device_name(0))
        except Exception:
            device_name = None
    return torch, {
        "model_device_requested": MODEL_DEVICE,
        "cuda_available": cuda_available,
        "cuda_device_count": cuda_device_count,
        "cuda_device_name": device_name,
    }


def _select_model_device() -> tuple[str, dict]:
    torch, diagnostics = _torch_cuda_diagnostics()
    requested = MODEL_DEVICE
    if not requested.startswith("cuda"):
        return requested, diagnostics
    if torch is None:
        raise RuntimeError(f"MODEL_DEVICE={requested} requires PyTorch.")
    if not diagnostics["cuda_available"]:
        raise RuntimeError(f"MODEL_DEVICE={requested} requires CUDA, but torch.cuda.is_available() is false.")
    if diagnostics["cuda_device_count"] < 1:
        raise RuntimeError(f"MODEL_DEVICE={requested} requires at least one CUDA device.")
    index = _cuda_device_index(requested)
    if index >= diagnostics["cuda_device_count"]:
        raise RuntimeError(f"MODEL_DEVICE={requested} requested CUDA device {index}, but only {diagnostics['cuda_device_count']} CUDA device(s) exist.")
    selected = f"cuda:{index}"
    try:
        diagnostics["cuda_device_name"] = _sanitize_device_name(torch.cuda.get_device_name(index))
    except Exception:
        diagnostics["cuda_device_name"] = None
    return selected, diagnostics


def _model_parameter_device(loaded_model: Any) -> str | None:
    try:
        parameters = loaded_model.model.parameters()
        return str(next(parameters).device)
    except Exception:
        return None


def get_model():
    global model, model_device_effective
    if model is None:
        print(f"[startup] Loading YOLO model from: {MODEL_PATH}")
        if not MODEL_PATH.exists():
            print(f"[startup] YOLO model file not found at: {MODEL_PATH}")
            raise HTTPException(status_code=500, detail="YOLO model file not found.")
        selected_device, diagnostics = _select_model_device()
        print(f"[startup] model device {json.dumps({**diagnostics, 'selected_model_device': selected_device}, sort_keys=True)}")
        loaded_model = YOLO(str(MODEL_PATH))
        model_to = getattr(loaded_model, "to", None)
        moved_model = model_to(selected_device) if callable(model_to) else loaded_model
        model = moved_model if moved_model is not None else loaded_model
        model_device_effective = _model_parameter_device(model) or selected_device
        if MODEL_DEVICE.startswith("cuda") and not str(model_device_effective).startswith("cuda"):
            raise RuntimeError(f"MODEL_DEVICE={MODEL_DEVICE} requested CUDA, but model loaded on {model_device_effective}.")
    return model


def _yolo_model_for_weights(model_weights: str | os.PathLike[str]):
    weights_path = Path(model_weights)
    if str(model_weights) == "best.pt" or weights_path == MODEL_PATH:
        return get_model()
    return YOLO(str(model_weights))


def predict_with_calibration(image_input, model_weights="best.pt"):
    model_instance = _yolo_model_for_weights(model_weights)
    results = model_instance.predict(image_input, conf=YOLO_CALIBRATION_CANDIDATE_CONFIDENCE, verbose=False)

    for result in results:
        if result.boxes is None or len(result.boxes) <= 0:
            continue
        filtered_indices = []
        for idx, (cls_tensor, conf_tensor) in enumerate(zip(result.boxes.cls, result.boxes.conf)):
            class_id = int(cls_tensor.item())
            confidence = float(conf_tensor.item())
            if confidence >= CLASS_THRESHOLDS.get(class_id, 0.25):
                filtered_indices.append(idx)

        result.boxes = result.boxes[filtered_indices]
        if result.masks is not None:
            result.masks.data = result.masks.data[filtered_indices]

    return results


def safe_startup_diagnostics() -> dict:
    browser_model_available = BROWSER_MODEL_PATH.exists()
    _torch, gpu = _torch_cuda_diagnostics()
    return {
        **deployment_identity(),
        "model_path": str(BROWSER_MODEL_PATH),
        "model_name": BROWSER_MODEL_NAME,
        "model_engine": BROWSER_INFERENCE_ENGINE,
        "model_available": browser_model_available,
        "browser_model_path": str(BROWSER_MODEL_PATH),
        "browser_model_available": browser_model_available,
        "backend_pytorch_model_path": str(MODEL_PATH),
        "backend_pytorch_model_available": MODEL_PATH.exists(),
        "supabase_configured": bool(SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY),
        "storage_bucket": PREVIEW_BUCKET,
        "storage_private": os.getenv("SUPABASE_STORAGE_PRIVATE", "false").lower() == "true",
        "drive_configured": bool(GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID),
        "ffmpeg_available": bool(shutil.which("ffmpeg")),
        "ffprobe_available": bool(shutil.which("ffprobe")),
        "model_loaded": model is not None,
        "model_device_effective": model_device_effective,
        **gpu,
        "allowed_origins": ALLOWED_ORIGINS,
    }


def deployment_identity() -> dict:
    app_commit_sha = (
        os.getenv("APP_COMMIT_SHA")
        or os.getenv("GIT_COMMIT_SHA")
        or os.getenv("VERCEL_GIT_COMMIT_SHA")
        or "unknown"
    )
    return {
        "app_commit_sha": app_commit_sha,
        "cloud_run_service": os.getenv("K_SERVICE") or "",
        "cloud_run_revision": os.getenv("K_REVISION") or "",
        "service_mode": SERVICE_MODE,
        "image_tag": os.getenv("IMAGE_TAG") or "",
        "image_digest": os.getenv("IMAGE_DIGEST") or "",
    }


def safe_drive_filename(original_filename: str | None) -> str:
    name = Path(original_filename or "uploaded-image.jpg").name
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("._") or "uploaded-image.jpg"
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return f"purityloop_{timestamp}_{safe_name}"


def upload_original_to_supabase_storage(
    file_bytes: bytes,
    original_filename: str | None,
    content_type: str,
    database: SupabaseExecutor | None = None,
    *,
    object_path: str | None = None,
) -> dict:
    database = database or SupabaseExecutor(supabase)
    if not database.client:
        raise RuntimeError("Supabase backend env is not configured")

    path = object_path or safe_drive_filename(original_filename)
    try:
        database.execute(lambda client: client.storage.from_(PREVIEW_BUCKET).upload(
            path=path,
            file=file_bytes,
            file_options={"content-type": content_type, "upsert": "false"},
        ))
    except Exception as exc:
        duplicate = (
            getattr(exc, "status_code", None) == 409
            or getattr(exc, "status", None) == 409
            or "duplicate" in str(exc).lower()
            or "already exists" in str(exc).lower()
        )
        if not object_path or not duplicate:
            raise
        # Deterministic browser paths make an existing object a successful retry.
        print(f"[Supabase Storage] Reusing deterministic object: {path}")
    public_url = supabase_storage_url(database, path)
    if not public_url:
        raise RuntimeError("Supabase Storage public URL is empty")
    return {"path": path, "public_url": str(public_url or "")}


def upload_file_to_supabase_storage(
    file_path: str | Path,
    object_path: str,
    content_type: str,
    database: SupabaseExecutor | None = None,
) -> dict:
    database = database or SupabaseExecutor(supabase)
    if not database.client:
        raise RuntimeError("Supabase backend env is not configured")

    path = str(object_path).lstrip("/")
    source = Path(file_path)
    with source.open("rb") as file:
        database.execute(lambda client: client.storage.from_(PREVIEW_BUCKET).upload(
            path=path,
            file=file,
            file_options={"content-type": content_type, "upsert": "true"},
        ))
    public_url = database.execute(lambda client: client.storage.from_(PREVIEW_BUCKET).get_public_url(path))
    if isinstance(public_url, dict):
        public_url = public_url.get("publicURL") or public_url.get("publicUrl") or public_url.get("signedURL") or ""
    if not public_url:
        raise RuntimeError("Supabase Storage public URL is empty")
    return {"path": path, "public_url": str(public_url or "")}


def supabase_storage_url(database: SupabaseExecutor, object_path: str, *, expires_in: int = 60 * 60 * 24 * 7) -> str:
    bucket = database.client.storage.from_(PREVIEW_BUCKET)
    private_bucket = os.getenv("SUPABASE_STORAGE_PRIVATE", "false").lower() == "true"
    if private_bucket and hasattr(bucket, "create_signed_url"):
        signed = database.execute(lambda _client: bucket.create_signed_url(object_path, expires_in))
        if isinstance(signed, dict):
            return str(signed.get("signedURL") or signed.get("signedUrl") or signed.get("url") or "")
    public_url = database.execute(lambda _client: bucket.get_public_url(object_path))
    if isinstance(public_url, dict):
        public_url = public_url.get("publicURL") or public_url.get("publicUrl") or public_url.get("signedURL") or ""
    return str(public_url or "")


def safe_error_message(exc: Exception) -> str:
    message = str(exc).replace(os.getenv("SUPABASE_SERVICE_ROLE_KEY") or "\0", "[redacted]")
    message = re.sub(r"[\w./ -]*google-service-account\.json", "[google-service-account.json]", message)
    message = re.sub(r"[\w./ -]*google-oauth-client\.json", "[google-oauth-client.json]", message)
    message = re.sub(r"[\w./ -]*google-oauth-token\.json", "[google-oauth-token.json]", message)
    message = re.sub(r"[\w./ -]*google-oauth-state\.json", "[google-oauth-state.json]", message)
    return message[:300]


def safe_worker_error_message(exc: Exception) -> str:
    return re.sub(r"/tmp/purityloop/[A-Za-z0-9._/-]+", "[worker-temp-path]", safe_error_message(exc))


def log_scan_stage_failure(stage: str, exc: Exception) -> None:
    print(f"[scans] {stage} failed: {type(exc).__name__}: {safe_error_message(exc)}")
    traceback.print_exc()


def execute_scan_read(stage: str, operation: Callable[[Any], T]) -> T:
    try:
        return SupabaseExecutor(client_factory=_new_supabase_client, attempts=2).execute(operation)
    except Exception as exc:
        log_scan_stage_failure(stage, exc)
        raise


def scan_history_filters(
    start_date: str | None,
    end_date: str | None,
    search: str | None,
    status: str | None,
) -> dict[str, str | None]:
    normalized_status = str(status or "").lower()
    return {
        "start_date": start_date,
        "end_date": end_date,
        "search": search.strip() if search and search.strip() else None,
        "status": normalized_status or None,
    }


def apply_scan_history_filters(query, filters: dict[str, str | None], status_override: str | None = None):
    query = query.neq("source_type", "video_frame")
    if filters.get("start_date"):
        query = query.gte("created_at", filters["start_date"])
    if filters.get("end_date"):
        query = query.lt("created_at", filters["end_date"])
    if filters.get("search"):
        query = query.ilike("source_name", f"%{filters['search']}%")
    normalized_status = str(status_override if status_override is not None else filters.get("status") or "").lower()
    if normalized_status == "review_needed":
        query = query.eq("human_review_required", True)
    elif normalized_status == "rejected":
        query = query.in_("overall_status", ["rejected", "quarantined"])
    elif normalized_status == "confirmed":
        query = query.eq("human_review_required", False)
    return query


def display_label(value: Any) -> str:
    text = re.sub(r"[_-]+", " ", str(value or "Unknown")).strip()
    return " ".join(word[:1].upper() + word[1:] for word in text.split()) or "Unknown"


def confidence_percent(value: Any) -> float:
    try:
        numeric = float(value or 0)
    except (TypeError, ValueError):
        return 0
    return numeric * 100 if numeric <= 1 else numeric


def latest_decisions_by_material(decisions: list[dict]) -> dict[str, dict]:
    latest: dict[str, dict] = {}
    for decision in sorted(decisions, key=lambda item: str(item.get("created_at", ""))):
        latest[str(decision.get("detected_material_id", ""))] = decision
    return latest


def attach_scan_children(scans: list[dict], materials: list[dict], decisions: list[dict]) -> list[dict]:
    latest = latest_decisions_by_material(decisions)
    materials_by_scan: dict[str, list[dict]] = {}
    for material in sorted(materials, key=lambda item: str(item.get("created_at", ""))):
        materials_by_scan.setdefault(str(material.get("scan_result_id", "")), []).append({
            **material,
            "review_decision": latest.get(str(material.get("id", ""))),
        })
    return [{**scan, "detected_materials": materials_by_scan.get(str(scan.get("id", "")), [])} for scan in scans]


def scan_history_page_from_rpc(rows: list[dict]) -> tuple[list[dict], int, dict[str, int]]:
    scans: list[dict] = []
    total = 0
    summary = {"confirmed": 0, "needs_review": 0, "rejected": 0, "total_objects": 0, "confirmed_objects": 0, "needs_review_objects": 0, "rejected_objects": 0}
    for row in rows:
        if row.get("total_count") is not None:
            total = int(row.get("total_count") or 0)
        for key in ("total_objects", "confirmed_objects", "needs_review_objects", "rejected_objects"):
            if row.get(key) is not None:
                summary[key] = int(row.get(key) or 0)
        scan = row.get("scan")
        if scan:
            scans.append(scan)
    summary["confirmed"] = summary["confirmed_objects"]
    summary["needs_review"] = summary["needs_review_objects"]
    summary["rejected"] = summary["rejected_objects"]
    return scans, total, summary


def export_material_decision(scan: dict) -> tuple[dict, dict | None]:
    material = (scan.get("detected_materials") or [{}])[0] or {}
    decision = material.get("review_decision") or None
    return material, decision


def export_final_category(scan: dict) -> str:
    material, decision = export_material_decision(scan)
    return canonical_category_key(scan.get("verified_category") or (decision or {}).get("chosen_category") or material.get("category") or material.get("material_name"))


def export_display_status(scan: dict, material: dict, decision: dict | None) -> tuple[str, str, str]:
    category = export_final_category(scan)
    material_class = (decision or {}).get("disposition") or material.get("material_class") or CATEGORY_CLASS_MAP.get(category, "unknown")
    final_status = derive_final_status(
        confidence=material.get("confidence") if material.get("confidence") is not None else scan.get("overall_confidence"),
        decision=decision,
        scan=scan,
    )
    if final_status == "rejected":
        return "Rejected", "rejected", material_class
    if normalize_status(scan.get("review_status") or scan.get("overall_status")) == "verified":
        return "Verified", "verified", material_class
    if final_status == "needs_review":
        return "Review Needed", "review_needed", material_class
    if material_class == "recyclable":
        return "Confirmed Recyclable", "confirmed", material_class
    if material_class == "contaminant":
        return "Confirmed Contaminant", "confirmed", material_class
    return "Confirmed", "confirmed", material_class


def export_weight(scan: dict, material: dict) -> str:
    for value in (
        material.get("estimated_weight_kg"), material.get("estimated_weight"), material.get("weight_kg"), material.get("weight"),
        scan.get("estimated_weight_kg"), scan.get("estimated_weight"), scan.get("weight_kg"), scan.get("weight"),
    ):
        if value not in (None, ""):
            try:
                return f"{float(value):.3f} kg"
            except (TypeError, ValueError):
                return str(value)
    category = material.get("category") or material.get("material_name")
    if category:
        key = analytics_category(category)
        estimate = ANALYTICS_MATERIAL_ESTIMATES.get(key)
        if estimate and estimate.get("average_weight_kg") is not None:
            return f"{float(estimate['average_weight_kg']):.3f} kg"
    return "-"


def export_quantity(scan: dict, material: dict) -> str:
    value = material.get("quantity", material.get("count", scan.get("quantity", "")))
    return "-" if value in (None, "") else str(value)


def export_reviewer(decision: dict | None) -> str:
    if not decision:
        return "-"
    return str(decision.get("reviewer") or decision.get("reviewer_id") or decision.get("reviewed_by") or "-")


def format_malaysia_datetime(value: Any) -> str:
    if value is None or value == "":
        return "-"
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, str):
        normalized = value.strip()
        if not normalized:
            return "-"
        if normalized.endswith("Z"):
            normalized = normalized[:-1] + "+00:00"
        try:
            parsed = datetime.fromisoformat(normalized)
        except ValueError:
            return "-"
    else:
        return "-"
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    malaysia_datetime = parsed.astimezone(MALAYSIA_TIMEZONE)
    weekday = malaysia_datetime.strftime("%a")
    day = malaysia_datetime.day
    month = malaysia_datetime.strftime("%b")
    hour = malaysia_datetime.strftime("%I").lstrip("0") or "0"
    minute = malaysia_datetime.strftime("%M")
    period = malaysia_datetime.strftime("%p")
    return f"{weekday} {day} {month} {hour}:{minute} {period}"


def export_scan_row(scan: dict) -> dict[str, str]:
    material, decision = export_material_decision(scan)
    category = export_final_category(scan)
    status, review_status, material_class = export_display_status(scan, material, decision)
    confidence = confidence_percent(scan.get("overall_confidence") if scan.get("overall_confidence") is not None else material.get("confidence"))
    image_url = scan.get("preview_image_url") or scan.get("image_url") or scan.get("drive_web_url") or ""
    return {
        "scan_id": str(scan.get("id") or ""),
        "datetime": format_malaysia_datetime(scan.get("created_at")),
        "file_name": str(scan.get("source_name") or "Uploaded image"),
        "predicted_category": display_label(material.get("category") or material.get("material_name")),
        "corrected_category": display_label(scan.get("verified_category") or (decision or {}).get("chosen_category") or category),
        "confidence": f"{confidence:.0f}%",
        "status": status,
        "quantity": export_quantity(scan, material),
        "estimated_weight": export_weight(scan, material),
        "recommended_route": (CATEGORY_ROUTES.get(category) or "Manual Audit Queue") if review_status != "review_needed" else "Manual Audit Queue",
        "review_status": review_status,
        "reviewer": export_reviewer(decision),
        "image_url": str(image_url),
        "material_class": display_label(material_class),
    }


def fetch_history_export_rows(
    start_date: str | None,
    end_date: str | None,
    search: str | None,
    category: str | None,
    status: str | None,
    sort: str,
    direction: str,
    principal: Principal,
) -> list[dict[str, str]]:
    filters = scan_history_filters(start_date, end_date, search, status)
    category_key = canonical_category_key(category) if category else ""
    if category and category_key == "unknown":
        return []
    order_column = "overall_confidence" if sort == "confidence" else "created_at"
    descending = str(direction).lower() != "asc"
    database = SupabaseExecutor(client_factory=_new_supabase_client, attempts=2)
    rows: list[dict[str, str]] = []
    offset = 0
    while True:
        def scan_query(client, offset=offset):
            query = client.table(SCAN_RESULTS_TABLE).select("*")
            query = apply_scan_history_filters(query, filters)
            return scoped_query(query.order(order_column, desc=descending).range(offset, offset + SCAN_HISTORY_EXPORT_BATCH_SIZE - 1), principal).execute()

        scans = database.execute(scan_query).data or []
        if not scans:
            break
        scan_ids = [str(scan.get("id")) for scan in scans if scan.get("id")]
        materials: list[dict] = []
        decisions: list[dict] = []
        for index in range(0, len(scan_ids), SCAN_HISTORY_EXPORT_CHILD_BATCH_SIZE):
            ids = scan_ids[index:index + SCAN_HISTORY_EXPORT_CHILD_BATCH_SIZE]
            materials.extend(database.execute(lambda client, ids=ids: client.table(DETECTED_MATERIALS_TABLE).select("*").in_("scan_result_id", ids).execute()).data or [])
            decisions.extend(database.execute(lambda client, ids=ids: client.table(REVIEW_DECISIONS_TABLE).select("*").in_("scan_result_id", ids).execute()).data or [])
        for scan in attach_scan_children(scans, materials, decisions):
            if category_key and export_final_category(scan) != category_key:
                continue
            rows.append(export_scan_row(scan))
        if len(scans) < SCAN_HISTORY_EXPORT_BATCH_SIZE:
            break
        offset += SCAN_HISTORY_EXPORT_BATCH_SIZE
    return rows


EXPORT_COLUMNS = [
    ("scan_id", "Scan ID"),
    ("datetime", "Date and time"),
    ("corrected_category", "Category"),
    ("confidence", "Confidence"),
    ("status", "Status"),
    ("recommended_route", "Recommended route"),
    ("image_preview", "Image Preview"),
]

PDF_EXPORT_COLUMNS = ["Scan ID", "Date/time", "Category", "Confidence", "Status", "Route", "Preview"]
EXCEL_COLUMN_WIDTHS = {
    "Scan ID": 18,
    "Date and time": 24,
    "Category": 20,
    "Confidence": 12,
    "Status": 22,
    "Recommended route": 30,
    "Image Preview": 14,
}


@dataclass(frozen=True)
class ThumbnailResult:
    url: str
    data: bytes | None = None
    error: str | None = None


def storage_signed_url(url: str) -> str | None:
    if not supabase or not SUPABASE_URL or not url.startswith(SUPABASE_URL):
        return None
    match = re.search(r"/storage/v1/object/(?:public|sign)/([^/]+)/(.+)$", url)
    if not match:
        return None
    bucket, object_path = match.group(1), unquote(match.group(2).split("?")[0])
    try:
        signed = supabase.storage.from_(bucket).create_signed_url(object_path, 3600)
        return signed.get("signedURL") or signed.get("signedUrl") or signed.get("signed_url")
    except Exception as exc:
        print(f"[history-export] signed URL fallback failed: {type(exc).__name__}")
        return None


def transformed_storage_url(url: str) -> str:
    if not SUPABASE_URL or not url.startswith(SUPABASE_URL):
        return url
    match = re.search(r"/storage/v1/object/public/([^/]+)/(.+)$", url)
    if not match:
        return url
    bucket, object_path = match.group(1), match.group(2).split("?")[0]
    return (
        f"{SUPABASE_URL}/storage/v1/render/image/public/{bucket}/{object_path}"
        f"?width={SCAN_HISTORY_THUMBNAIL_SIZE[0]}&height={SCAN_HISTORY_THUMBNAIL_SIZE[1]}"
        "&resize=contain&quality=50&format=origin"
    )


def fetch_image_bytes(client: httpx.Client, url: str) -> bytes:
    request_url = transformed_storage_url(url)
    for attempt in range(2):
        try:
            response = client.get(request_url, follow_redirects=True)
            if response.status_code >= 500 and attempt == 0:
                continue
            break
        except (httpx.TimeoutException, httpx.TransportError):
            if attempt == 0:
                continue
            raise
    if response.status_code in {403, 404}:
        signed_url = storage_signed_url(url)
        if signed_url and signed_url != url:
            response = client.get(signed_url, follow_redirects=True)
    response.raise_for_status()
    content_type = response.headers.get("content-type", "").lower()
    if content_type and not any(kind in content_type for kind in ("image/jpeg", "image/jpg", "image/png", "image/webp")):
        raise ValueError("unsupported image content type")
    content = response.content
    if len(content) > SCAN_HISTORY_IMAGE_MAX_BYTES:
        raise ValueError("source image too large")
    return content


def thumbnail_png(image_bytes: bytes) -> bytes:
    with Image.open(BytesIO(image_bytes)) as image:
        image.thumbnail(SCAN_HISTORY_THUMBNAIL_SIZE)
        canvas = Image.new("RGB", SCAN_HISTORY_THUMBNAIL_SIZE, "white")
        if image.mode in {"RGBA", "LA"} or (image.mode == "P" and "transparency" in image.info):
            image = image.convert("RGBA")
            x = (SCAN_HISTORY_THUMBNAIL_SIZE[0] - image.width) // 2
            y = (SCAN_HISTORY_THUMBNAIL_SIZE[1] - image.height) // 2
            canvas.paste(image, (x, y), image)
        else:
            image = image.convert("RGB")
            x = (SCAN_HISTORY_THUMBNAIL_SIZE[0] - image.width) // 2
            y = (SCAN_HISTORY_THUMBNAIL_SIZE[1] - image.height) // 2
            canvas.paste(image, (x, y))
        output = BytesIO()
        canvas.save(output, format="PNG", optimize=True)
        return output.getvalue()


def load_thumbnail(client: httpx.Client, url: str) -> ThumbnailResult:
    if not url:
        return ThumbnailResult(url, error="missing URL")
    try:
        return ThumbnailResult(url, data=thumbnail_png(fetch_image_bytes(client, url)))
    except Exception as exc:
        return ThumbnailResult(url, error=type(exc).__name__)


def fetch_history_thumbnails(rows: list[dict[str, str]]) -> tuple[dict[str, bytes], dict[str, int]]:
    urls = [row.get("image_url", "") for row in rows if row.get("image_url")]
    unique_urls = list(dict.fromkeys(urls))
    stats = {"requested": len(urls), "unique": len(unique_urls), "cache_hits": max(0, len(urls) - len(unique_urls)), "failed": 0, "processed": 0}
    if not unique_urls:
        return {}, stats
    thumbnails: dict[str, bytes] = {}
    limits = httpx.Limits(max_connections=SCAN_HISTORY_THUMBNAIL_WORKERS, max_keepalive_connections=SCAN_HISTORY_THUMBNAIL_WORKERS)
    with httpx.Client(timeout=SCAN_HISTORY_IMAGE_TIMEOUT, limits=limits) as client:
        with ThreadPoolExecutor(max_workers=SCAN_HISTORY_THUMBNAIL_WORKERS) as executor:
            futures = [executor.submit(load_thumbnail, client, url) for url in unique_urls]
            for future in as_completed(futures):
                result = future.result()
                stats["processed"] += 1
                if result.data:
                    thumbnails[result.url] = result.data
                else:
                    stats["failed"] += 1
                if stats["processed"] == stats["unique"] or stats["processed"] % 250 == 0:
                    print(
                        "[history-export] image progress "
                        f"processed={stats['processed']}/{stats['unique']} failed={stats['failed']} "
                        f"cache_hits={stats['cache_hits']}"
                    )
    return thumbnails, stats


def build_history_excel(rows: list[dict[str, str]], thumbnails: dict[str, bytes] | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PurityLoop History"
    worksheet.append([label for _, label in EXPORT_COLUMNS])
    preview_col = len(EXPORT_COLUMNS)
    for index, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = EXCEL_COLUMN_WIDTHS.get(label, 16)
    for row in rows:
        worksheet.append(["" if key == "image_preview" else row.get(key, "") for key, _ in EXPORT_COLUMNS])
        excel_row = worksheet.max_row
        worksheet.row_dimensions[excel_row].height = 64
        image_bytes = (thumbnails or {}).get(row.get("image_url", ""))
        if image_bytes:
            image = ExcelImage(BytesIO(image_bytes))
            image.width = SCAN_HISTORY_THUMBNAIL_SIZE[0]
            image.height = SCAN_HISTORY_THUMBNAIL_SIZE[1]
            worksheet.add_image(image, f"{get_column_letter(preview_col)}{excel_row}")
        else:
            cell = worksheet.cell(excel_row, preview_col)
            cell.value = "Image unavailable"
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    workbook.save(output)
    return output.getvalue()


def build_history_pdf(rows: list[dict[str, str]], filters: dict[str, str | None], thumbnails: dict[str, bytes] | None = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Image as PdfImage
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    body = styles["BodyText"]
    body.fontSize = 7
    body.leading = 8
    applied = ", ".join(f"{key}: {value}" for key, value in filters.items() if value) or "None"
    data = [PDF_EXPORT_COLUMNS]
    for row in rows:
        image_bytes = (thumbnails or {}).get(row.get("image_url", ""))
        preview = PdfImage(BytesIO(image_bytes), width=18 * mm, height=18 * mm) if image_bytes else Paragraph("Image unavailable", body)
        data.append([
            row["scan_id"][:8],
            row["datetime"],
            row["corrected_category"],
            row["confidence"],
            row["status"],
            row["recommended_route"],
            preview,
        ])
    if len(data) == 1:
        data.append(["No scans to export.", "", "", "", "", "", ""])
    table_data = [[cell if hasattr(cell, "wrapOn") else Paragraph(str(cell), body) for cell in record] for record in data]
    table = Table(table_data, repeatRows=1, colWidths=[24 * mm, 44 * mm, 34 * mm, 24 * mm, 42 * mm, 77 * mm, 28 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#edf5f0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#17251e")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9e4dc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fbfdfb")]),
    ]))

    def page_number(canvas, document):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(document.pagesize[0] - 10 * mm, 6 * mm, f"Page {document.page}")
        canvas.restoreState()

    story = [
        Paragraph("PurityLoop AI Audit History", styles["Title"]),
        Paragraph(f"Exported {format_malaysia_datetime(datetime.now(timezone.utc))} · {len(rows)} records", styles["Normal"]),
        Paragraph(f"Applied filters: {applied}", styles["Normal"]),
        Spacer(1, 6 * mm),
        table,
    ]
    doc.build(story, onFirstPage=page_number, onLaterPages=page_number)
    return output.getvalue()


def config_path(env_name: str, default_relative: str) -> Path:
    raw_path = os.getenv(env_name, default_relative)
    path = Path(raw_path).expanduser()
    return path if path.is_absolute() else BACKEND_ROOT / path


def google_credentials_path() -> Path | None:
    raw_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
    if not raw_path:
        return None
    path = Path(raw_path)
    if path.is_absolute():
        return path

    candidates = [APP_ROOT / path, Path.cwd() / path, BACKEND_ROOT / path]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def upload_original_to_drive(file_bytes: bytes, original_filename: str | None, content_type: str | None) -> dict:
    if not GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID:
        raise RuntimeError("GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID is not configured")

    credentials_path = google_credentials_path()
    if not credentials_path or not credentials_path.exists():
        raise RuntimeError("Google service account credentials file is not available")

    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload

    drive_file_name = safe_drive_filename(original_filename)
    mimetype = content_type or guess_type(drive_file_name)[0] or "application/octet-stream"
    credentials = service_account.Credentials.from_service_account_file(str(credentials_path), scopes=DRIVE_SCOPES)
    service = build("drive", "v3", credentials=credentials, cache_discovery=False)
    media = MediaIoBaseUpload(BytesIO(file_bytes), mimetype=mimetype, resumable=False)
    metadata = {
        "name": drive_file_name,
        "parents": [GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID],
    }

    created = (
        service.files()
        .create(body=metadata, media_body=media, fields="id, webViewLink, webContentLink", supportsAllDrives=True)
        .execute()
    )
    return {
        "drive_file_id": created.get("id"),
        "drive_file_name": drive_file_name,
        "drive_web_url": created.get("webViewLink"),
        "image_url": created.get("webViewLink"),
    }


def google_oauth_client_path() -> Path:
    return config_path("GOOGLE_OAUTH_CLIENT_SECRET_FILE", "google-oauth-client.json")


def google_oauth_token_path() -> Path:
    return config_path("GOOGLE_OAUTH_TOKEN_FILE", "google-oauth-token.json")


def google_oauth_state_path() -> Path:
    return config_path("GOOGLE_OAUTH_STATE_FILE", "google-oauth-state.json")


def oauth_flow():
    client_path = google_oauth_client_path()
    if not client_path.exists():
        raise RuntimeError("Google OAuth client file is not available")

    os.environ.setdefault("OAUTHLIB_INSECURE_TRANSPORT", "1")
    from google_auth_oauthlib.flow import Flow

    return Flow.from_client_secrets_file(
        str(client_path),
        scopes=OAUTH_DRIVE_SCOPES,
        redirect_uri=GOOGLE_OAUTH_REDIRECT_URI,
    )


def save_oauth_state(state: str, code_verifier: str | None) -> None:
    if not state or not code_verifier:
        raise RuntimeError("Google OAuth state is not available")

    state_path = google_oauth_state_path()
    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(json.dumps({"state": state, "code_verifier": code_verifier}))
    try:
        state_path.chmod(0o600)
    except OSError:
        pass


def load_oauth_state(expected_state: str | None) -> dict:
    safe_detail = "OAuth session expired. Please open /api/google/auth again."
    state_path = google_oauth_state_path()
    if not state_path.exists():
        raise HTTPException(status_code=400, detail=safe_detail)

    try:
        payload = json.loads(state_path.read_text())
    except Exception as exc:
        print(f"[google-callback] OAuth state read failed: {type(exc).__name__}: {safe_error_message(exc)}")
        raise HTTPException(status_code=400, detail=safe_detail) from exc

    saved_state = payload.get("state")
    code_verifier = payload.get("code_verifier")
    if not saved_state or not code_verifier or saved_state != expected_state:
        raise HTTPException(status_code=400, detail=safe_detail)
    return payload


def save_oauth_token(credentials) -> None:
    token_path = google_oauth_token_path()
    token_path.parent.mkdir(parents=True, exist_ok=True)
    token_path.write_text(credentials.to_json())
    try:
        token_path.chmod(0o600)
    except OSError:
        pass


def oauth_drive_credentials():
    from google.auth.transport.requests import Request as GoogleAuthRequest
    from google.oauth2.credentials import Credentials

    env_client_id = os.getenv("GOOGLE_CLIENT_ID")
    env_client_secret = os.getenv("GOOGLE_CLIENT_SECRET")
    env_refresh_token = os.getenv("GOOGLE_REFRESH_TOKEN")
    if env_client_id and env_client_secret and env_refresh_token:
        # Refresh tokens already carry their granted scopes. Re-requesting an
        # expanded scope set here can make Google's refresh endpoint reject it.
        credentials = Credentials(
            token=None,
            refresh_token=env_refresh_token,
            token_uri="https://oauth2.googleapis.com/token",
            client_id=env_client_id,
            client_secret=env_client_secret,
        )
        credentials.refresh(GoogleAuthRequest())
        if not credentials.valid:
            raise RuntimeError("Google OAuth environment credentials are not valid")
        return credentials

    token_path = google_oauth_token_path()
    if not token_path.exists():
        raise RuntimeError("Google OAuth token file is not available")

    credentials = Credentials.from_authorized_user_file(str(token_path))
    if credentials.expired and credentials.refresh_token:
        credentials.refresh(GoogleAuthRequest())
        save_oauth_token(credentials)
    if not credentials.valid:
        raise RuntimeError("Google OAuth token is not valid")
    return credentials


def upload_original_to_drive_oauth(
    file_bytes: bytes,
    original_filename: str | None,
    content_type: str | None,
    *,
    submission_id: UUID | None = None,
) -> dict:
    if not GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID:
        raise RuntimeError("GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID is not configured")

    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload

    drive_file_name = safe_drive_filename(original_filename)
    mimetype = content_type or guess_type(drive_file_name)[0] or "application/octet-stream"
    service = build("drive", "v3", credentials=oauth_drive_credentials(), cache_discovery=False)
    if submission_id:
        existing = service.files().list(
            q=(
                f"trashed = false and '{GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID}' in parents and "
                f"appProperties has {{ key='purityloop_submission_id' and value='{submission_id}' }}"
            ),
            spaces="drive",
            fields="files(id,name,webViewLink,webContentLink)",
            pageSize=1,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute().get("files", [])
        if existing:
            created = existing[0]
            return {
                "drive_file_id": created.get("id"),
                "drive_file_name": created.get("name") or drive_file_name,
                "drive_web_url": created.get("webViewLink"),
                "image_url": created.get("webViewLink"),
            }
    media = MediaIoBaseUpload(BytesIO(file_bytes), mimetype=mimetype, resumable=False)
    metadata = {
        "name": drive_file_name,
        "parents": [GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID],
    }
    if submission_id:
        metadata["appProperties"] = {"purityloop_submission_id": str(submission_id)}

    created = (
        service.files()
        .create(body=metadata, media_body=media, fields="id, webViewLink, webContentLink", supportsAllDrives=True)
        .execute()
    )
    return {
        "drive_file_id": created.get("id"),
        "drive_file_name": drive_file_name,
        "drive_web_url": created.get("webViewLink"),
        "image_url": created.get("webViewLink"),
    }


def legacy_scan_row(scan_row: dict, original_filename: str | None) -> dict:
    keep_keys = {
        "image_url",
        "preview_image_url",
        "drive_file_id",
        "drive_file_name",
        "drive_web_url",
        "source_type",
        "upload_status",
        "processing_status",
        "overall_status",
        "contamination_risk",
        "recommended_action",
        "human_review_required",
        "overall_confidence",
        "user_id",
    }
    row = {key: value for key, value in scan_row.items() if key in keep_keys}
    row["image_url"] = row.get("image_url")
    row["upload_status"] = row.get("upload_status") or "uploaded"
    return row


def material_category(name: str) -> str:
    text = re.sub(r"[_-]+", " ", str(name or "").lower()).strip()
    if "battery" in text:
        return "battery"
    if "food" in text or "organic" in text:
        return "food_organics"
    if "trash" in text or "waste" in text:
        return "general_trash"
    if "textile" in text or "fabric" in text or "cloth" in text:
        return "textile"
    if "glass" in text or "jar" in text:
        return "glass"
    if "cardboard" in text or "box" in text:
        return "cardboard"
    if "paper" in text:
        return "paper"
    if "metal" in text or "aluminum" in text or "aluminium" in text or "can" in text:
        return "metal"
    if "plastic" in text or "bottle" in text or "pet" in text:
        return "plastic"
    return "unknown"


def material_status(category: str) -> tuple[str, str]:
    if CATEGORY_CLASS_MAP.get(category) == "contaminant":
        return "non_recyclable", "contaminated"
    if CATEGORY_CLASS_MAP.get(category) == "recyclable":
        return "recyclable", "clean"
    return "unknown", "unknown"


def normalized_confidence(value: Any) -> float:
    try:
        confidence = float(value)
    except (TypeError, ValueError):
        return 0.0
    if not math.isfinite(confidence):
        return 0.0
    if confidence > 1:
        confidence = confidence / 100
    return max(0.0, min(1.0, confidence))


def normalize_status(value: Any) -> str:
    return str(value or "").strip().lower().replace("-", "_").replace(" ", "_")


def derive_final_status(*, confidence: Any, decision: dict | None = None, scan: dict | None = None) -> str:
    """Canonical object status. Human decisions win; auto confidence threshold is inclusive."""
    decision_status = normalize_status((decision or {}).get("outcome") or (decision or {}).get("review_outcome"))
    scan_status = normalize_status((scan or {}).get("review_status") or (scan or {}).get("overall_status"))
    if decision_status == "rejected" or scan_status in {"rejected", "quarantined"}:
        return "rejected"
    if decision_status == "confirmed" or scan_status in {"verified", "corrected"}:
        return "confirmed"
    try:
        numeric = float(confidence)
    except (TypeError, ValueError):
        return "needs_review"
    if not math.isfinite(numeric):
        return "needs_review"
    return "confirmed" if normalized_confidence(numeric) >= DECISION_CONFIDENCE_THRESHOLD else "needs_review"


def object_metrics_from_rows(scans: list[dict], materials: list[dict], decisions: list[dict]) -> dict[str, int]:
    latest = latest_decisions_by_material(decisions)
    scan_by_id = {str(scan.get("id")): scan for scan in scans if scan.get("id")}
    counts = {"total_objects": 0, "confirmed_objects": 0, "needs_review_objects": 0, "rejected_objects": 0}
    for material in materials:
        scan = scan_by_id.get(str(material.get("scan_result_id") or ""), {})
        status = derive_final_status(confidence=material.get("confidence"), decision=latest.get(str(material.get("id") or "")), scan=scan)
        counts["total_objects"] += 1
        counts[f"{status}_objects"] += 1
    return counts


def _looks_like_uuid(value: Any) -> bool:
    try:
        UUID(str(value))
        return True
    except (TypeError, ValueError):
        return False


def determine_detection_status(confidence: float, is_contaminant: bool, category: str = "unknown") -> dict[str, str]:
    if derive_final_status(confidence=confidence) == "needs_review":
        return {
            "review_status": "needs_review",
            "ai_status": "low_confidence_detection",
        }
    return {
        "review_status": "confirmed",
        "ai_status": "confirmed_contaminant" if is_contaminant else "confirmed_recyclable",
    }


def evaluate_material(category: str, confidence: float) -> dict:
    material_class = CATEGORY_CLASS_MAP.get(category, "unknown")
    confidence = normalized_confidence(confidence)
    status = determine_detection_status(confidence, material_class == "contaminant", category)
    review_required = status["review_status"] == "needs_review"
    decision_status = "review_needed" if review_required else "confirmed"
    if review_required:
        display_status = "Review Needed"
        disposal_route = "Manual Audit Queue"
    elif material_class == "recyclable":
        display_status = "Confirmed Recyclable"
        disposal_route = CATEGORY_ROUTES[category]
    elif material_class == "contaminant":
        display_status = "Confirmed Contaminant"
        disposal_route = CATEGORY_ROUTES[category]
    else:
        display_status = "Confirmed"
        disposal_route = "Manual Audit Queue"
    return {
        "material_class": material_class,
        "review_required": review_required,
        "decision_status": decision_status,
        "display_status": display_status,
        "disposal_route": disposal_route,
    }


def to_detected_materials(result) -> list[dict]:
    names = result.names
    image_height, image_width = result.orig_shape
    materials = []
    for box in result.boxes:
        xyxy = box.xyxy[0].tolist()
        confidence = float(box.conf[0])
        class_id = int(box.cls[0])
        material_name = str(names.get(class_id, f"class_{class_id}"))
        category = material_category(material_name)
        recyclable_status, contaminant_status = material_status(category)
        materials.append(
            {
                "material_name": material_name,
                "category": category,
                "confidence": round(confidence, 4),
                "recyclable_status": recyclable_status,
                "contaminant_status": contaminant_status,
                **evaluate_material(category, confidence),
                "bbox_x": round((float(xyxy[0]) / image_width) * 100, 2),
                "bbox_y": round((float(xyxy[1]) / image_height) * 100, 2),
                "bbox_width": round((float(xyxy[2] - xyxy[0]) / image_width) * 100, 2),
                "bbox_height": round((float(xyxy[3] - xyxy[1]) / image_height) * 100, 2),
            }
        )
    return materials


def _valid_xyxy(box: list[float], width: int, height: int) -> bool:
    if len(box) < 4:
        return False
    x1, y1, x2, y2 = [_coerce_float(value) for value in box[:4]]
    return (
        math.isfinite(x1)
        and math.isfinite(y1)
        and math.isfinite(x2)
        and math.isfinite(y2)
        and x2 > x1
        and y2 > y1
        and x1 < width
        and y1 < height
        and x2 > 0
        and y2 > 0
    )


def _normalized_xyxy_to_pixels(box: list[float], image_width: int, image_height: int) -> list[float]:
    if len(box) < 4 or not all(0 <= _coerce_float(value) <= 1 for value in box[:4]):
        raise ValueError("Expected normalized xyxy values between 0 and 1")
    return [
        _coerce_float(box[0]) * image_width,
        _coerce_float(box[1]) * image_height,
        _coerce_float(box[2]) * image_width,
        _coerce_float(box[3]) * image_height,
    ]


def _percentage_xywh_to_pixels(x: float, y: float, width: float, height: float, image_width: int, image_height: int) -> list[float]:
    values = [_coerce_float(x), _coerce_float(y), _coerce_float(width), _coerce_float(height)]
    if not all(0 <= value <= 100 for value in values):
        raise ValueError("Expected percentage xywh values between 0 and 100")
    return [
        (values[0] / 100) * image_width,
        (values[1] / 100) * image_height,
        ((values[0] + values[2]) / 100) * image_width,
        ((values[1] + values[3]) / 100) * image_height,
    ]


def _detection_box_to_pixels(material: dict, image_width: int, image_height: int) -> tuple[list[float], str]:
    best_box = material.get("best_box")
    if isinstance(best_box, dict):
        xyxy = best_box.get("xyxy")
        if isinstance(xyxy, list) and _valid_xyxy(xyxy, image_width, image_height):
            return [_coerce_float(value) for value in xyxy[:4]], "pixel_xyxy:best_box.xyxy"
        normalized = best_box.get("normalized_xyxy") or best_box.get("xyxy_normalized")
        if isinstance(normalized, list):
            pixels = _normalized_xyxy_to_pixels(normalized, image_width, image_height)
            if _valid_xyxy(pixels, image_width, image_height):
                return pixels, "normalized_xyxy:best_box"
        if all(key in best_box for key in ("x", "y", "width", "height")):
            pixels = _percentage_xywh_to_pixels(best_box.get("x"), best_box.get("y"), best_box.get("width"), best_box.get("height"), image_width, image_height)
            if _valid_xyxy(pixels, image_width, image_height):
                return pixels, "percentage_xywh:best_box"
    best_bbox_norm = material.get("best_bbox_norm")
    if isinstance(best_bbox_norm, list):
        pixels = _normalized_xyxy_to_pixels(best_bbox_norm, image_width, image_height)
        if _valid_xyxy(pixels, image_width, image_height):
            return pixels, "normalized_xyxy:best_bbox_norm"
    bbox = material.get("bbox")
    if isinstance(bbox, list):
        pixels = _normalized_xyxy_to_pixels(bbox, image_width, image_height)
        if _valid_xyxy(pixels, image_width, image_height):
            return pixels, "normalized_xyxy:bbox"
    if all(material.get(key) is not None for key in ("bbox_x", "bbox_y", "bbox_width", "bbox_height")):
        pixels = _percentage_xywh_to_pixels(material.get("bbox_x"), material.get("bbox_y"), material.get("bbox_width"), material.get("bbox_height"), image_width, image_height)
        if _valid_xyxy(pixels, image_width, image_height):
            return pixels, "percentage_xywh:material_bbox"
    raise ValueError("No valid detection box metadata is available")


def _material_bbox_xyxy(material: dict, image_width: int, image_height: int) -> list[float]:
    box, _format = _detection_box_to_pixels(material, image_width, image_height)
    return box


def _material_preview_detection(material: dict, image_width: int, image_height: int) -> dict:
    x1, y1, x2, y2 = _material_bbox_xyxy(material, image_width, image_height)
    return {
        "track_id": material.get("track_id"),
        "category": material.get("category"),
        "material_name": material.get("material_name") or material.get("category"),
        "confidence": material.get("confidence"),
        "best_box": {"xyxy": [x1, y1, x2, y2]},
        "bbox": [
            x1 / image_width if image_width else 0,
            y1 / image_height if image_height else 0,
            x2 / image_width if image_width else 0,
            y2 / image_height if image_height else 0,
        ],
        "mask": material.get("segmentation_mask") or material.get("mask"),
    }


def _clean_preview_detections(detections: list[dict], image_width: int, image_height: int) -> list[dict]:
    candidates = []
    for detection in detections:
        if _coerce_float(detection.get("confidence")) < PREVIEW_BOX_CONFIDENCE_THRESHOLD:
            continue
        try:
            box, _box_format = _detection_box_to_pixels(detection, image_width, image_height)
        except Exception:
            continue
        if _preview_box_is_extreme_edge_strip(box, image_width, image_height):
            continue
        candidates.append((detection, box))
    kept: list[tuple[dict, list[float]]] = []
    for detection, box in sorted(candidates, key=lambda item: _coerce_float(item[0].get("confidence")), reverse=True):
        if all(_bbox_iou(box, kept_box) < PREVIEW_BOX_NMS_IOU_THRESHOLD for _kept, kept_box in kept):
            kept.append((detection, box))
    return [detection for detection, _box in kept]


def _preview_box_is_extreme_edge_strip(box: list[float], image_width: int, image_height: int) -> bool:
    if image_width <= 0 or image_height <= 0 or len(box) < 4:
        return False
    x1, y1, x2, y2 = [_coerce_float(value) for value in box[:4]]
    box_width = max(0.0, x2 - x1)
    box_height = max(0.0, y2 - y1)
    if box_width <= 0 or box_height <= 0:
        return False
    aspect_ratio = box_width / box_height
    extreme_strip = aspect_ratio < PREVIEW_EDGE_STRIP_ASPECT_RATIO_MIN or aspect_ratio > PREVIEW_EDGE_STRIP_ASPECT_RATIO_MAX
    if not extreme_strip:
        return False
    tolerance_x = max(PREVIEW_EDGE_TOLERANCE_MIN_PIXELS, image_width * PREVIEW_EDGE_TOLERANCE_RATIO)
    tolerance_y = max(PREVIEW_EDGE_TOLERANCE_MIN_PIXELS, image_height * PREVIEW_EDGE_TOLERANCE_RATIO)
    return x1 <= tolerance_x or y1 <= tolerance_y or x2 >= image_width - tolerance_x or y2 >= image_height - tolerance_y


def _translate_mask_to_crop(mask, image_width: int, image_height: int, crop_x: int, crop_y: int, crop_width: int, crop_height: int) -> tuple[list | None, str]:
    if not mask:
        return None, "unavailable"
    translated = []
    source_format = "pixel"
    for point in mask:
        if not isinstance(point, (list, tuple)) or len(point) < 2:
            continue
        x = _coerce_float(point[0])
        y = _coerce_float(point[1])
        if 0 <= x <= 1 and 0 <= y <= 1:
            source_format = "normalized"
            x *= image_width
            y *= image_height
        translated.append([
            max(0, min(crop_width - 1, round(x - crop_x))),
            max(0, min(crop_height - 1, round(y - crop_y))),
        ])
    return (translated, source_format) if len(translated) >= 3 else (None, source_format)


def _tracked_object_crop_bounds(frame, material: dict) -> tuple[int, int, int, int, list[int]]:
    image_height, image_width = frame.shape[:2]
    box, _format = _detection_box_to_pixels(material, image_width, image_height)
    x1, y1, x2, y2 = _clip_box(box, image_width, image_height)
    box_width = max(1, x2 - x1)
    box_height = max(1, y2 - y1)
    padding = max(8, round(max(box_width, box_height) * 0.18))
    crop_x1 = max(0, x1 - padding)
    crop_y1 = max(0, y1 - padding)
    crop_x2 = min(image_width, x2 + padding)
    crop_y2 = min(image_height, y2 + padding)
    min_crop_width = min(image_width, max(180, box_width + padding * 2))
    min_crop_height = min(image_height, max(120, box_height + padding * 2))
    extra_width = max(0, min_crop_width - (crop_x2 - crop_x1))
    extra_height = max(0, min_crop_height - (crop_y2 - crop_y1))
    grow_left = min(crop_x1, extra_width // 2)
    crop_x1 -= grow_left
    crop_x2 = min(image_width, crop_x2 + (extra_width - grow_left))
    if crop_x2 - crop_x1 < min_crop_width:
        crop_x1 = max(0, crop_x2 - min_crop_width)
    grow_top = min(crop_y1, extra_height // 2)
    crop_y1 -= grow_top
    crop_y2 = min(image_height, crop_y2 + (extra_height - grow_top))
    if crop_y2 - crop_y1 < min_crop_height:
        crop_y1 = max(0, crop_y2 - min_crop_height)
    return crop_x1, crop_y1, crop_x2, crop_y2, [x1, y1, x2, y2]


def _tracked_object_crop_preview(frame, material: dict) -> tuple[bytes, dict]:
    import cv2

    image_height, image_width = frame.shape[:2]
    bbox_input = {
        "best_box": material.get("best_box"),
        "best_bbox_norm": material.get("best_bbox_norm"),
        "bbox": material.get("bbox"),
        "bbox_x": material.get("bbox_x"),
        "bbox_y": material.get("bbox_y"),
        "bbox_width": material.get("bbox_width"),
        "bbox_height": material.get("bbox_height"),
    }
    _box, bbox_format = _detection_box_to_pixels(material, image_width, image_height)
    if bbox_format == "percentage_xywh:material_bbox" and (material.get("stable_object_id") or material.get("object_uid") or material.get("result_type") == "video_track_object"):
        raise ValueError("Tracked-object preview requires representative-frame bbox metadata, not material percentage bbox fallback")
    crop_x1, crop_y1, crop_x2, crop_y2, box_xyxy = _tracked_object_crop_bounds(frame, material)
    x1, y1, x2, y2 = box_xyxy
    crop = frame[crop_y1:crop_y2, crop_x1:crop_x2]
    if crop.size == 0:
        raise ValueError("Tracked-object preview crop is empty")
    crop_height, crop_width = crop.shape[:2]
    translated_mask, mask_format = _translate_mask_to_crop(
        material.get("segmentation_mask") or material.get("mask"),
        image_width,
        image_height,
        crop_x1,
        crop_y1,
        crop_width,
        crop_height,
    )
    detection = {
        "track_id": material.get("track_id"),
        "category": material.get("category"),
        "material_name": material.get("material_name") or material.get("category"),
        "confidence": material.get("confidence") or material.get("track_max_confidence") or material.get("track_avg_confidence"),
        "best_box": {"xyxy": [x1 - crop_x1, y1 - crop_y1, x2 - crop_x1, y2 - crop_y1]},
        "bbox": [
            (x1 - crop_x1) / crop_width if crop_width else 0,
            (y1 - crop_y1) / crop_height if crop_height else 0,
            (x2 - crop_x1) / crop_width if crop_width else 0,
            (y2 - crop_y1) / crop_height if crop_height else 0,
        ],
        "mask": translated_mask,
    }
    annotated = _annotate_video_frame(crop, [detection])
    ok, encoded = cv2.imencode(".jpg", annotated)
    if not ok:
        raise ValueError("Unable to encode tracked-object annotated preview")
    return encoded.tobytes(), {
        "format": "representative_frame_annotation",
        "source_width": image_width,
        "source_height": image_height,
        "crop_x": crop_x1,
        "crop_y": crop_y1,
        "crop_width": crop_width,
        "crop_height": crop_height,
        "bbox_input": bbox_input,
        "bbox_format": bbox_format,
        "mask_format": mask_format,
        "box_xyxy": box_xyxy,
        "translated_box_xyxy": detection["best_box"]["xyxy"],
        "frame": (material.get("best_box") or {}).get("frame") if isinstance(material.get("best_box"), dict) else None,
        "timestamp": (material.get("best_box") or {}).get("timestamp") if isinstance(material.get("best_box"), dict) else None,
    }


def _encode_tracked_object_preview(frame_bytes: bytes, filename: str | None, material: dict) -> tuple[bytes, dict]:
    import cv2
    import numpy as np

    frame = cv2.imdecode(np.frombuffer(frame_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)
    if frame is None:
        raise ValueError("Unable to decode tracked-object representative frame")
    return _tracked_object_crop_preview(frame, material)


def _track_id_values(material: dict) -> set[str]:
    values: set[str] = set()
    for value in material.get("source_track_ids") or []:
        if value is not None and str(value) != "":
            values.add(str(value))
    for value in str(material.get("track_id") or "").split(","):
        cleaned = value.strip()
        if cleaned:
            values.add(cleaned)
    return values


def _annotated_detection_observation(detection: dict, frame_index: int, timestamp: float, width: int, height: int) -> dict | None:
    track_id = detection.get("track_id")
    if track_id is None or str(track_id) == "":
        return None
    box_xyxy, bbox_format = _detection_box_to_pixels(detection, width, height)
    x1, y1, x2, y2 = _clip_box(box_xyxy, width, height)
    if x2 <= x1 or y2 <= y1:
        return None
    category = material_category(detection.get("category") or detection.get("material_name"))
    confidence = max(0.0, min(1.0, _coerce_float(detection.get("confidence"))))
    return {
        "track_id": str(track_id),
        "annotated_frame_index": int(frame_index),
        "source_frame_index": int(frame_index),
        "timestamp": round(timestamp, 3),
        "box_xyxy": [x1, y1, x2, y2],
        "bbox_format": bbox_format,
        "bbox": [
            x1 / width if width else 0,
            y1 / height if height else 0,
            x2 / width if width else 0,
            y2 / height if height else 0,
        ],
        "category": category,
        "material_name": detection.get("material_name") or category,
        "confidence": round(confidence, 4),
        "mask": detection.get("mask"),
        "track_hazard_status": "hazard" if CATEGORY_CLASS_MAP.get(category) == "contaminant" else "clear",
        "frame_width": int(width),
        "frame_height": int(height),
    }


def _record_annotated_frame_observations(observations_by_track: dict[str, list[dict]], detections: list[dict], frame_index: int, timestamp: float, width: int, height: int) -> None:
    for detection in detections:
        try:
            observation = _annotated_detection_observation(detection, frame_index, timestamp, width, height)
        except Exception as exc:
            _video_processing_log(
                "annotated_frame_observation_invalid",
                frame_index=frame_index,
                track_id=detection.get("track_id"),
                width=width,
                height=height,
                error_type=type(exc).__name__,
                error=safe_error_message(exc),
            )
            continue
        if not observation:
            continue
        observations_by_track.setdefault(observation["track_id"], []).append(observation)


def _track_frame_fragments(observations: list[dict]) -> dict[str, list[dict]]:
    frames_by_track: dict[str, list[int]] = {}
    for observation in observations or []:
        if not isinstance(observation, dict):
            continue
        track_id = str(observation.get("track_id") or "").strip()
        frame = int(_coerce_float(observation.get("frame", observation.get("source_frame_index")), -1))
        if not track_id or frame < 0:
            continue
        frames_by_track.setdefault(track_id, []).append(frame)

    fragments: dict[str, list[dict]] = {}
    for track_id, frames in frames_by_track.items():
        sorted_frames = sorted(set(frames))
        ranges = []
        start = previous = sorted_frames[0]
        for frame in sorted_frames[1:]:
            if frame == previous + 1:
                previous = frame
                continue
            ranges.append({"first_frame": start, "last_frame": previous})
            start = previous = frame
        ranges.append({"first_frame": start, "last_frame": previous})
        fragments[track_id] = ranges
    return fragments


def _material_track_fragments(material: dict) -> dict[str, list[tuple[int, int]]]:
    debug = material.get("track_debug") if isinstance(material.get("track_debug"), dict) else {}
    raw_fragments = debug.get("accepted_track_fragments") if isinstance(debug, dict) else None
    if not isinstance(raw_fragments, dict):
        return {}
    fragments: dict[str, list[tuple[int, int]]] = {}
    for track_id, ranges in raw_fragments.items():
        if not isinstance(ranges, list):
            continue
        normalized = []
        for item in ranges:
            if not isinstance(item, dict):
                continue
            start = int(_coerce_float(item.get("first_frame"), -1))
            end = int(_coerce_float(item.get("last_frame"), -1))
            if start < 0 or end < start:
                continue
            normalized.append((start, end))
        if normalized:
            fragments[str(track_id)] = normalized
    return fragments


def _frame_matches_material_track(material: dict, track_id: str, frame: int) -> bool:
    fragments = _material_track_fragments(material)
    if fragments:
        return any(start <= frame <= end for start, end in fragments.get(str(track_id), []))

    first_frame = int(_coerce_float(material.get("track_first_frame"), -1))
    last_frame = int(_coerce_float(material.get("track_last_frame"), -1))
    if first_frame >= 0 and frame < first_frame:
        return False
    if last_frame >= 0 and frame > last_frame:
        return False
    return True


def _select_annotated_preview_observation(material: dict, observations_by_track: dict[str, list[dict]] | None) -> dict | None:
    track_ids = _track_id_values(material)
    if not track_ids or not observations_by_track:
        return None
    candidates = []
    for track_id in track_ids:
        for observation in observations_by_track.get(str(track_id), []):
            if str(observation.get("track_id") or "") not in track_ids:
                continue
            frame = int(_coerce_float(observation.get("source_frame_index", observation.get("frame")), -1))
            if not _frame_matches_material_track(material, str(track_id), frame):
                continue
            width = int(_coerce_float(observation.get("frame_width"), 0))
            height = int(_coerce_float(observation.get("frame_height"), 0))
            box = observation.get("box_xyxy")
            if not isinstance(box, list) or not _valid_xyxy(box, width, height):
                continue
            candidates.append(observation)
    if not candidates:
        return None
    return max(candidates, key=lambda item: (_coerce_float(item.get("confidence")), int(_coerce_float(item.get("source_frame_index"), -1))))


def _material_observation_fallback(material: dict) -> dict | None:
    track_ids = _track_id_values(material)
    best_box = material.get("best_box") if isinstance(material.get("best_box"), dict) else {}
    if not track_ids or best_box.get("frame") is None:
        return None
    box = best_box.get("xyxy")
    if not isinstance(box, list):
        return None
    width = int(_coerce_float((material.get("track_debug") or {}).get("representative_frame_dimensions", {}).get("width"), 0)) if isinstance(material.get("track_debug"), dict) else 0
    height = int(_coerce_float((material.get("track_debug") or {}).get("representative_frame_dimensions", {}).get("height"), 0)) if isinstance(material.get("track_debug"), dict) else 0
    if width <= 0 or height <= 0:
        width = max(1, math.ceil(max(_coerce_float(value) for value in box[:4])))
        height = width
    if not _valid_xyxy(box, width, height):
        return None
    track_id = sorted(track_ids)[0]
    return {
        "track_id": track_id,
        "annotated_frame_index": int(_coerce_float(best_box.get("frame"))),
        "source_frame_index": int(_coerce_float(best_box.get("frame"))),
        "timestamp": _coerce_float(best_box.get("timestamp"), -1),
        "box_xyxy": [_coerce_float(value) for value in box[:4]],
        "bbox_format": "pixel_xyxy:best_box.xyxy",
        "category": material.get("category"),
        "material_name": material.get("material_name") or material.get("category"),
        "confidence": material.get("confidence") or material.get("track_max_confidence") or material.get("track_avg_confidence") or 0,
        "mask": material.get("segmentation_mask") or material.get("mask"),
        "track_hazard_status": material.get("track_hazard_status"),
        "frame_width": width,
        "frame_height": height,
    }


def _material_from_annotated_observation(material: dict, observation: dict) -> dict:
    width = int(_coerce_float(observation.get("frame_width"), 0))
    height = int(_coerce_float(observation.get("frame_height"), 0))
    x1, y1, x2, y2 = [_coerce_float(value) for value in observation["box_xyxy"][:4]]
    merged = dict(material)
    merged["category"] = material.get("category") or observation.get("category")
    merged["material_name"] = material.get("material_name") or material.get("category") or observation.get("material_name")
    merged["confidence"] = material.get("confidence") if material.get("confidence") is not None else observation.get("confidence")
    merged["best_box"] = {
        "xyxy": [x1, y1, x2, y2],
        "frame": observation.get("annotated_frame_index"),
        "source_frame": observation.get("source_frame_index"),
        "timestamp": observation.get("timestamp"),
        "track_id": observation.get("track_id"),
    }
    if width and height:
        merged["best_bbox_norm"] = [x1 / width, y1 / height, x2 / width, y2 / height]
    if observation.get("mask") is not None:
        merged["segmentation_mask"] = observation.get("mask")
    return merged


def _extract_annotated_video_object_preview(video_path: str | Path | None, material: dict, observation: dict | None = None, raw_video_path: str | Path | None = None) -> tuple[bytes, dict] | None:
    if not video_path and not raw_video_path:
        return None
    import cv2

    target_video = Path(raw_video_path) if raw_video_path and Path(raw_video_path).exists() else Path(video_path or "")
    if not target_video.exists():
        return None
    observation = observation or _material_observation_fallback(material)
    if not observation:
        return None
    track_ids = _track_id_values(material)
    observation_track_id = str(observation.get("track_id") or "")
    if not observation_track_id or (track_ids and observation_track_id not in track_ids):
        return None
    if observation.get("annotated_frame_index") is None:
        return None
    if observation.get("bbox_format") == "percentage_xywh:material_bbox":
        return None
    frame_index = int(_coerce_float(observation.get("annotated_frame_index")))
    timestamp = _coerce_float(observation.get("timestamp"), -1)
    capture = cv2.VideoCapture(str(target_video))
    try:
        if not capture.isOpened():
            return None
        fps = float(capture.get(cv2.CAP_PROP_FPS) or 0)
        frame_count = int(capture.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
        source_width = int(capture.get(cv2.CAP_PROP_FRAME_WIDTH) or 0)
        source_height = int(capture.get(cv2.CAP_PROP_FRAME_HEIGHT) or 0)
        frame = None
        extraction_method = None
        if frame_index >= 0:
            capture.set(cv2.CAP_PROP_POS_FRAMES, max(0, frame_index))
            ok, candidate = capture.read()
            if ok and candidate is not None:
                frame = candidate
                extraction_method = "frame_index"
        if frame is None and timestamp >= 0:
            capture.set(cv2.CAP_PROP_POS_MSEC, timestamp * 1000)
            ok, candidate = capture.read()
            if ok and candidate is not None:
                frame = candidate
                extraction_method = "timestamp"
        if frame is None:
            return None
        extraction_material = _material_from_annotated_observation(material, observation)
        bbox_format = observation.get("bbox_format")
        try:
            crop_x1, crop_y1, crop_x2, crop_y2, box_xyxy = _tracked_object_crop_bounds(frame, extraction_material)
            category = extraction_material.get("category") or material.get("category") or "Detected object"
            confidence = float(extraction_material.get("confidence") or material.get("confidence") or 0)
            conf_pct = round(confidence * 100) if confidence <= 1.0 else round(confidence)
            conf_label = f"{category} | {conf_pct}%"
            color = (0, 200, 80) if category.lower() not in ("trash", "quarantine", "rejected") else (50, 50, 220)
            bx1, by1, bx2, by2 = [int(v) for v in box_xyxy]
            if bx2 > bx1 and by2 > by1:
                (lw, lh), _ = cv2.getTextSize(conf_label, cv2.FONT_HERSHEY_SIMPLEX, 0.55, 2)
                hy1 = max(0, by1 - lh - 8)
                cv2.rectangle(frame, (bx1, hy1), (min(frame.shape[1], bx1 + lw + 10), by1), color, -1)
                cv2.putText(frame, conf_label, (bx1 + 4, max(16, by1 - 5)), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 2)
                cv2.rectangle(frame, (bx1, by1), (bx2, by2), color, 2)
            crop = frame[crop_y1:crop_y2, crop_x1:crop_x2]
            if crop.size == 0:
                raise ValueError("Tracked-object annotated-video preview crop is empty")
        except Exception as exc:
            _video_processing_log(
                "tracked_preview_crop_failed",
                track_id=observation_track_id,
                stable_object_id=material.get("stable_object_id") or material.get("object_uid"),
                video_path=str(target_video),
                frame_index=frame_index,
                timestamp=timestamp,
                fps=fps,
                frame_count=frame_count,
                source_width=source_width,
                source_height=source_height,
                bbox=observation.get("box_xyxy"),
                stage="crop",
                worker_build_revision=_worker_build_revision(),
                error_type=type(exc).__name__,
                error=safe_error_message(exc),
            )
            return None
        ok, encoded = cv2.imencode(".jpg", crop)
        if not ok:
            raise ValueError("Unable to encode annotated-video object preview")
        crop_height, crop_width = crop.shape[:2]
        return encoded.tobytes(), {
            "format": "annotated_video_frame",
            "preview_source": "annotated_video_frame",
            "extraction_method": extraction_method,
            "video_path": str(target_video),
            "source_width": source_width or frame.shape[1],
            "source_height": source_height or frame.shape[0],
            "fps": fps,
            "frame_count": frame_count,
            "crop_x": crop_x1,
            "crop_y": crop_y1,
            "crop_x2": crop_x2,
            "crop_y2": crop_y2,
            "crop_width": crop_width,
            "crop_height": crop_height,
            "bbox_format": bbox_format,
            "box_xyxy": box_xyxy,
            "track_id": observation_track_id,
            "frame": frame_index,
            "source_frame": observation.get("source_frame_index"),
            "timestamp": timestamp,
        }
    finally:
        capture.release()


def _encode_annotated_image_preview(file_bytes: bytes, filename: str | None, materials: list[dict]) -> bytes:
    if not materials:
        return file_bytes
    try:
        import cv2
        import numpy as np

        frame = cv2.imdecode(np.frombuffer(file_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)
        if frame is None:
            raise ValueError("Unable to decode image preview for annotation")
        image_height, image_width = frame.shape[:2]
        detections = _clean_preview_detections(
            [_material_preview_detection(material, image_width, image_height) for material in materials],
            image_width,
            image_height,
        )
        annotated = _annotate_video_frame(frame, detections, footer_count=len(detections))
        suffix = (Path(filename or "upload.jpg").suffix or ".jpg").lower()
        extension = ".jpg" if suffix == ".jpeg" else suffix
        if extension not in {".jpg", ".png", ".webp"}:
            extension = ".jpg"
        ok, encoded = cv2.imencode(extension, annotated)
        if not ok:
            raise ValueError(f"Unable to encode annotated image preview as {extension}")
        return encoded.tobytes()
    except Exception as exc:
        _video_processing_log(
            "image_preview_annotation_failed",
            filename=filename,
            detection_count=len(materials),
            error_type=type(exc).__name__,
            error=safe_error_message(exc),
        )
        return file_bytes


def summarize(materials: list[dict]) -> dict:
    if not materials:
        return {
            "overall_status": "review_required",
            "contamination_risk": "medium",
            "recommended_action": "Human review recommended before sorting.",
            "human_review_required": True,
            "overall_confidence": 0,
        }

    avg_confidence = sum(item["confidence"] for item in materials) / len(materials)
    contaminated = any(item["contaminant_status"] == "contaminated" for item in materials)
    review_required = any(item["review_required"] for item in materials)

    return {
        "overall_status": "review_required" if review_required else "accepted",
        "contamination_risk": "medium" if contaminated else "low",
        "recommended_action": "Human review required before sorting." if review_required else "Confirmed sorting routes applied.",
        "human_review_required": review_required,
        "overall_confidence": round(avg_confidence, 4),
    }


def _image_content_type(filename: str | None, content_type: str | None) -> tuple[str, str]:
    suffix = Path(filename or "").suffix.lower()
    suffix_types = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp"}
    mime_types = {"image/jpeg": ".jpg", "image/jpg": ".jpg", "image/png": ".png", "image/webp": ".webp"}
    mime = str(content_type or "").split(";", 1)[0].strip().lower()
    generic_mime = mime in {"", "application/octet-stream", "binary/octet-stream"}
    normalized = suffix_types.get(suffix)
    if normalized:
        if mime and not (mime.startswith("image/") or generic_mime):
            raise HTTPException(status_code=400, detail="Upload one JPG, PNG, or WebP image file.")
        return suffix, normalized
    if mime in mime_types:
        inferred_suffix = mime_types[mime]
        return inferred_suffix, suffix_types[inferred_suffix]
    if not normalized:
        raise HTTPException(status_code=400, detail="Upload one JPG, PNG, or WebP image file.")
    return suffix, normalized


def _load_scan(database: SupabaseExecutor, scan_result_id: UUID | str) -> dict | None:
    response = database.execute(
        lambda client: client.table(SCAN_RESULTS_TABLE).select("*").eq("id", str(scan_result_id)).maybe_single().execute()
    )
    return response.data if response and response.data else None


def _load_scan_materials(database: SupabaseExecutor, scan_result_id: UUID | str) -> list[dict]:
    response = database.execute(
        lambda client: client.table(DETECTED_MATERIALS_TABLE).select("*").eq("scan_result_id", str(scan_result_id)).execute()
    )
    return response.data or []


def _load_review_decisions(database: SupabaseExecutor, scan_result_id: UUID | str) -> list[dict]:
    response = database.execute(
        lambda client: client.table(REVIEW_DECISIONS_TABLE).select("*").eq(
            "scan_result_id", str(scan_result_id)
        ).execute()
    )
    return response.data or []


def _scan_response(scan: dict, materials: list[dict]) -> dict:
    video_summary = scan.get("video_tracking_summary") if isinstance(scan.get("video_tracking_summary"), dict) else {}
    return {
        "scan_result_id": scan["id"],
        "overall_status": scan.get("overall_status"),
        "contamination_risk": scan.get("contamination_risk"),
        "recommended_action": scan.get("recommended_action"),
        "human_review_required": scan.get("human_review_required"),
        "overall_confidence": scan.get("overall_confidence"),
        "storage_provider": scan.get("storage_provider"),
        "upload_status": scan.get("upload_status"),
        "drive_upload_status": scan.get("drive_upload_status"),
        "preview_upload_status": scan.get("preview_upload_status"),
        "drive_file_id": scan.get("drive_file_id"),
        "drive_file_name": scan.get("drive_file_name"),
        "drive_web_url": scan.get("drive_web_url"),
        "image_url": scan.get("image_url"),
        "preview_image_url": scan.get("preview_image_url"),
        "result_kind": scan.get("result_kind"),
        "legacy_result": scan.get("legacy_result"),
        "total_unique_objects": scan.get("total_unique_objects"),
        "counts_by_class": video_summary.get("counts_by_class"),
        "hazards": video_summary.get("hazards"),
        "annotated_video_url": scan.get("annotated_video_url") or video_summary.get("annotated_video_url"),
        "annotated_video_storage_path": scan.get("annotated_video_storage_path") or video_summary.get("annotated_video_storage_path"),
        "annotated_video_status": scan.get("annotated_video_status") or video_summary.get("annotated_video_status"),
        "annotated_video_error": scan.get("annotated_video_error") or video_summary.get("annotated_video_error"),
        "annotated_video_probe": video_summary.get("annotated_video_probe"),
        "tracked_objects": materials if str(scan.get("result_kind") or "") in {"tracked_video_object", "video_track_object"} else None,
        "detected_materials": materials,
        "backend_build_version": BACKEND_BUILD_VERSION,
    }


def _browser_storage_path(submission_id: UUID, filename: str | None) -> str:
    suffix = (Path(filename or "upload.jpg").suffix or ".jpg").lower()
    return f"browser-onnx/{submission_id}{suffix}"


def persist_scan(
    file_bytes: bytes,
    filename: str | None,
    source_type: str,
    materials: list[dict],
    summary: dict,
    *,
    source_ref: str | None = None,
    batch_id: str | None = None,
    principal: Principal | None = None,
    content_type: str | None = None,
    existing_drive_metadata: dict | None = None,
    frame_time_seconds: float | None = None,
    database: SupabaseExecutor | None = None,
    scan_result_id: UUID | None = None,
    model_version: str | None = None,
    verified: bool = False,
) -> dict:
    """Persist one canonical scan. A supplied UUID enables browser retry recovery."""
    database = database or SupabaseExecutor(supabase)
    if not database.client:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    suffix, normalized_type = _image_content_type(filename, content_type)
    existing_scan = _load_scan(database, scan_result_id) if scan_result_id else None
    if existing_scan and existing_scan.get("processing_status") == "complete":
        stored_materials = _load_scan_materials(database, scan_result_id)
        stored_decisions = _load_review_decisions(database, scan_result_id) if verified else []
        storage_complete = (
            existing_scan.get("drive_upload_status") == "uploaded"
            and existing_scan.get("preview_upload_status") == "uploaded"
        )
        if (
            len(stored_materials) == len(materials)
            and (not verified or len(stored_decisions) == len(materials))
            and storage_complete
        ):
            return _scan_response(existing_scan, stored_materials)
    if scan_result_id and not existing_scan:
        reservation = {
            "id": str(scan_result_id),
            "source_type": source_type,
            "source_name": filename or "uploaded-image",
            "source_ref": source_ref,
            "model_version": model_version or BROWSER_MODEL_VERSION,
            "processing_status": "pending",
            "upload_status": "pending",
            "overall_status": "review_required",
            "human_review_required": True,
        }
        principal_profile_id = scan_user_id(principal, database)
        if principal_profile_id:
            reservation["user_id"] = principal_profile_id
        try:
            inserted = database.execute(
                lambda client: client.table(SCAN_RESULTS_TABLE).insert(reservation).execute().data or []
            )
            existing_scan = inserted[0] if inserted else _load_scan(database, scan_result_id)
        except Exception as exc:
            if getattr(exc, "code", "") != "23505":
                raise
            existing_scan = _load_scan(database, scan_result_id)
        if not existing_scan:
            raise HTTPException(status_code=500, detail="Unable to reserve browser scan result.")

    drive_metadata = {
        "storage_provider": "google_drive_and_supabase_storage",
        "upload_status": "pending",
        "drive_upload_status": "pending",
        "preview_upload_status": "pending",
        "drive_file_id": None,
        "drive_file_name": filename or "uploaded-image",
        "drive_web_url": None,
        "image_url": None,
        "preview_image_url": None,
    }
    if existing_scan:
        drive_metadata.update({
            key: existing_scan.get(key)
            for key in drive_metadata
            if existing_scan.get(key) is not None
        })

    if existing_drive_metadata:
        drive_metadata.update(existing_drive_metadata)
        drive_metadata["drive_upload_status"] = "uploaded" if drive_metadata.get("drive_file_id") else "failed"
    elif not drive_metadata.get("drive_file_id"):
        try:
            drive_metadata.update(
                upload_original_to_drive_oauth(
                    file_bytes,
                    filename,
                    normalized_type,
                    submission_id=scan_result_id,
                )
            )
            drive_metadata["drive_upload_status"] = "uploaded"
        except Exception as exc:
            drive_metadata["drive_upload_status"] = "failed"
            print(f"[Google Drive] Upload failed: {type(exc).__name__}: {safe_error_message(exc)}")

    if not drive_metadata.get("preview_image_url"):
        try:
            preview_upload = upload_original_to_supabase_storage(
                file_bytes,
                filename,
                normalized_type,
                database,
                object_path=_browser_storage_path(scan_result_id, filename) if scan_result_id else None,
            )
            drive_metadata["preview_image_url"] = preview_upload["public_url"]
            drive_metadata["preview_upload_status"] = "uploaded"
        except Exception as exc:
            drive_metadata["preview_upload_status"] = "failed"
            print(f"[Supabase Storage] Upload failed: {type(exc).__name__}: {safe_error_message(exc)}")

    drive_metadata["upload_status"] = (
        "uploaded" if drive_metadata["preview_upload_status"] == "uploaded" else "preview_upload_failed"
    )
    scan_row = {
        **drive_metadata,
        "source_type": source_type,
        "source_name": filename or "uploaded-image",
        "source_ref": source_ref,
        "batch_id": batch_id,
        "model_version": model_version or os.getenv("MODEL_VERSION", "yolov8-purityloop"),
        "processing_status": "complete",
        **summary,
    }
    if verified:
        scan_row.update({
            "overall_status": "verified",
            "human_review_required": False,
            "review_status": "verified",
            "recommended_action": "Verified after operator review.",
            "reviewed_at": datetime.now(timezone.utc).isoformat(),
        })
    principal_profile_id = scan_user_id(principal, database)
    if principal_profile_id:
        scan_row["user_id"] = principal_profile_id

    def recover_scan(client):
        if scan_result_id:
            response = client.table(SCAN_RESULTS_TABLE).select("*").eq("id", str(scan_result_id)).maybe_single().execute()
            return [response.data] if response and response.data else None
        if source_type == "video_frame" and batch_id:
            response = client.table(SCAN_RESULTS_TABLE).select("*").eq("batch_id", batch_id).eq(
                "source_name", scan_row["source_name"]
            ).maybe_single().execute()
            return [response.data] if response and response.data else None
        return None

    if existing_scan:
        scan_data = database.execute(
            lambda client: client.table(SCAN_RESULTS_TABLE).update(scan_row).eq(
                "id", str(scan_result_id)
            ).execute().data or []
        )
    else:
        insert_row = {**scan_row, **({"id": str(scan_result_id)} if scan_result_id else {})}
        try:
            scan_data = database.execute(
                lambda client: client.table(SCAN_RESULTS_TABLE).insert(insert_row).execute().data or [],
                recover=recover_scan,
            )
        except Exception as exc:
            if getattr(exc, "code", "") in {"23505"} and scan_result_id:
                scan_data = recover_scan(database.client) or []
            elif getattr(exc, "code", "") in {"PGRST204", "PGRST205", "42703"} and not scan_result_id:
                scan_data = database.execute(
                    lambda client: client.table(SCAN_RESULTS_TABLE).insert(
                        legacy_scan_row(scan_row, filename)
                    ).execute().data or [],
                    recover=recover_scan,
                )
            else:
                raise
    if not scan_data:
        raise HTTPException(status_code=500, detail="Unable to save scan result.")

    saved_scan_id = scan_data[0]["id"]
    stored_material_keys = {
        "material_name", "category", "confidence", "recyclable_status", "contaminant_status",
        "bbox_x", "bbox_y", "bbox_width", "bbox_height", "original_category", "stable_object_id",
        "object_uid", "source_track_ids", "track_id", "track_first_frame", "track_last_frame", "track_first_timestamp",
        "track_last_timestamp", "track_duration_seconds", "track_avg_confidence",
        "track_max_confidence", "track_frame_count", "track_hazard_status", "track_counted",
        "track_start_center", "track_end_center", "track_avg_width", "track_avg_height",
        "track_avg_aspect_ratio", "track_debug", "track_path", "segmentation_mask", "best_box",
        "best_bbox_norm", "result_type",
    }
    legacy_material_keys = {
        "material_name", "category", "confidence", "recyclable_status", "contaminant_status",
        "bbox_x", "bbox_y", "bbox_width", "bbox_height", "original_category",
    }
    linked_materials = []
    for index, item in enumerate(materials):
        linked = {
            key: value for key, value in item.items() if key in stored_material_keys
        } | {"scan_result_id": saved_scan_id}
        if scan_result_id:
            linked["id"] = str(uuid5(scan_result_id, f"material:{item.get('object_uid') or index}"))
        if frame_time_seconds is not None:
            linked["frame_time_seconds"] = frame_time_seconds
        linked_materials.append(linked)

    existing_materials = _load_scan_materials(database, saved_scan_id) if scan_result_id else []
    existing_material_ids = {str(item.get("id")) for item in existing_materials}
    missing_materials = [
        item for item in linked_materials if not item.get("id") or str(item["id"]) not in existing_material_ids
    ]
    if missing_materials:
        def recover_materials(client):
            response = client.table(DETECTED_MATERIALS_TABLE).select("*").eq(
                "scan_result_id", saved_scan_id
            ).execute()
            rows = response.data or []
            return rows if len(rows) >= len(linked_materials) else None

        try:
            database.execute(
                lambda client: client.table(DETECTED_MATERIALS_TABLE).insert(missing_materials).execute().data or [],
                recover=recover_materials,
            )
        except Exception as exc:
            if getattr(exc, "code", "") in {"PGRST204", "PGRST205", "42703"}:
                database.execute(
                    lambda client: client.table(DETECTED_MATERIALS_TABLE).insert(
                        [{key: value for key, value in item.items() if key in legacy_material_keys or key in {"id", "scan_result_id"}} for item in missing_materials]
                    ).execute().data or [],
                    recover=recover_materials,
                )
            elif getattr(exc, "code", "") != "23505":
                raise

    stored_materials = _load_scan_materials(database, saved_scan_id)
    if len(stored_materials) != len(linked_materials):
        raise HTTPException(status_code=500, detail="Unable to retrieve saved detected materials.")

    if verified:
        existing_decisions = _load_review_decisions(database, saved_scan_id)
        existing_decision_ids = {str(item.get("id")) for item in existing_decisions}
        decisions = []
        materials_by_id = {str(item["id"]): item for item in stored_materials}
        for index, material in enumerate(materials):
            material_id = str(uuid5(scan_result_id, f"material:{index}"))
            decision_id = str(uuid5(scan_result_id, f"review:{index}"))
            if decision_id in existing_decision_ids:
                continue
            category = materials_by_id[material_id]["category"]
            decisions.append({
                "id": decision_id,
                "scan_result_id": saved_scan_id,
                "detected_material_id": material_id,
                "chosen_category": category,
                "disposition": CATEGORY_CLASS_MAP[category],
                "outcome": "confirmed",
                "reviewer_email": principal.id if principal else "public",
            })
        if decisions:
            try:
                database.execute(
                    lambda client: client.table(REVIEW_DECISIONS_TABLE).insert(decisions).execute().data or []
                )
            except Exception as exc:
                if getattr(exc, "code", "") != "23505":
                    raise

    saved_scan = _load_scan(database, saved_scan_id) or scan_data[0]
    clear_analytics_summary_cache()
    return _scan_response(saved_scan, stored_materials)


def run_scan(
    file_bytes: bytes,
    filename: str | None,
    source_type: str,
    *,
    source_ref: str | None = None,
    batch_id: str | None = None,
    principal: Principal | None = None,
    content_type: str | None = None,
    existing_drive_metadata: dict | None = None,
    frame_time_seconds: float | None = None,
    database: SupabaseExecutor | None = None,
) -> dict:
    """Run one image through the existing YOLO path and persist one canonical scan."""
    suffix, _ = _image_content_type(filename, content_type)
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        result = predict_with_calibration(tmp_path)[0]
        materials = to_detected_materials(result)
        preview_bytes = _encode_annotated_image_preview(file_bytes, filename, materials)
        return persist_scan(
            preview_bytes,
            filename,
            source_type,
            materials,
            summarize(materials),
            source_ref=source_ref,
            batch_id=batch_id,
            principal=principal,
            content_type=content_type,
            existing_drive_metadata=existing_drive_metadata,
            frame_time_seconds=frame_time_seconds,
            database=database,
        )
    finally:
        if tmp_path:
            Path(tmp_path).unlink(missing_ok=True)


class ReviewDecisionInput(BaseModel):
    scan_result_id: UUID
    detected_material_id: UUID
    action: str
    manual_category: str | None = None
    reviewer_email: str | None = None


class FalsePositiveReportInput(BaseModel):
    detected_material_id: UUID | None = None
    expected_category: str
    reason: str
    note: str | None = None


class FalsePositiveDismissInput(BaseModel):
    reason: str | None = None


class BrowserVerifiedDetection(BaseModel):
    detection_index: int
    class_id: int
    model_class_name: str
    verified_class: str
    confidence: float
    x1: float
    y1: float
    x2: float
    y2: float
    verification_status: str


class BrowserDetectedDetection(BaseModel):
    detection_index: int
    class_id: int
    model_class_name: str
    confidence: float
    x1: float
    y1: float
    x2: float
    y2: float


def validate_detection_confidence(value: object) -> float:
    try:
        confidence = float(value)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="Detection confidence must be a valid number.") from exc
    if not math.isfinite(confidence):
        raise HTTPException(status_code=400, detail="Detection confidence must be finite.")
    if not 0.0 <= confidence <= 1.0:
        raise HTTPException(status_code=400, detail=BROWSER_CONFIDENCE_DETAIL)
    return confidence


def validate_browser_detections(
    raw_detections: Any,
    image_width: int,
    image_height: int,
) -> list[dict]:
    if not isinstance(raw_detections, list) or not raw_detections:
        raise HTTPException(status_code=400, detail="At least one verified detection is required.")

    validated = []
    indexes = set()
    for raw in raw_detections:
        try:
            detection = BrowserVerifiedDetection(**raw)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Verified detection JSON is invalid.") from exc
        if detection.detection_index < 0 or detection.detection_index in indexes:
            raise HTTPException(status_code=400, detail="Detection indexes must be unique non-negative integers.")
        indexes.add(detection.detection_index)
        if detection.class_id < 0 or detection.class_id >= len(BROWSER_MODEL_CLASSES):
            raise HTTPException(status_code=400, detail="Detection class ID is outside the fixed model contract.")
        expected_name = BROWSER_MODEL_CLASSES[detection.class_id]
        if detection.model_class_name != expected_name:
            raise HTTPException(status_code=400, detail="Detection class ID and model class name do not match.")
        if detection.verified_class not in BROWSER_MODEL_CLASSES:
            raise HTTPException(status_code=400, detail="Verified class is outside the fixed model contract.")
        confidence = validate_detection_confidence(detection.confidence)
        coordinates = (detection.x1, detection.y1, detection.x2, detection.y2)
        if not all(math.isfinite(value) for value in coordinates):
            raise HTTPException(status_code=400, detail="Detection coordinates must be finite.")
        x1 = min(max(detection.x1, 0), image_width)
        y1 = min(max(detection.y1, 0), image_height)
        x2 = min(max(detection.x2, 0), image_width)
        y2 = min(max(detection.y2, 0), image_height)
        if x2 <= x1 or y2 <= y1:
            raise HTTPException(status_code=400, detail="Detection bounding box must have positive area.")
        battery_detected = detection.model_class_name == "battery" or detection.verified_class == "battery"
        required_status = "battery-confirmed" if battery_detected else "verified"
        if detection.verification_status != required_status:
            detail = (
                "Battery detections require explicit human confirmation."
                if battery_detected
                else "Every detection must be human verified."
            )
            raise HTTPException(status_code=400, detail=detail)

        category = "food_organics" if detection.verified_class == "food_organic" else detection.verified_class
        original_category = "food_organics" if detection.model_class_name == "food_organic" else detection.model_class_name
        recyclable_status, contaminant_status = material_status(category)
        validated.append({
            "_detection_index": detection.detection_index,
            "material_name": detection.verified_class,
            "category": category,
            "original_category": original_category if original_category != category else None,
            "confidence": round(confidence, 4),
            "recyclable_status": recyclable_status,
            "contaminant_status": contaminant_status,
            "bbox_x": round((x1 / image_width) * 100, 4),
            "bbox_y": round((y1 / image_height) * 100, 4),
            "bbox_width": round(((x2 - x1) / image_width) * 100, 4),
            "bbox_height": round(((y2 - y1) / image_height) * 100, 4),
            **evaluate_material(category, confidence),
        })
    return sorted(validated, key=lambda item: item["_detection_index"])


def validate_browser_detected_detections(raw_detections: Any, image_width: int, image_height: int) -> list[dict]:
    if not isinstance(raw_detections, list):
        raise HTTPException(status_code=400, detail="Detection JSON must be an array.")
    validated = []
    indexes = set()
    for raw in raw_detections:
        try:
            detection = BrowserDetectedDetection(**raw)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Detection JSON is invalid.") from exc
        if detection.detection_index < 0 or detection.detection_index in indexes:
            raise HTTPException(status_code=400, detail="Detection indexes must be unique non-negative integers.")
        indexes.add(detection.detection_index)
        if detection.class_id < 0 or detection.class_id >= len(BROWSER_MODEL_CLASSES):
            raise HTTPException(status_code=400, detail="Detection class ID is outside the fixed model contract.")
        category_name = BROWSER_MODEL_CLASSES[detection.class_id]
        if detection.model_class_name != category_name:
            raise HTTPException(status_code=400, detail="Detection class ID and model class name do not match.")
        confidence = validate_detection_confidence(detection.confidence)
        coordinates = (detection.x1, detection.y1, detection.x2, detection.y2)
        if not all(math.isfinite(value) for value in coordinates):
            raise HTTPException(status_code=400, detail="Detection coordinates must be finite.")
        x1, y1 = min(max(detection.x1, 0), image_width), min(max(detection.y1, 0), image_height)
        x2, y2 = min(max(detection.x2, 0), image_width), min(max(detection.y2, 0), image_height)
        if x2 <= x1 or y2 <= y1:
            raise HTTPException(status_code=400, detail="Detection bounding box must have positive area.")
        category = "food_organics" if category_name == "food_organic" else category_name
        recyclable_status, contaminant_status = material_status(category)
        material = {
            "_detection_index": detection.detection_index,
            "material_name": category_name,
            "category": category,
            "confidence": round(confidence, 4),
            "recyclable_status": recyclable_status,
            "contaminant_status": contaminant_status,
            "bbox_x": round((x1 / image_width) * 100, 4),
            "bbox_y": round((y1 / image_height) * 100, 4),
            "bbox_width": round(((x2 - x1) / image_width) * 100, 4),
            "bbox_height": round(((y2 - y1) / image_height) * 100, 4),
            **evaluate_material(category, confidence),
        }
        if category == "battery":
            material.update({
                "review_required": True,
                "decision_status": "review_needed",
                "display_status": "Review Needed",
                "disposal_route": "Manual Audit Queue",
            })
        validated.append(material)
    return sorted(validated, key=lambda item: item["_detection_index"])


FALSE_POSITIVE_REASONS = {
    "wrong_class",
    "incorrect_object",
    "background_false_detection",
    "duplicate_detection",
    "low_quality_prediction",
    "other",
}
FALSE_POSITIVE_ACTIVE_STATUSES = {"reported", "queued", "reprocessing", "failed"}
FALSE_POSITIVE_TERMINAL_STATUSES = {"resolved", "dismissed"}
FALSE_POSITIVE_REPROCESS_ACTIVE_STATUSES = {"queued", "reprocessing"}
FALSE_POSITIVE_REVIEW_ROLES = {"development_team", "plant_manager"}


def false_positive_log(event: str, **fields: Any) -> None:
    safe_fields = {
        key: ("[redacted]" if key.lower() in {"token", "authorization", "service_role_key", "api_key", "password", "signed_url"} else value)
        for key, value in fields.items()
    }
    print(f"[false-positive] {event} {json.dumps(safe_fields, default=str, sort_keys=True)}")


def _load_active_profile(database: SupabaseExecutor, principal: Principal) -> dict | None:
    if principal.kind != "user":
        return None
    response = database.execute(
        lambda client: client.table("user_profiles")
        .select("id,auth_user_id,email,role,status,deleted_at")
        .eq("auth_user_id", principal.id)
        .maybe_single()
        .execute()
    )
    profile = response.data if response else None
    if not profile or profile.get("status") != "active" or profile.get("deleted_at"):
        return None
    return profile


def _require_false_positive_principal(database: SupabaseExecutor, principal: Principal) -> dict:
    profile = _load_active_profile(database, principal)
    if not profile:
        raise HTTPException(status_code=403, detail="Active workspace profile is required.")
    return profile


def _require_false_positive_reviewer(database: SupabaseExecutor, principal: Principal) -> dict:
    profile = _require_false_positive_principal(database, principal)
    if profile.get("role") not in FALSE_POSITIVE_REVIEW_ROLES:
        raise HTTPException(status_code=403, detail="Development Team or Plant Manager access is required.")
    return profile


def _first_material(scan: dict, material_id: str | None, database: SupabaseExecutor) -> dict:
    def select_material(client: Any) -> Any:
        query = client.table(DETECTED_MATERIALS_TABLE).select("*").eq("scan_result_id", str(scan["id"]))
        if material_id:
            query = query.eq("id", material_id)
        return query.limit(1).execute()

    response = database.execute(select_material)
    rows = response.data or []
    if not rows:
        raise HTTPException(status_code=404, detail="Detected material was not found for this scan.")
    return rows[0]


def _load_scan_for_false_positive(scan_id: str, database: SupabaseExecutor, principal: Principal) -> dict:
    response = database.execute(
        lambda client: scoped_query(client.table(SCAN_RESULTS_TABLE).select("*").eq("id", scan_id), principal).execute()
    )
    if not response.data:
        raise HTTPException(status_code=404, detail="Scan result was not found.")
    return response.data[0]


def _safe_false_positive_report(row: dict | None) -> dict | None:
    if not row:
        return None
    return {
        key: row.get(key)
        for key in (
            "id",
            "original_scan_id",
            "original_detected_material_id",
            "batch_id",
            "processing_job_id",
            "reported_by",
            "predicted_category",
            "predicted_confidence",
            "expected_category",
            "reason",
            "note",
            "source_type",
            "source_name",
            "source_storage_path",
            "source_drive_file_id",
            "original_model_version",
            "original_model_hash",
            "status",
            "reprocess_job_id",
            "reprocessed_scan_id",
            "reprocess_model_version",
            "reprocess_model_hash",
            "failure_reason",
            "created_at",
            "updated_at",
            "resolved_at",
            "dismissed_at",
        )
    }


def resolve_original_upload(scan_id: str, database: SupabaseExecutor, principal: Principal, detected_material_id: str | None = None) -> dict:
    try:
        UUID(str(scan_id))
    except ValueError as exc:
        raise HTTPException(status_code=404, detail="Scan result was not found.") from exc
    scan = _load_scan_for_false_positive(str(scan_id), database, principal)
    material = _first_material(scan, str(detected_material_id) if detected_material_id else None, database)
    source_type = str(scan.get("source_type") or scan.get("result_kind") or "image")
    if source_type == "tracked_video" or scan.get("result_kind") in {"video_track_object", "tracked_video_object"}:
        source_type = "video/mp4"
    source = {
        "scan": scan,
        "material": material,
        "batch_id": scan.get("batch_id"),
        "processing_job_id": scan.get("processing_job_id"),
        "source_type": source_type,
        "source_name": scan.get("source_name") or scan.get("drive_file_name"),
        "source_storage_path": scan.get("source_storage_path") or scan.get("source_ref"),
        "source_drive_file_id": scan.get("drive_file_id"),
        "original_model_version": scan.get("model_version"),
        "original_model_hash": scan.get("model_hash"),
        "source_owner": scan.get("user_id"),
        "original_category": canonical_category_key(material.get("category") or material.get("material_name")),
        "original_confidence": normalized_confidence(material.get("confidence") if material.get("confidence") is not None else scan.get("overall_confidence")),
    }
    if not source["source_drive_file_id"] and not source["source_storage_path"]:
        false_positive_log("false_positive_source_resolution_failed", original_scan_id=scan_id, original_detected_material_id=material.get("id"), safe_error_code="SOURCE_REFERENCE_MISSING")
        raise HTTPException(status_code=409, detail={"code": "SOURCE_REFERENCE_MISSING", "message": "Original source reference is missing."})
    false_positive_log(
        "false_positive_source_resolved",
        original_scan_id=scan_id,
        original_detected_material_id=material.get("id"),
        batch_id=source["batch_id"],
        source_type=source["source_type"],
        source_filename=source["source_name"],
        source_drive_file_id=source["source_drive_file_id"],
        original_category=source["original_category"],
        original_confidence=source["original_confidence"],
    )
    return source


def _load_false_positive_report(database: SupabaseExecutor, report_id: str) -> dict:
    try:
        UUID(str(report_id))
    except ValueError as exc:
        raise HTTPException(status_code=404, detail="False-positive report was not found.") from exc
    response = database.execute(
        lambda client: client.table(FALSE_POSITIVE_REPORTS_TABLE).select("*").eq("id", report_id).maybe_single().execute()
    )
    report = response.data if response else None
    if not report:
        raise HTTPException(status_code=404, detail="False-positive report was not found.")
    return report


def _update_false_positive_report(database: SupabaseExecutor, report_id: str, **fields: Any) -> dict:
    response = database.execute(
        lambda client: client.table(FALSE_POSITIVE_REPORTS_TABLE)
        .update({**fields, "updated_at": datetime.now(timezone.utc).isoformat()})
        .eq("id", report_id)
        .execute()
    )
    rows = response.data or []
    if not rows:
        raise HTTPException(status_code=500, detail="Unable to update false-positive report.")
    return rows[0]


def _active_false_positive_for_material(database: SupabaseExecutor, scan_id: str, material_id: str, expected_category: str) -> dict | None:
    response = database.execute(
        lambda client: client.table(FALSE_POSITIVE_REPORTS_TABLE)
        .select("*")
        .eq("original_scan_id", scan_id)
        .eq("original_detected_material_id", material_id)
        .eq("expected_category", expected_category)
        .in_("status", list(FALSE_POSITIVE_ACTIVE_STATUSES))
        .limit(1)
        .execute()
    )
    rows = response.data or []
    return rows[0] if rows else None


def _find_active_reprocess_job(database: SupabaseExecutor, report: dict) -> dict | None:
    job_id = report.get("reprocess_job_id")
    if not job_id:
        return None
    response = database.execute(lambda client: client.table(JOBS_TABLE).select("*").eq("id", str(job_id)).maybe_single().execute())
    job = response.data if response else None
    if job and str(job.get("status") or "") in {"upload_pending", "queued", "processing"}:
        return job
    return None


def _create_false_positive_reprocess_job(database: SupabaseExecutor, report: dict, principal: Principal) -> dict:
    if _find_active_reprocess_job(database, report):
        false_positive_log("false_positive_reprocess_duplicate_blocked", report_id=report.get("id"), reprocess_job_id=report.get("reprocess_job_id"))
        raise HTTPException(status_code=409, detail={"message": "A reprocessing job is already queued or running.", "job_id": report.get("reprocess_job_id")})
    drive_file_id = report.get("source_drive_file_id")
    if not drive_file_id:
        updated = _update_false_positive_report(database, str(report["id"]), status="failed", failure_reason="UNSUPPORTED_SOURCE_TYPE")
        false_positive_log("false_positive_source_resolution_failed", report_id=report.get("id"), safe_error_code="UNSUPPORTED_SOURCE_TYPE")
        raise HTTPException(status_code=409, detail={"code": "UNSUPPORTED_SOURCE_TYPE", "message": "Only Drive-backed sources can be reprocessed by the current pipeline.", "report": _safe_false_positive_report(updated)})
    job_row = {
        "source": "drive_file",
        "source_ref": drive_file_id,
        "options": {
            "job_type": "false_positive_reprocess",
            "parent_report_id": str(report["id"]),
            "parent_scan_id": str(report["original_scan_id"]),
            "parent_job_id": str(report.get("processing_job_id") or ""),
            "expected_category": report.get("expected_category"),
        },
        "created_by": principal.id,
        "created_by_type": principal.kind,
    }
    inserted = database.execute(lambda client: client.table(JOBS_TABLE).insert(job_row).execute().data or [])
    if not inserted:
        raise HTTPException(status_code=500, detail="Unable to create reprocessing job.")
    job = inserted[0]
    updated = _update_false_positive_report(database, str(report["id"]), status="queued", reprocess_job_id=str(job["id"]), failure_reason=None)
    if PROCESSING_BACKEND == "cloud-tasks":
        try:
            task = enqueue_processing_task(str(job["id"]))
            database.execute(
                lambda client: client.table(JOBS_TABLE).update(
                    {"status": "queued", "dispatched_at": datetime.now(timezone.utc).isoformat(), "dispatch_error": None}
                ).eq("id", str(job["id"])).execute()
            )
        except Exception as exc:
            updated = _update_false_positive_report(database, str(report["id"]), status="failed", failure_reason=f"QUEUE_FAILURE: {safe_error_message(exc)}")
            database.execute(
                lambda client: client.table(JOBS_TABLE).update(
                    {"status": "queued", "dispatch_error": safe_error_message(exc), "updated_at": datetime.now(timezone.utc).isoformat()}
                ).eq("id", str(job["id"])).execute()
            )
            false_positive_log("false_positive_reprocess_failed", report_id=report.get("id"), reprocess_job_id=job.get("id"), safe_error_code="QUEUE_FAILURE", safe_error_message=safe_error_message(exc))
            raise HTTPException(status_code=503, detail={"message": "Unable to dispatch reprocessing task.", "report": _safe_false_positive_report(updated)}) from exc
    false_positive_log("false_positive_reprocess_queued", report_id=report.get("id"), reprocess_job_id=job.get("id"), original_scan_id=report.get("original_scan_id"), source_drive_file_id=drive_file_id)
    return {"job": job, "report": updated}


def _finalize_false_positive_reprocess(database: SupabaseExecutor, job_id: str, scan_ids: list[str]) -> None:
    response = database.execute(
        lambda client: client.table(FALSE_POSITIVE_REPORTS_TABLE).select("*").eq("reprocess_job_id", job_id).limit(1).execute()
    )
    reports = response.data or []
    if not reports:
        return
    report = reports[0]
    fields: dict[str, Any] = {
        "reprocessed_scan_id": scan_ids[0] if scan_ids else None,
        "reprocess_model_version": os.getenv("MODEL_VERSION", "yolov8-purityloop"),
        "reprocess_model_hash": os.getenv("MODEL_HASH") or "",
        "failure_reason": None,
    }
    status = "failed"
    if scan_ids:
        try:
            scan = _load_scan_for_false_positive(str(scan_ids[0]), database, Principal("api_key", "worker", frozenset({"scan:read"})))
            material = _first_material(scan, None, database)
            new_category = canonical_category_key(material.get("category") or material.get("material_name"))
            new_confidence = normalized_confidence(material.get("confidence") if material.get("confidence") is not None else scan.get("overall_confidence"))
            expected = canonical_category_key(report.get("expected_category"))
            if new_confidence >= DECISION_CONFIDENCE_THRESHOLD and new_category == expected:
                status = "resolved"
                fields["resolved_at"] = datetime.now(timezone.utc).isoformat()
            else:
                status = "reported"
            false_positive_log(
                "false_positive_reprocess_completed" if status == "resolved" else "false_positive_reprocess_unresolved",
                report_id=report.get("id"),
                reprocess_job_id=job_id,
                reprocessed_scan_id=scan_ids[0],
                expected_category=expected,
                new_category=new_category,
                new_confidence=new_confidence,
                decision_threshold=DECISION_CONFIDENCE_THRESHOLD,
                report_resolution_decision=status,
            )
        except Exception as exc:
            status = "failed"
            fields["failure_reason"] = safe_worker_error_message(exc)
    else:
        fields["failure_reason"] = "REPROCESSING_CREATED_NO_SCAN"
    _update_false_positive_report(database, str(report["id"]), status=status, **fields)


class UploadStartInput(BaseModel):
    filename: str
    size_bytes: int
    mime: str


class UploadStartFailure(RuntimeError):
    def __init__(self, code: str, message: str, *, status_code: int = 500, stage: str = "unknown"):
        super().__init__(message)
        self.code = code
        self.status_code = status_code
        self.stage = stage


def _upload_start_log(event: str, **fields) -> None:
    safe_fields = {}
    for key, value in fields.items():
        lowered = key.lower()
        if lowered in {"token", "authorization", "upload_url", "signed_url", "service_role_key", "api_key", "password"}:
            safe_fields[key] = "[redacted]"
        elif isinstance(value, Path):
            safe_fields[key] = str(value)
        else:
            safe_fields[key] = value
    print(f"[upload-start] {event} {json.dumps(safe_fields, sort_keys=True, default=str)}")


def _upload_start_error_response(exc: UploadStartFailure) -> HTTPException:
    return HTTPException(
        status_code=exc.status_code,
        detail={
            "code": exc.code,
            "message": str(exc),
            "stage": exc.stage,
        },
    )


def _validate_upload_start_payload(payload: UploadStartInput) -> tuple[str, int, str]:
    filename = str(payload.filename or "").strip()
    mime = str(payload.mime or "").strip().lower()
    try:
        size_bytes = int(payload.size_bytes)
    except Exception as exc:
        raise UploadStartFailure("INVALID_UPLOAD_PAYLOAD", "Upload size must be an integer.", status_code=400, stage="validate_payload") from exc
    if not filename:
        raise UploadStartFailure("INVALID_UPLOAD_PAYLOAD", "Filename is required.", status_code=400, stage="validate_payload")
    if size_bytes <= 0:
        raise UploadStartFailure("INVALID_UPLOAD_PAYLOAD", "MP4 upload must be non-empty.", status_code=400, stage="validate_payload")
    if MAX_VIDEO_UPLOAD_BYTES and size_bytes > MAX_VIDEO_UPLOAD_BYTES:
        raise UploadStartFailure("UPLOAD_TOO_LARGE", "MP4 upload exceeds the configured size limit.", status_code=413, stage="validate_payload")
    if mime != "video/mp4":
        raise UploadStartFailure("UNSUPPORTED_MEDIA_TYPE", "Only video/mp4 uploads are supported by /api/uploads/start.", status_code=415, stage="validate_payload")
    return filename, size_bytes, mime


def _create_drive_resumable_upload(filename: str, size_bytes: int, mime: str) -> str:
    if not GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID:
        raise UploadStartFailure(
            "MISSING_ENVIRONMENT_VARIABLE",
            "GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID is not configured.",
            status_code=503,
            stage="validate_environment",
        )
    try:
        from google.auth.transport.requests import AuthorizedSession
        credentials = oauth_drive_credentials()
        session = AuthorizedSession(credentials)
    except UploadStartFailure:
        raise
    except Exception as exc:
        raise UploadStartFailure(
            "MISSING_ENVIRONMENT_VARIABLE",
            safe_error_message(exc) or "Google Drive OAuth credentials are not configured.",
            status_code=503,
            stage="load_drive_credentials",
        ) from exc

    try:
        response = session.post(
            "https://www.googleapis.com/upload/drive/v3/files?uploadType=resumable&supportsAllDrives=true",
            headers={
                "Content-Type": "application/json; charset=UTF-8",
                "X-Upload-Content-Type": mime,
                "X-Upload-Content-Length": str(size_bytes),
            },
            json={"name": filename, "parents": [GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID]},
            timeout=30,
        )
        response.raise_for_status()
    except Exception as exc:
        status_code = getattr(getattr(exc, "response", None), "status_code", None) or getattr(locals().get("response", None), "status_code", None)
        raise UploadStartFailure(
            "STORAGE_UPLOAD_INIT_FAILED",
            f"Google Drive resumable upload session failed{f' with status {status_code}' if status_code else ''}.",
            status_code=502,
            stage="create_drive_resumable_upload",
        ) from exc

    upload_url = response.headers.get("Location")
    if not upload_url:
        raise UploadStartFailure(
            "STORAGE_UPLOAD_INIT_FAILED",
            "Google Drive did not return a resumable upload URL.",
            status_code=502,
            stage="create_drive_resumable_upload",
        )
    return upload_url


class IngestInput(BaseModel):
    source: str
    ref: str
    options: dict = {}


class WorkerJobInput(BaseModel):
    job_id: UUID


def _content_range_end(content_range: str) -> int | None:
    match = re.match(r"^bytes\s+(\d+)-(\d+)/(\d+|\*)$", str(content_range or "").strip(), re.I)
    if not match:
        return None
    return int(match.group(2)) + 1


def _upload_session_owner_fields(principal: Principal, database: SupabaseExecutor) -> dict:
    fields = {"owner_auth_user_id": principal.id}
    profile_id = scan_user_id(principal, database)
    if profile_id:
        fields["owner_user_id"] = profile_id
    return fields


def _upload_session_response(row: dict) -> dict:
    return {
        "upload_id": row["id"],
        "filename": row["original_filename"],
        "chunk_size": UPLOAD_CHUNK_SIZE_BYTES,
        "status": row.get("status", "upload_pending"),
        "received_size": row.get("received_size", 0),
        "drive_file": {"id": row.get("drive_file_id")} if row.get("drive_file_id") else None,
    }


def _insert_upload_session(database: SupabaseExecutor, principal: Principal, *, upload_id: str, filename: str, mime: str, size_bytes: int, upload_url: str) -> dict:
    now = datetime.now(timezone.utc)
    row = {
        "id": upload_id,
        **_upload_session_owner_fields(principal, database),
        "original_filename": filename,
        "content_type": mime,
        "total_size": size_bytes,
        "received_size": 0,
        "drive_resumable_url": upload_url,
        "status": "upload_pending",
        "expires_at": (now + timedelta(days=1)).isoformat(),
    }
    inserted = database.execute(lambda client: client.table(UPLOAD_SESSIONS_TABLE).insert(row).execute().data or [])
    if not inserted:
        raise HTTPException(status_code=500, detail="Unable to persist upload session.")
    return inserted[0]


def _load_upload_session(database: SupabaseExecutor, upload_id: str, principal: Principal) -> dict | None:
    def query(client):
        return (
            client.table(UPLOAD_SESSIONS_TABLE)
            .select("*")
            .eq("id", upload_id)
            .eq("owner_auth_user_id", principal.id)
            .maybe_single()
            .execute()
        )
    response = database.execute(query)
    return response.data if response else None


def _update_upload_session(database: SupabaseExecutor, upload_id: str, fields: dict) -> dict | None:
    payload = {**fields, "updated_at": datetime.now(timezone.utc).isoformat()}
    response = database.execute(
        lambda client: client.table(UPLOAD_SESSIONS_TABLE).update(payload).eq("id", upload_id).execute()
    )
    rows = response.data or []
    return rows[0] if rows else None


def _verify_completed_upload_for_drive_file(database: SupabaseExecutor, file_id: str, principal: Principal) -> None:
    response = database.execute(
        lambda client: client.table(UPLOAD_SESSIONS_TABLE)
        .select("id,status,drive_file_id")
        .eq("drive_file_id", file_id)
        .eq("owner_auth_user_id", principal.id)
        .limit(1)
        .execute()
    )
    rows = response.data or []
    if rows and rows[0].get("status") != "completed":
        raise HTTPException(status_code=409, detail="Upload has not completed yet.")


def _cloud_tasks_required_env() -> dict:
    values = {
        "project": CLOUD_TASKS_PROJECT_ID,
        "location": CLOUD_TASKS_LOCATION,
        "queue": CLOUD_TASKS_QUEUE,
        "worker_url": CLOUD_TASKS_WORKER_URL,
        "oidc_audience": CLOUD_TASKS_OIDC_AUDIENCE,
        "caller_service_account": CLOUD_TASKS_CALLER_SERVICE_ACCOUNT,
    }
    missing = [key for key, value in values.items() if not value]
    if missing:
        raise RuntimeError(f"Cloud Tasks environment is incomplete: {', '.join(missing)}")
    values["oidc_audience"] = _normalize_cloud_tasks_oidc_audience(str(values["oidc_audience"]))
    return values


def _normalize_cloud_tasks_oidc_audience(audience: str) -> str:
    parsed = urlparse(str(audience or "").strip())
    if (
        not parsed.scheme
        or parsed.scheme != "https"
        or not parsed.netloc
        or parsed.path not in {"", "/"}
        or parsed.params
        or parsed.query
        or parsed.fragment
    ):
        raise RuntimeError("CLOUD_TASKS_OIDC_AUDIENCE must be the HTTPS worker service origin.")
    hostname = parsed.hostname or ""
    if hostname in {"localhost", "127.0.0.1"} or hostname.endswith(".localhost"):
        raise RuntimeError("CLOUD_TASKS_OIDC_AUDIENCE must not be localhost.")
    if is_production() and not hostname.endswith(".run.app"):
        raise RuntimeError("Production CLOUD_TASKS_OIDC_AUDIENCE must use a run.app hostname.")
    return f"{parsed.scheme}://{parsed.netloc}"


def _cloud_task_name(client: Any, parent: str, job_id: str) -> str:
    safe_id = re.sub(r"[^A-Za-z0-9_-]+", "-", job_id).strip("-")
    return client.task_path(CLOUD_TASKS_PROJECT_ID, CLOUD_TASKS_LOCATION, CLOUD_TASKS_QUEUE, f"process-{safe_id}")


def enqueue_processing_task(job_id: str) -> dict:
    values = _cloud_tasks_required_env()
    try:
        from google.api_core.exceptions import AlreadyExists
        from google.cloud import tasks_v2
        from google.protobuf.duration_pb2 import Duration
    except Exception as exc:
        raise RuntimeError("google-cloud-tasks is not installed.") from exc

    client = tasks_v2.CloudTasksClient()
    parent = client.queue_path(values["project"], values["location"], values["queue"])
    payload = json.dumps({"job_id": job_id}, separators=(",", ":")).encode("utf-8")
    dispatch_deadline = Duration(seconds=CLOUD_TASKS_DISPATCH_DEADLINE_SECONDS)
    task = {
        "name": _cloud_task_name(client, parent, job_id),
        "http_request": {
            "http_method": tasks_v2.HttpMethod.POST,
            "url": values["worker_url"],
            "headers": {"Content-Type": "application/json"},
            "body": payload,
            "oidc_token": {
                "service_account_email": values["caller_service_account"],
                "audience": values["oidc_audience"],
            },
        },
        "dispatch_deadline": dispatch_deadline,
    }
    try:
        created = client.create_task(request={"parent": parent, "task": task})
        return {"task_name": created.name, "duplicate": False}
    except AlreadyExists:
        return {"task_name": task["name"], "duplicate": True}


def _claim_processing_job(database: SupabaseExecutor, job_id: str) -> dict:
    worker_id = f"{os.getenv('K_SERVICE', SERVICE_MODE)}:{os.getenv('K_REVISION', 'local')}:{os.getpid()}"
    response = database.execute(
        lambda client: client.rpc(
            "claim_processing_job",
            {"p_job_id": job_id, "p_lease_seconds": JOB_LEASE_SECONDS, "p_worker_id": worker_id},
        ).execute()
    )
    data = response.data if response else None
    if isinstance(data, list):
        return data[0] if data else {}
    return data or {}


@app.get("/api/health")
def health():
    return {
        "ok": True,
        "mode": "public_demo",
        "backend_build_version": BACKEND_BUILD_VERSION,
        **safe_startup_diagnostics(),
    }


def _drive_service():
    from googleapiclient.discovery import build
    return build("drive", "v3", credentials=oauth_drive_credentials(), cache_discovery=False)


def _drive_file_info(file_id: str) -> dict:
    return _drive_service().files().get(
        fileId=file_id,
        fields="id,name,mimeType,size,webViewLink,parents,trashed",
        supportsAllDrives=True,
    ).execute()


def _download_drive_file(file_id: str) -> bytes:
    from googleapiclient.http import MediaIoBaseDownload
    output = BytesIO()
    request = _drive_service().files().get_media(fileId=file_id, supportsAllDrives=True)
    downloader = MediaIoBaseDownload(output, request, chunksize=8 * 1024 * 1024)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return output.getvalue()


def _find_video_frame_scan(database: SupabaseExecutor, job_id: str, filename: str) -> str | None:
    response = database.execute(
        lambda client: client.table(SCAN_RESULTS_TABLE).select("id").eq("batch_id", job_id).eq("source_name", filename).maybe_single().execute()
    )
    return str(response.data["id"]) if response and response.data else None


def _clip_box(box: list[float], width: int, height: int) -> tuple[int, int, int, int]:
    x1 = max(0, min(width - 1, int(round(_coerce_float(box[0]))))) if width else 0
    y1 = max(0, min(height - 1, int(round(_coerce_float(box[1]))))) if height else 0
    x2 = max(x1 + 1, min(width, int(round(_coerce_float(box[2]))))) if width else 0
    y2 = max(y1 + 1, min(height, int(round(_coerce_float(box[3]))))) if height else 0
    return x1, y1, x2, y2


def _encode_detection_crop(frame, box: list[float]) -> bytes | None:
    import cv2
    height, width = frame.shape[:2]
    x1, y1, x2, y2 = _clip_box(box, width, height)
    crop = frame[y1:y2, x1:x2]
    if crop.size == 0:
        crop = frame
    ok, encoded = cv2.imencode(".jpg", crop)
    return encoded.tobytes() if ok else None


def _encode_frame_jpeg(frame) -> bytes | None:
    import cv2
    ok, encoded = cv2.imencode(".jpg", frame)
    return encoded.tobytes() if ok else None


def _mask_polygon(result, index: int) -> list | None:
    masks = getattr(result, "masks", None)
    if not masks:
        return None
    polygons = getattr(masks, "xyn", None) or getattr(masks, "xy", None)
    if polygons is None or index >= len(polygons):
        return None
    polygon = polygons[index]
    if hasattr(polygon, "tolist"):
        polygon = polygon.tolist()
    return polygon


def _result_track_observations(result, frame, frame_index: int, timestamp: float) -> list[dict]:
    boxes = getattr(result, "boxes", None)
    if boxes is None:
        return []
    names = getattr(result, "names", {}) or {}
    image_height, image_width = getattr(result, "orig_shape", frame.shape[:2])
    track_ids = getattr(boxes, "id", None)
    detections = []
    frame_bytes = _encode_frame_jpeg(frame)
    for index, box in enumerate(boxes):
        xyxy = [float(value) for value in box.xyxy[0].tolist()]
        confidence = float(box.conf[0])
        class_id = int(box.cls[0])
        material_name = str(names.get(class_id, f"class_{class_id}"))
        track_id = None
        if track_ids is not None:
            try:
                track_id = int(track_ids[index].item())
            except (AttributeError, IndexError, TypeError, ValueError):
                track_id = None
        detections.append({
            "track_id": track_id,
            "material_name": material_name,
            "category": material_category(material_name),
            "confidence": confidence,
            "bbox": [
                xyxy[0] / image_width if image_width else 0,
                xyxy[1] / image_height if image_height else 0,
                xyxy[2] / image_width if image_width else 0,
                xyxy[3] / image_height if image_height else 0,
            ],
            "bbox_percent": {
                "x": round((xyxy[0] / image_width) * 100, 2) if image_width else 0,
                "y": round((xyxy[1] / image_height) * 100, 2) if image_height else 0,
                "width": round(((xyxy[2] - xyxy[0]) / image_width) * 100, 2) if image_width else 0,
                "height": round(((xyxy[3] - xyxy[1]) / image_height) * 100, 2) if image_height else 0,
            },
            "best_box": {
                "xyxy": [round(value, 2) for value in xyxy],
                "frame": frame_index,
                "timestamp": round(timestamp, 3),
            },
            "mask": _mask_polygon(result, index),
            "frame_bytes": frame_bytes,
            "frame_width": int(image_width or frame.shape[1]),
            "frame_height": int(image_height or frame.shape[0]),
            "crop_bytes": _encode_detection_crop(frame, xyxy),
        })
    return detections


def _video_class_color(category: str) -> tuple[int, int, int]:
    palette = {
        "plastic": (40, 180, 99),
        "metal": (245, 158, 11),
        "glass": (14, 165, 233),
        "paper": (234, 179, 8),
        "cardboard": (168, 85, 247),
        "battery": (37, 99, 235),
        "textile": (236, 72, 153),
        "food_organics": (34, 197, 94),
        "general_trash": (239, 68, 68),
    }
    return palette.get(material_category(category), (20, 184, 166))


def _mask_to_points(mask, width: int, height: int):
    import numpy as np
    if not mask:
        return None
    points = []
    for point in mask:
        if not isinstance(point, (list, tuple)) or len(point) < 2:
            continue
        x = _coerce_float(point[0])
        y = _coerce_float(point[1])
        if x <= 1 and y <= 1:
            x *= width
            y *= height
        points.append([max(0, min(width - 1, int(round(x)))), max(0, min(height - 1, int(round(y))))])
    if len(points) < 3:
        return None
    return np.array(points, dtype=np.int32)


def _annotate_video_frame(frame, detections: list[dict], *, footer_count: int | None = None):
    import cv2
    height, width = frame.shape[:2]
    if not detections:
        return frame
    detections = _clean_preview_detections(detections, width, height)
    if not detections:
        return frame
    annotated = frame.copy()
    mask_layer = annotated.copy()
    has_mask = False
    line_width = max(2, round(min(width, height) / 360))
    font_scale = max(0.45, min(1.1, min(width, height) / 900))
    label_padding = max(4, round(line_width * 2))
    for detection in detections:
        category = material_category(detection.get("category") or detection.get("material_name"))
        color_rgb = _video_class_color(category)
        color = (color_rgb[2], color_rgb[1], color_rgb[0])
        try:
            box, _box_format = _detection_box_to_pixels(detection, width, height)
        except Exception as exc:
            _video_processing_log(
                "annotation_detection_box_invalid",
                image_width=width,
                image_height=height,
                track_id=detection.get("track_id"),
                category=category,
                error_type=type(exc).__name__,
                error=safe_error_message(exc),
            )
            continue
        x1, y1, x2, y2 = _clip_box(box, width, height)
        mask_points = _mask_to_points(detection.get("mask"), width, height)
        if mask_points is not None:
            cv2.fillPoly(mask_layer, [mask_points], color)
            has_mask = True
        cv2.rectangle(annotated, (x1, y1), (x2, y2), color, line_width)
        confidence = _coerce_float(detection.get("confidence"))
        track_id = detection.get("track_id")
        hazard = " | HAZARD" if CATEGORY_CLASS_MAP.get(category) == "contaminant" else ""
        label = f"{display_label(category)} | {confidence:.2f} | ID {track_id or '-'}{hazard}"
        font = cv2.FONT_HERSHEY_SIMPLEX
        (label_width, label_height), baseline = cv2.getTextSize(label, font, font_scale, line_width)
        label_width = min(label_width + label_padding * 2, width)
        label_height = label_height + baseline + label_padding * 2
        label_x = max(0, min(x1, width - label_width))
        label_y = y1 - label_height if y1 - label_height >= 0 else min(height - label_height, y2 + line_width)
        label_y = max(0, label_y)
        cv2.rectangle(annotated, (label_x, label_y), (label_x + label_width, label_y + label_height), color, -1)
        text_x = label_x + label_padding
        text_y = label_y + label_padding + label_height - baseline - label_padding
        cv2.putText(annotated, label, (text_x, text_y), font, font_scale, (255, 255, 255), max(1, line_width - 1), cv2.LINE_AA)
    if has_mask:
        annotated = cv2.addWeighted(mask_layer, 0.28, annotated, 0.72, 0)
    if footer_count is not None:
        label = f"{footer_count} object{'s' if footer_count != 1 else ''} detected"
        footer_height = max(26, round(height * 0.055))
        footer_y = max(0, height - footer_height)
        cv2.rectangle(annotated, (0, footer_y), (width, height), (4, 8, 6), -1)
        cv2.putText(
            annotated,
            label,
            (max(8, round(width * 0.025)), min(height - 8, footer_y + round(footer_height * 0.68))),
            cv2.FONT_HERSHEY_SIMPLEX,
            max(0.45, min(0.9, min(width, height) / 850)),
            (46, 204, 113),
            max(1, round(line_width * 0.8)),
            cv2.LINE_AA,
        )
    return annotated


class VideoDecodeError(RuntimeError):
    pass


def _video_processing_log(event: str, **fields) -> None:
    safe_fields = {}
    for key, value in fields.items():
        if key.lower() in {"token", "authorization", "api_key", "service_role_key", "password"}:
            safe_fields[key] = "[redacted]"
        elif isinstance(value, Path):
            safe_fields[key] = str(value)
        else:
            safe_fields[key] = value
    print(f"[video-processing] {event} {json.dumps(safe_fields, sort_keys=True, default=str)}")


def _worker_build_revision() -> str:
    return (
        os.getenv("APP_COMMIT_SHA")
        or os.getenv("K_REVISION")
        or os.getenv("GIT_COMMIT_SHA")
        or os.getenv("VERCEL_GIT_COMMIT_SHA")
        or "local"
    )


def _video_job_dir(scan_id: str) -> Path:
    safe_id = re.sub(r"[^A-Za-z0-9._-]+", "_", str(scan_id or "scan")).strip("._") or "scan"
    path = VIDEO_WORK_ROOT / safe_id
    path.mkdir(parents=True, exist_ok=True)
    return path


def _source_video_path(job_dir: Path, filename: str | None) -> Path:
    safe_name = safe_drive_filename(filename or "source.mp4")
    if not safe_name.lower().endswith(".mp4"):
        safe_name = f"{safe_name}.mp4"
    return job_dir / f"source-{safe_name}"


def _even_video_dimensions(width: int, height: int) -> tuple[int, int]:
    return max(2, int(width) - int(width) % 2), max(2, int(height) - int(height) % 2)


def _normalize_video_frame(frame, width: int, height: int):
    import cv2
    normalized = frame
    if normalized.shape[1] != width or normalized.shape[0] != height:
        normalized = cv2.resize(normalized, (width, height), interpolation=cv2.INTER_AREA)
    if len(normalized.shape) == 2:
        normalized = cv2.cvtColor(normalized, cv2.COLOR_GRAY2BGR)
    elif normalized.shape[2] == 4:
        normalized = cv2.cvtColor(normalized, cv2.COLOR_BGRA2BGR)
    return normalized


def _require_executable(name: str) -> str:
    path = shutil.which(name)
    if not path:
        raise RuntimeError(f"{name} is not installed or is not available on PATH")
    return path


def _encode_browser_mp4(input_path: str | Path, output_path: str | Path) -> list[str]:
    ffmpeg = _require_executable("ffmpeg")
    command = [
        ffmpeg,
        "-y",
        "-i", str(input_path),
        "-an",
        "-c:v", "libx264",
        "-preset", "medium",
        "-crf", "23",
        "-pix_fmt", "yuv420p",
        "-movflags", "+faststart",
        str(output_path),
    ]
    _video_processing_log("ffmpeg_started", ffmpeg_path=ffmpeg, arguments=command[1:], input_path=str(input_path), output_path=str(output_path))
    completed = subprocess.run(command, capture_output=True, text=True, timeout=900)
    output_exists = Path(output_path).exists()
    output_size = Path(output_path).stat().st_size if output_exists else 0
    _video_processing_log(
        "ffmpeg_completed",
        exit_code=completed.returncode,
        stderr=(completed.stderr or "")[-1200:],
        stdout=(completed.stdout or "")[-400:],
        output_exists=output_exists,
        output_size=output_size,
    )
    if completed.returncode != 0:
        raise RuntimeError(f"FFmpeg H.264 encoding failed with exit code {completed.returncode}: {(completed.stderr or completed.stdout or '').strip()[-1200:]}")
    if not output_exists or output_size <= 0:
        raise RuntimeError("FFmpeg H.264 encoding produced an empty output file.")
    return command


def _ffprobe_mp4(path: str | Path) -> dict:
    ffprobe = _require_executable("ffprobe")
    command = [
        ffprobe,
        "-v", "error",
        "-print_format", "json",
        "-show_format",
        "-show_streams",
        str(path),
    ]
    completed = subprocess.run(command, capture_output=True, text=True, timeout=60)
    if completed.returncode != 0:
        raise RuntimeError((completed.stderr or completed.stdout or "ffprobe validation failed").strip()[-500:])
    payload = json.loads(completed.stdout or "{}")
    streams = payload.get("streams") or []
    video = next((stream for stream in streams if stream.get("codec_type") == "video"), {})
    audio = next((stream for stream in streams if stream.get("codec_type") == "audio"), {})
    diagnostics = {
        "container": (payload.get("format") or {}).get("format_name"),
        "video_codec": video.get("codec_name"),
        "pixel_format": video.get("pix_fmt"),
        "dimensions": f"{video.get('width')}x{video.get('height')}" if video.get("width") and video.get("height") else None,
        "frame_rate": video.get("avg_frame_rate") or video.get("r_frame_rate"),
        "duration": (payload.get("format") or {}).get("duration") or video.get("duration"),
        "audio_codec": audio.get("codec_name") if audio else None,
    }
    print(f"[video-annotation] ffprobe {json.dumps(diagnostics, sort_keys=True)}")
    if diagnostics["video_codec"] != "h264" or diagnostics["pixel_format"] != "yuv420p":
        raise RuntimeError(f"Annotated MP4 is not browser-compatible H.264/yuv420p: {diagnostics}")
    return diagnostics


def _video_tracking_summary(tracked_objects: list[dict]) -> dict:
    counts_by_class: dict[str, int] = {}
    hazards = []
    public_objects = []
    for item in tracked_objects:
        category = str(item.get("category") or "unknown")
        counts_by_class[category] = counts_by_class.get(category, 0) + 1
        public = {key: value for key, value in item.items() if not key.startswith("_")}
        public_objects.append(public)
        if item.get("track_hazard_status") == "hazard" or item.get("contaminant_status") == "contaminated":
            hazards.append(public)
    return {
        "total_unique_objects": len(tracked_objects),
        "counts_by_class": counts_by_class,
        "hazards": hazards,
        "tracked_objects": public_objects,
        "result_kind": "tracked_video",
    }


def _track_time_overlap(first: dict, second: dict) -> bool:
    return not (
        _coerce_float(first.get("track_last_frame"), -1) < _coerce_float(second.get("track_first_frame"), 0)
        or _coerce_float(second.get("track_last_frame"), -1) < _coerce_float(first.get("track_first_frame"), 0)
    )


def _track_temporal_gap(first: dict, second: dict) -> int:
    return max(0, int(_coerce_float(second.get("track_first_frame"))) - int(_coerce_float(first.get("track_last_frame"))))


def _track_overlap_frames(first: dict, second: dict) -> int:
    start = max(int(_coerce_float(first.get("track_first_frame"))), int(_coerce_float(second.get("track_first_frame"))))
    end = min(int(_coerce_float(first.get("track_last_frame"))), int(_coerce_float(second.get("track_last_frame"))))
    return max(0, end - start + 1)


def _track_observations(track: dict) -> list[dict]:
    debug = track.get("track_debug") or {}
    return sorted(
        [item for item in debug.get("frame_observations", []) if isinstance(item, dict)],
        key=lambda item: int(_coerce_float(item.get("frame"))),
    )


def _observation_quality(observation: dict) -> dict:
    bbox_quality = _bbox_quality(observation.get("bbox"))
    confidence = _coerce_float(observation.get("confidence"))
    valid = bool(bbox_quality.get("valid")) and confidence >= VIDEO_DUPLICATE_MIN_OBSERVATION_CONFIDENCE
    return {
        **bbox_quality,
        "confidence": round(confidence, 4),
        "reliable": valid,
    }


def _track_evidence_quality(track: dict) -> dict:
    observations = _track_observations(track)
    qualities = [_observation_quality(item) for item in observations]
    reliable = [item for item in qualities if item.get("reliable")]
    boxes = [item.get("bbox") for item in observations if _observation_quality(item).get("reliable")]
    widths = [max(0.0, _coerce_float(box[2]) - _coerce_float(box[0])) for box in boxes if isinstance(box, list) and len(box) >= 4]
    heights = [max(0.0, _coerce_float(box[3]) - _coerce_float(box[1])) for box in boxes if isinstance(box, list) and len(box) >= 4]
    width_steps = [min(first, second) / max(first, second) for first, second in zip(widths, widths[1:]) if max(first, second) > 0]
    height_steps = [min(first, second) / max(first, second) for first, second in zip(heights, heights[1:]) if max(first, second) > 0]
    width_stability = _median(width_steps) if width_steps else (1.0 if widths else 0.0)
    height_stability = _median(height_steps) if height_steps else (1.0 if heights else 0.0)
    representative = _bbox_quality(track.get("best_bbox_norm"))
    near_full = representative.get("reason") == "bbox near full frame"
    edge_fraction = (
        sum(bool(item.get("touches_edge")) for item in qualities) / len(qualities)
        if qualities else 0.0
    )
    reliable_track = (
        len(reliable) >= VIDEO_DUPLICATE_MIN_RELIABLE_OBSERVATIONS
        and min(width_stability, height_stability) >= 0.35
    )
    return {
        "observation_count": len(observations),
        "valid_observation_count": len(reliable),
        "box_stability": round(min(width_stability, height_stability), 4),
        "edge_fraction": round(edge_fraction, 4),
        "near_full_frame": near_full,
        "partial_or_broad": near_full or bool(representative.get("touches_edge")) or edge_fraction >= 0.5,
        "reliable_track": reliable_track,
        "representative": representative,
    }


def _track_class_votes(track: dict) -> dict[str, float]:
    debug = track.get("track_debug") or {}
    votes = {str(key): _coerce_float(value) for key, value in (debug.get("class_votes") or {}).items()}
    if not votes and track.get("category"):
        votes[str(track["category"])] = _coerce_float(track.get("track_max_confidence") or track.get("confidence"), 1.0)
    return votes


def _class_vote_similarity(first: dict, second: dict) -> float:
    first_votes = _track_class_votes(first)
    second_votes = _track_class_votes(second)
    keys = set(first_votes) | set(second_votes)
    if not keys:
        return 0.0
    dot = sum(first_votes.get(key, 0.0) * second_votes.get(key, 0.0) for key in keys)
    first_norm = math.sqrt(sum(value * value for value in first_votes.values()))
    second_norm = math.sqrt(sum(value * value for value in second_votes.values()))
    return dot / (first_norm * second_norm) if first_norm and second_norm else 0.0


def _track_endpoint_boxes(first: dict, second: dict) -> tuple[list[float] | None, list[float] | None]:
    ordered = sorted([first, second], key=lambda item: int(_coerce_float(item.get("track_first_frame"))))
    previous, current = ordered
    previous_observations = _track_observations(previous)
    current_observations = _track_observations(current)
    previous_box = (previous_observations[-1].get("bbox") if previous_observations else previous.get("best_bbox_norm")) or []
    current_box = (current_observations[0].get("bbox") if current_observations else current.get("best_bbox_norm")) or []
    return previous_box, current_box


def _trajectory_distance(previous: dict, current: dict) -> dict:
    previous_raw = [item for item in previous.get("track_path") or [] if isinstance(item, dict)]
    current_raw = [item for item in current.get("track_path") or [] if isinstance(item, dict)]
    previous_scene = [
        item for item in (previous.get("track_debug") or {}).get("stabilized_track_path") or []
        if isinstance(item, dict) and str(item.get("motion_status") or "ok") in {"ok", "reference"}
    ]
    current_scene = [
        item for item in (current.get("track_debug") or {}).get("stabilized_track_path") or []
        if isinstance(item, dict) and str(item.get("motion_status") or "ok") in {"ok", "reference"}
    ]

    def window_distance(first_path: list[dict], second_path: list[dict]) -> float | None:
        pairs = []
        for first_item in first_path[-5:]:
            for second_item in second_path[:5]:
                if first_item.get("x") is None or first_item.get("y") is None or second_item.get("x") is None or second_item.get("y") is None:
                    continue
                pairs.append(math.hypot(
                    _coerce_float(first_item.get("x")) - _coerce_float(second_item.get("x")),
                    _coerce_float(first_item.get("y")) - _coerce_float(second_item.get("y")),
                ))
        return _median(sorted(pairs)[: min(5, len(pairs))]) if pairs else None

    raw_distance = window_distance(previous_raw, current_raw)
    if raw_distance is None:
        previous_center = previous.get("track_end_center") or {}
        current_center = current.get("track_start_center") or {}
        raw_distance = math.hypot(
            _coerce_float(previous_center.get("x")) - _coerce_float(current_center.get("x")),
            _coerce_float(previous_center.get("y")) - _coerce_float(current_center.get("y")),
        )
    stabilized_distance = window_distance(previous_scene, current_scene)
    if stabilized_distance is None:
        previous_center = previous.get("track_end_scene_center") or {}
        current_center = current.get("track_start_scene_center") or {}
        if all(center.get("x") is not None and center.get("y") is not None for center in (previous_center, current_center)):
            stabilized_distance = math.hypot(
                _coerce_float(previous_center.get("x")) - _coerce_float(current_center.get("x")),
                _coerce_float(previous_center.get("y")) - _coerce_float(current_center.get("y")),
            )
    return {
        "raw": raw_distance,
        "stabilized": stabilized_distance,
        "selected": stabilized_distance if stabilized_distance is not None else raw_distance,
        "stabilized_used": stabilized_distance is not None,
        "raw_observations": [len(previous_raw), len(current_raw)],
        "stabilized_observations": [len(previous_scene), len(current_scene)],
    }


def _track_overlap_iou(first: dict, second: dict) -> dict:
    first_by_frame = {
        int(_coerce_float(item.get("frame"))): item
        for item in _track_observations(first)
        if item.get("frame") is not None
    }
    second_by_frame = {
        int(_coerce_float(item.get("frame"))): item
        for item in _track_observations(second)
        if item.get("frame") is not None
    }
    values = []
    reliable_values = []
    for frame in sorted(set(first_by_frame) & set(second_by_frame)):
        first_item, second_item = first_by_frame[frame], second_by_frame[frame]
        value = _bbox_iou(first_item.get("bbox"), second_item.get("bbox"))
        values.append(value)
        if _observation_quality(first_item).get("reliable") and _observation_quality(second_item).get("reliable"):
            reliable_values.append(value)
    if not values:
        return {"max": None, "avg": None, "median": None, "reliable_pair_count": 0, "reliable_median": None}
    return {
        "max": round(max(values), 4),
        "avg": round(sum(values) / len(values), 4),
        "median": round(_median(values) or 0.0, 4),
        "reliable_pair_count": len(reliable_values),
        "reliable_median": round(_median(reliable_values), 4) if reliable_values else None,
    }


def _track_appearance_score(first: dict, second: dict) -> float | None:
    compared = _best_appearance_similarity(first, second)
    if compared.get("score") is not None:
        return _coerce_float(compared.get("score"))
    first_bytes = first.get("_best_crop_bytes")
    second_bytes = second.get("_best_crop_bytes")
    if first_bytes and second_bytes:
        return 1.0 if hashlib.sha256(first_bytes).hexdigest() == hashlib.sha256(second_bytes).hexdigest() else 0.0
    return None


def _track_verified(track: dict) -> bool:
    debug = track.get("track_debug") or {}
    values = {
        str(track.get("review_status") or "").lower(),
        str(track.get("verification_status") or "").lower(),
        str(track.get("overall_status") or "").lower(),
        str(debug.get("review_status") or "").lower(),
        str(debug.get("verification_status") or "").lower(),
    }
    return bool(debug.get("human_verified")) or bool(track.get("human_verified")) or "verified" in values or "confirmed" in values


def _duplicate_canonical_sort_key(track: dict) -> tuple[int, int, float, float, int, float, str]:
    preview_quality = _bbox_quality(track.get("best_bbox_norm"))
    evidence_quality = _track_evidence_quality(track)
    return (
        1 if _track_verified(track) else 0,
        1 if preview_quality.get("valid") else 0,
        _coerce_float(preview_quality.get("score")),
        _coerce_float(track.get("track_max_confidence") or track.get("confidence")),
        int(_coerce_float(evidence_quality.get("valid_observation_count"))),
        _coerce_float(track.get("track_avg_confidence")),
        "".join(chr(255 - ord(ch)) for ch in str(track.get("stable_object_id") or track.get("object_uid") or track.get("track_id") or "")),
    )


def _duplicate_evidence(first: dict, second: dict) -> dict:
    ordered = sorted(
        [first, second],
        key=lambda item: (int(_coerce_float(item.get("track_first_frame"))), _object_id(item)),
    )
    previous, current = ordered
    gap = 0 if _track_time_overlap(previous, current) else _track_temporal_gap(previous, current)
    overlap = _track_overlap_frames(previous, current)
    trajectory = _trajectory_distance(previous, current)
    raw_distance = _coerce_float(trajectory.get("raw"))
    stabilized_distance = trajectory.get("stabilized")
    distance = _coerce_float(trajectory.get("selected"))
    previous_box, current_box = _track_endpoint_boxes(previous, current)
    iou = _bbox_iou(previous_box, current_box)
    overlap_iou = _track_overlap_iou(previous, current)
    overlap_iou_max = _coerce_float(overlap_iou.get("max"), 0.0)
    overlap_iou_reliable_median = _coerce_float(overlap_iou.get("reliable_median"), 0.0)
    reliable_overlap_pairs = int(_coerce_float(overlap_iou.get("reliable_pair_count")))
    simultaneously_visible = overlap > 0
    width_ratio = min(_coerce_float(first.get("track_avg_width")), _coerce_float(second.get("track_avg_width"))) / max(_coerce_float(first.get("track_avg_width")), _coerce_float(second.get("track_avg_width")), 1e-9)
    height_ratio = min(_coerce_float(first.get("track_avg_height")), _coerce_float(second.get("track_avg_height"))) / max(_coerce_float(first.get("track_avg_height")), _coerce_float(second.get("track_avg_height")), 1e-9)
    aspect_ratio = min(_coerce_float(first.get("track_avg_aspect_ratio")), _coerce_float(second.get("track_avg_aspect_ratio"))) / max(_coerce_float(first.get("track_avg_aspect_ratio")), _coerce_float(second.get("track_avg_aspect_ratio")), 1e-9)
    size_ratio = min(width_ratio, height_ratio)
    area_ratio = min(
        _coerce_float(first.get("track_avg_width")) * _coerce_float(first.get("track_avg_height")),
        _coerce_float(second.get("track_avg_width")) * _coerce_float(second.get("track_avg_height")),
    ) / max(
        _coerce_float(first.get("track_avg_width")) * _coerce_float(first.get("track_avg_height")),
        _coerce_float(second.get("track_avg_width")) * _coerce_float(second.get("track_avg_height")),
        1e-9,
    )
    appearance_detail = _appearance_evidence(first, second)
    appearance = _coerce_float(appearance_detail.get("score")) if appearance_detail.get("score") is not None else _track_appearance_score(first, second)
    class_vote_similarity = _class_vote_similarity(first, second)
    same_category = str(first.get("category")) == str(second.get("category"))
    first_track_quality = _track_evidence_quality(first)
    second_track_quality = _track_evidence_quality(second)
    first_quality = first_track_quality["representative"]
    second_quality = second_track_quality["representative"]
    partial_pair = first_track_quality.get("partial_or_broad") or second_track_quality.get("partial_or_broad")
    best_appearance = appearance_detail.get("best_valid_similarity")
    appearance_strong = best_appearance is not None and _coerce_float(best_appearance) >= VIDEO_DUPLICATE_APPEARANCE_SIMILARITY
    appearance_positive = appearance_strong and int(_coerce_float(appearance_detail.get("agreeing_pair_count"))) > 0
    appearance_mismatch = best_appearance is not None and _coerce_float(best_appearance) < VIDEO_DUPLICATE_MIN_APPEARANCE_SIMILARITY
    reliable_coexistence = (
        first_track_quality.get("reliable_track")
        and second_track_quality.get("reliable_track")
        and reliable_overlap_pairs >= VIDEO_DUPLICATE_MIN_RELIABLE_OBSERVATIONS
    )
    simultaneous_low_iou_separate = (
        overlap >= VIDEO_DUPLICATE_MEANINGFUL_OVERLAP_FRAMES
        and reliable_coexistence
        and overlap_iou_reliable_median < VIDEO_DUPLICATE_OVERLAP_IOU
    )
    stable_tracks_coexist = (
        overlap >= VIDEO_DUPLICATE_MEANINGFUL_OVERLAP_FRAMES
        and reliable_coexistence
        and int(_coerce_float(first_track_quality.get("valid_observation_count"))) >= VIDEO_DUPLICATE_STABLE_TRACK_FRAMES
        and int(_coerce_float(second_track_quality.get("valid_observation_count"))) >= VIDEO_DUPLICATE_STABLE_TRACK_FRAMES
    )
    quality_handover = (
        0 < overlap <= VIDEO_DUPLICATE_STRONG_OVERLAP_MAX_FRAMES
        and reliable_overlap_pairs >= VIDEO_DUPLICATE_MIN_RELIABLE_OBSERVATIONS
        and overlap_iou_reliable_median >= VIDEO_DUPLICATE_HANDOVER_MEDIAN_IOU
        and (same_category or class_vote_similarity >= 0.5 or appearance_positive)
        and not appearance_mismatch
    )
    strong_overlap_switch = (
        0 < overlap <= VIDEO_DUPLICATE_STRONG_OVERLAP_MAX_FRAMES
        and (
            quality_handover
            or (
                overlap_iou_max >= VIDEO_DUPLICATE_STRONG_OVERLAP_IOU
                and min(size_ratio, aspect_ratio) >= VIDEO_DUPLICATE_STRONG_SIZE_RATIO
                and (same_category or class_vote_similarity > 0 or appearance_positive)
            )
        )
    )
    near_full_bridge = (
        not overlap
        and gap <= min(15, VIDEO_DUPLICATE_WEAK_FRAGMENT_MAX_GAP)
        and (first_track_quality.get("near_full_frame") or second_track_quality.get("near_full_frame"))
        and class_vote_similarity >= 0.55
        and not appearance_mismatch
    )
    clipped_class_bridge = (
        not overlap
        and gap <= VIDEO_DUPLICATE_WEAK_FRAGMENT_MAX_GAP
        and partial_pair
        and same_category
        and class_vote_similarity >= 0.80
        and not appearance_mismatch
    )
    clipped_appearance_bridge = (
        not overlap
        and gap <= VIDEO_DUPLICATE_APPEARANCE_FRAGMENT_MAX_GAP
        and partial_pair
        and appearance_positive
        and class_vote_similarity >= 0.75
        and not appearance_mismatch
    )
    partial_fragment_supported = near_full_bridge or clipped_class_bridge or clipped_appearance_bridge
    reject_reason = None
    if simultaneous_low_iou_separate:
        reject_reason = f"simultaneous low-IoU boxes for {overlap} frames"
    elif stable_tracks_coexist and not strong_overlap_switch:
        reject_reason = f"two stable tracks coexist for {overlap} frames"
    elif overlap >= VIDEO_DUPLICATE_MEANINGFUL_OVERLAP_FRAMES:
        reject_reason = f"temporal overlap {overlap} is meaningful"
    elif overlap > VIDEO_DUPLICATE_MAX_OVERLAP_FRAMES and not strong_overlap_switch:
        reject_reason = f"temporal overlap {overlap} exceeds {VIDEO_DUPLICATE_MAX_OVERLAP_FRAMES}"
    elif gap > VIDEO_DUPLICATE_MAX_GAP:
        reject_reason = f"temporal gap {gap} exceeds {VIDEO_DUPLICATE_MAX_GAP}"
    elif appearance_mismatch:
        reject_reason = "appearance differs"
    elif partial_pair and not partial_fragment_supported and not strong_overlap_switch:
        reject_reason = "weak fragment lacks bridge evidence"
    elif min(size_ratio, aspect_ratio) < VIDEO_DUPLICATE_SIZE_RATIO and not partial_fragment_supported and not strong_overlap_switch:
        reject_reason = "size/aspect mismatch"
    elif overlap and not strong_overlap_switch and (distance > VIDEO_DUPLICATE_STRONG_CENTER_DISTANCE or max(iou, overlap_iou_max) < VIDEO_DUPLICATE_OVERLAP_IOU):
        reject_reason = "overlap lacks strong spatial agreement"
    elif not overlap and not partial_fragment_supported and distance > VIDEO_DUPLICATE_CENTER_DISTANCE and iou < VIDEO_DUPLICATE_IOU:
        reject_reason = "trajectory discontinuity"
    elif not same_category and class_vote_similarity <= 0 and not (
        strong_overlap_switch
        or (appearance_positive and distance <= VIDEO_DUPLICATE_STRONG_CENTER_DISTANCE and min(size_ratio, aspect_ratio) >= VIDEO_DUPLICATE_STRONG_SIZE_RATIO)
    ):
        reject_reason = "class votes incompatible without strong tracking/appearance evidence"
    accepted = reject_reason is None
    acceptance_score = 0.0
    if accepted:
        acceptance_score += max(0.0, 1.0 - min(distance, 1.0)) * 0.35
        acceptance_score += max(iou, overlap_iou_max) * 0.25
        acceptance_score += (appearance or 0.0) * 0.25
        acceptance_score += min(size_ratio, aspect_ratio, area_ratio) * 0.10
        acceptance_score += class_vote_similarity * 0.05
    return {
        "object_ids": [str(first.get("stable_object_id") or first.get("object_uid")), str(second.get("stable_object_id") or second.get("object_uid"))],
        "source_track_ids": [first.get("source_track_ids"), second.get("source_track_ids")],
        "categories": [first.get("category"), second.get("category")],
        "confidence_values": [first.get("track_max_confidence") or first.get("confidence"), second.get("track_max_confidence") or second.get("confidence")],
        "category_compatible": same_category or class_vote_similarity > 0,
        "class_vote_similarity": round(class_vote_similarity, 4),
        "temporal_gap": gap,
        "simultaneously_visible": simultaneously_visible,
        "overlap_frames": overlap,
        "trajectory_distance": round(distance, 4),
        "centre_distance": round(distance, 4),
        "raw_trajectory_distance": round(raw_distance, 4),
        "stabilized_trajectory_distance": round(stabilized_distance, 4) if stabilized_distance is not None else None,
        "stabilized_coordinates_used": bool(trajectory.get("stabilized_used")),
        "trajectory_observation_counts": {
            "raw": trajectory.get("raw_observations"),
            "stabilized": trajectory.get("stabilized_observations"),
        },
        "iou": round(iou, 4),
        "overlapping_frame_iou": overlap_iou,
        "size_ratio": round(size_ratio, 4),
        "area_ratio": round(area_ratio, 4),
        "aspect_ratio": round(aspect_ratio, 4),
        "width_ratio": round(width_ratio, 4),
        "height_ratio": round(height_ratio, 4),
        "preview_quality": [first_quality, second_quality],
        "observation_quality": [first_track_quality, second_track_quality],
        "appearance_status": appearance_detail.get("status", "appearance unavailable"),
        "appearance_score": appearance,
        "appearance_best_valid_similarity": appearance_detail.get("best_valid_similarity"),
        "appearance_strongest_median_similarity": appearance_detail.get("strongest_median_similarity"),
        "appearance_valid_pair_count": appearance_detail.get("valid_pair_count"),
        "appearance_agreeing_pair_count": appearance_detail.get("agreeing_pair_count"),
        "appearance_conflicting_pair_count": appearance_detail.get("conflicting_pair_count"),
        "appearance_strong": appearance_strong,
        "appearance_positive": appearance_positive,
        "strong_overlap_switch": strong_overlap_switch,
        "quality_handover": quality_handover,
        "partial_fragment_supported": partial_fragment_supported,
        "partial_fragment_bridge": {
            "near_full": near_full_bridge,
            "clipped_class": clipped_class_bridge,
            "clipped_appearance": clipped_appearance_bridge,
        },
        "simultaneous_low_iou_separate": simultaneous_low_iou_separate,
        "stable_tracks_coexist": stable_tracks_coexist,
        "association_score": round(acceptance_score, 4),
        "accepted": accepted,
        "final_reason": "same physical object evidence accepted" if accepted else reject_reason,
        "thresholds": {
            "max_gap": VIDEO_DUPLICATE_MAX_GAP,
            "max_overlap_frames": VIDEO_DUPLICATE_MAX_OVERLAP_FRAMES,
            "strong_overlap_max_frames": VIDEO_DUPLICATE_STRONG_OVERLAP_MAX_FRAMES,
            "meaningful_overlap_frames": VIDEO_DUPLICATE_MEANINGFUL_OVERLAP_FRAMES,
            "center_distance": VIDEO_DUPLICATE_CENTER_DISTANCE,
            "strong_center_distance": VIDEO_DUPLICATE_STRONG_CENTER_DISTANCE,
            "iou": VIDEO_DUPLICATE_IOU,
            "overlap_iou": VIDEO_DUPLICATE_OVERLAP_IOU,
            "strong_overlap_iou": VIDEO_DUPLICATE_STRONG_OVERLAP_IOU,
            "handover_median_iou": VIDEO_DUPLICATE_HANDOVER_MEDIAN_IOU,
            "size_ratio": VIDEO_DUPLICATE_SIZE_RATIO,
            "strong_size_ratio": VIDEO_DUPLICATE_STRONG_SIZE_RATIO,
            "appearance_similarity": VIDEO_DUPLICATE_APPEARANCE_SIMILARITY,
            "min_appearance_similarity": VIDEO_DUPLICATE_MIN_APPEARANCE_SIMILARITY,
            "min_observation_confidence": VIDEO_DUPLICATE_MIN_OBSERVATION_CONFIDENCE,
            "min_reliable_observations": VIDEO_DUPLICATE_MIN_RELIABLE_OBSERVATIONS,
            "weak_fragment_max_gap": VIDEO_DUPLICATE_WEAK_FRAGMENT_MAX_GAP,
            "appearance_fragment_max_gap": VIDEO_DUPLICATE_APPEARANCE_FRAGMENT_MAX_GAP,
        },
    }


def _merge_duplicate_group(group: list[dict], decisions: list[dict]) -> dict:
    canonical = max(group, key=_duplicate_canonical_sort_key)
    merged = canonical
    for track in sorted(group, key=lambda item: str(item.get("stable_object_id") or item.get("object_uid") or item.get("track_id") or "")):
        if track is canonical:
            continue
        merged = _merge_two_tracks(merged, track)
    discarded_ids = sorted(str(item.get("stable_object_id") or item.get("object_uid")) for item in group if item is not canonical)
    merged["stable_object_id"] = canonical.get("stable_object_id")
    merged["object_uid"] = canonical.get("object_uid")
    for key in ("persisted_scan_id", "persisted_material_id"):
        if canonical.get(key):
            merged[key] = canonical[key]
    if _track_verified(canonical):
        merged["review_status"] = canonical.get("review_status", "verified")
    debug = merged.setdefault("track_debug", {})
    debug["deduplicated_object_ids"] = discarded_ids
    debug["duplicate_reconciliation"] = decisions
    debug["physical_object_reconciliation"] = {
        "cluster_object_ids": sorted(str(item.get("stable_object_id") or item.get("object_uid")) for item in group),
        "accepted_edges": decisions,
        "canonical_selection": "human_verified > valid_preview > preview_quality > track_max_confidence > valid_observation_count > track_avg_confidence > stable_id",
    }
    debug["canonical_object_id"] = canonical.get("stable_object_id") or canonical.get("object_uid")
    return merged


def _association_hard_blocker(evidence: dict) -> bool:
    reason = str(evidence.get("final_reason") or "")
    return bool(
        evidence.get("simultaneous_low_iou_separate")
        or evidence.get("stable_tracks_coexist")
        or reason.startswith("two stable tracks coexist")
        or reason.startswith("simultaneous low-IoU boxes")
        or reason == "appearance differs"
        or reason == "overlap lacks strong spatial agreement"
    )


def _edge_key(evidence: dict) -> tuple[float, str]:
    return (
        _coerce_float(evidence.get("association_score")),
        "|".join(str(item) for item in evidence.get("object_ids") or []),
    )


def _object_id(track: dict) -> str:
    return str(track.get("stable_object_id") or track.get("object_uid") or track.get("track_id") or "")


def _validated_physical_components(objects: list[dict], evaluated: list[dict], batch_id: str) -> tuple[list[list[int]], list[dict]]:
    index_by_id = {_object_id(item): index for index, item in enumerate(objects)}
    blocker_pairs = {
        frozenset(str(item) for item in evidence.get("object_ids") or [])
        for evidence in evaluated
        if not evidence.get("accepted") and _association_hard_blocker(evidence)
    }
    accepted_edges = sorted(
        [item for item in evaluated if item.get("accepted")],
        key=_edge_key,
        reverse=True,
    )
    parent = list(range(len(objects)))

    def find(index: int) -> int:
        while parent[index] != index:
            parent[index] = parent[parent[index]]
            index = parent[index]
        return index

    def members_after_union(first: int, second: int) -> list[int]:
        roots = {find(first), find(second)}
        return [index for index in range(len(objects)) if find(index) in roots]

    def has_cluster_blocker(indices: list[int]) -> bool:
        ids = [_object_id(objects[index]) for index in indices]
        for i, first_id in enumerate(ids):
            for second_id in ids[i + 1:]:
                if frozenset({first_id, second_id}) in blocker_pairs:
                    return True
        return False

    split_events = []
    for edge in accepted_edges:
        ids = [str(item) for item in edge.get("object_ids") or []]
        if len(ids) != 2 or ids[0] not in index_by_id or ids[1] not in index_by_id:
            continue
        first, second = index_by_id[ids[0]], index_by_id[ids[1]]
        if find(first) == find(second):
            continue
        candidate_members = members_after_union(first, second)
        if has_cluster_blocker(candidate_members):
            event = {
                "object_ids": ids,
                "reason": "accepted edge skipped because connected component would contain hard separation evidence",
                "association_score": edge.get("association_score"),
            }
            split_events.append(event)
            _video_processing_log("physical_cluster_split", batch_id=batch_id, **event)
            continue
        root_first, root_second = find(first), find(second)
        parent[max(root_first, root_second)] = min(root_first, root_second)
        _video_processing_log("physical_association_accepted", batch_id=batch_id, **edge)

    groups_by_root: dict[int, list[int]] = {}
    for index in range(len(objects)):
        groups_by_root.setdefault(find(index), []).append(index)
    components = [sorted(indices, key=lambda idx: _object_id(objects[idx])) for indices in groups_by_root.values()]
    components.sort(key=lambda indices: min(_object_id(objects[index]) for index in indices))
    for indices in components:
        _video_processing_log(
            "physical_cluster_validated",
            batch_id=batch_id,
            object_ids=[_object_id(objects[index]) for index in indices],
            member_count=len(indices),
        )
    return components, split_events


def reconcile_duplicate_tracked_objects(logical_objects: list[dict], batch_id: str, *, dry_run: bool = False) -> tuple[list[dict], dict]:
    objects = sorted(logical_objects, key=lambda item: str(item.get("stable_object_id") or item.get("object_uid") or item.get("track_id") or ""))
    evaluated = []
    for i, first in enumerate(objects):
        for j in range(i + 1, len(objects)):
            evidence = _duplicate_evidence(first, objects[j])
            evaluated.append(evidence)
            _video_processing_log("physical_association_candidate", batch_id=batch_id, **evidence)
            if not evidence["accepted"]:
                _video_processing_log("physical_association_rejected", batch_id=batch_id, **evidence)

    components, split_events = _validated_physical_components(objects, evaluated, batch_id)
    output = []
    confirmed_groups = []
    rejected = [item for item in evaluated if not item["accepted"]]
    for indices in components:
        group = [objects[index] for index in indices]
        if len(group) == 1:
            output.append(group[0])
            continue
        group_object_ids = [str(item.get("stable_object_id") or item.get("object_uid")) for item in group]
        group_decisions = [item for item in evaluated if item["accepted"] and set(item["object_ids"]).issubset(set(group_object_ids))]
        merged = _merge_duplicate_group(group, group_decisions)
        confirmed_groups.append({
            "object_ids": group_object_ids,
            "source_track_ids": [item.get("source_track_ids") for item in group],
            "categories": [item.get("category") for item in group],
            "selected_canonical_object": merged.get("stable_object_id") or merged.get("object_uid"),
            "evidence": group_decisions,
        })
        _video_processing_log(
            "physical_canonical_selected",
            batch_id=batch_id,
            logical_object_ids=group_object_ids,
            selected_canonical_object=merged.get("stable_object_id") or merged.get("object_uid"),
            representative_quality=_bbox_quality(merged.get("best_bbox_norm")),
            confidence=merged.get("track_max_confidence") or merged.get("confidence"),
        )
        output.append(merged)

    output.sort(key=lambda item: int(_coerce_float(item.get("track_first_frame"))))
    for index, item in enumerate(output, start=1):
        item.setdefault("track_debug", {})["duplicate_reconciliation_completed"] = True
        item.setdefault("track_debug", {})["physical_reconciliation_completed"] = True
        item.setdefault("track_debug", {}).setdefault("physical_object_reconciliation", {
            "cluster_object_ids": [_object_id(item)],
            "accepted_edges": [],
            "canonical_selection": "singleton final physical cluster",
        })
        item["object_uid"] = item.get("object_uid") or f"{batch_id}-object-{index:04d}"
        item["stable_object_id"] = item.get("stable_object_id") or item["object_uid"]
    report = {
        "batch_id": batch_id,
        "input_count": len(logical_objects),
        "output_count": len(output),
        "evaluated_candidates": evaluated,
        "confirmed_groups": confirmed_groups,
        "rejected_candidates": rejected,
        "cluster_split_events": split_events,
        "dry_run": dry_run,
    }
    _video_processing_log(
        "physical_reconciliation_completed",
        batch_id=batch_id,
        input_count=len(logical_objects),
        output_count=len(output),
        final_cluster_count=len(output),
        groups_confirmed=len(confirmed_groups),
        dry_run=dry_run,
    )
    return output, report


def _track_merge_score(first: dict, second: dict) -> tuple[bool, str]:
    if first.get("category") != second.get("category"):
        return False, "class mismatch"
    if _track_time_overlap(first, second):
        return False, "tracks overlap in time"
    ordered = sorted([first, second], key=lambda item: int(_coerce_float(item.get("track_first_frame"))))
    previous, current = ordered
    gap = _track_temporal_gap(previous, current)
    if gap > VIDEO_LOGICAL_MERGE_MAX_GAP:
        return False, f"temporal gap {gap} exceeds {VIDEO_LOGICAL_MERGE_MAX_GAP}"
    previous_end = previous.get("track_end_center") or {}
    current_start = current.get("track_start_center") or {}
    distance = math.hypot(
        _coerce_float(previous_end.get("x")) - _coerce_float(current_start.get("x")),
        _coerce_float(previous_end.get("y")) - _coerce_float(current_start.get("y")),
    )
    if distance > VIDEO_LOGICAL_MERGE_CENTER_DISTANCE:
        return False, f"centre distance {distance:.3f} exceeds {VIDEO_LOGICAL_MERGE_CENTER_DISTANCE}"
    width_a, width_b = _coerce_float(previous.get("track_avg_width")), _coerce_float(current.get("track_avg_width"))
    height_a, height_b = _coerce_float(previous.get("track_avg_height")), _coerce_float(current.get("track_avg_height"))
    aspect_a, aspect_b = _coerce_float(previous.get("track_avg_aspect_ratio")), _coerce_float(current.get("track_avg_aspect_ratio"))
    width_ratio = min(width_a, width_b) / max(width_a, width_b) if max(width_a, width_b) > 0 else 1
    height_ratio = min(height_a, height_b) / max(height_a, height_b) if max(height_a, height_b) > 0 else 1
    aspect_ratio = min(aspect_a, aspect_b) / max(aspect_a, aspect_b) if max(aspect_a, aspect_b) > 0 else 1
    if min(width_ratio, height_ratio, aspect_ratio) < VIDEO_LOGICAL_MERGE_SIZE_RATIO:
        return False, "size/aspect mismatch"
    return True, f"class compatible; gap={gap}; centre_distance={distance:.3f}; size ratios ok"


def _merge_two_tracks(first: dict, second: dict) -> dict:
    tracks = sorted([first, second], key=lambda item: int(_coerce_float(item.get("track_first_frame"))))
    primary = max(tracks, key=_duplicate_canonical_sort_key)
    representative = max(
        tracks,
        key=lambda item: (
            1 if _bbox_quality(item.get("best_bbox_norm")).get("valid") else 0,
            _coerce_float(_bbox_quality(item.get("best_bbox_norm")).get("score")),
            _coerce_float(item.get("track_max_confidence") or item.get("confidence")),
            int(_coerce_float(item.get("track_frame_count"))),
            "".join(chr(255 - ord(ch)) for ch in str(item.get("stable_object_id") or item.get("object_uid") or "")),
        ),
    )
    merged_observations = []
    merged_path = []
    merged_scene_path = []
    appearance_fingerprints = []
    class_votes: dict[str, float] = {}
    source_track_ids = []
    for track in tracks:
        source_track_ids.extend([str(item) for item in track.get("source_track_ids") or str(track.get("track_id") or "").split(",") if item])
        debug = track.get("track_debug") or {}
        merged_observations.extend(pl for pl in debug.get("frame_observations", []) if isinstance(pl, dict))
        merged_path.extend(pl for pl in track.get("track_path", []) if isinstance(pl, dict))
        merged_scene_path.extend(pl for pl in debug.get("stabilized_track_path", []) if isinstance(pl, dict))
        appearance_fingerprints.extend(pl for pl in debug.get("appearance_fingerprints", []) if isinstance(pl, dict))
        for category, value in (debug.get("class_votes") or {}).items():
            class_votes[category] = class_votes.get(category, 0.0) + _coerce_float(value)
    source_track_ids = sorted(set(source_track_ids), key=str)
    first_frame = min(int(_coerce_float(track.get("track_first_frame"))) for track in tracks)
    last_frame = max(int(_coerce_float(track.get("track_last_frame"))) for track in tracks)
    first_timestamp = min(_coerce_float(track.get("track_first_timestamp")) for track in tracks)
    last_timestamp = max(_coerce_float(track.get("track_last_timestamp")) for track in tracks)
    deduplicated_observations = {
        (
            int(_coerce_float(item.get("frame"))),
            tuple(round(_coerce_float(value), 6) for value in (item.get("bbox") or [])[:4]),
        ): item
        for item in merged_observations
    }
    sorted_observations = sorted(deduplicated_observations.values(), key=lambda item: int(_coerce_float(item.get("frame"))))
    frame_count = len(deduplicated_observations)
    max_confidence = max(_coerce_float(track.get("track_max_confidence") or track.get("confidence")) for track in tracks)
    weighted_category = max(class_votes, key=class_votes.get, default=primary.get("category") or "unknown")
    observation_confidences = [_coerce_float(item.get("confidence")) for item in deduplicated_observations.values()]
    avg_confidence = (
        sum(observation_confidences) / len(observation_confidences)
        if observation_confidences else _coerce_float(primary.get("track_avg_confidence"))
    )
    recyclable_status, contaminant_status = material_status(weighted_category)
    hazard_status = "hazard" if any(track.get("track_hazard_status") == "hazard" for track in tracks) or CATEGORY_CLASS_MAP.get(weighted_category) == "contaminant" else "clear"
    sorted_path = sorted(merged_path, key=lambda point: int(_coerce_float(point.get("frame"))))
    sorted_scene_path = sorted(merged_scene_path, key=lambda point: int(_coerce_float(point.get("frame"))))
    merged = {
        **primary,
        "best_box": representative.get("best_box"),
        "best_bbox_norm": representative.get("best_bbox_norm"),
        "bbox_x": representative.get("bbox_x"),
        "bbox_y": representative.get("bbox_y"),
        "bbox_width": representative.get("bbox_width"),
        "bbox_height": representative.get("bbox_height"),
        "segmentation_mask": representative.get("segmentation_mask"),
        "source_track_ids": source_track_ids,
        "track_id": ",".join(source_track_ids),
        "category": weighted_category,
        "material_name": weighted_category,
        "confidence": round(max_confidence, 4),
        "track_avg_confidence": round(avg_confidence, 4),
        "track_max_confidence": round(max_confidence, 4),
        "track_first_frame": first_frame,
        "track_last_frame": last_frame,
        "track_first_timestamp": round(first_timestamp, 3),
        "track_last_timestamp": round(last_timestamp, 3),
        "track_duration_seconds": round(max(0.0, last_timestamp - first_timestamp), 3),
        "track_frame_count": frame_count,
        "track_hazard_status": hazard_status,
        "recyclable_status": recyclable_status,
        "contaminant_status": contaminant_status,
        "track_path": sorted_path,
        **({"track_start_center": {"x": sorted_path[0].get("x"), "y": sorted_path[0].get("y")}} if sorted_path else {}),
        **({"track_end_center": {"x": sorted_path[-1].get("x"), "y": sorted_path[-1].get("y")}} if sorted_path else {}),
        **({"track_start_scene_center": {"x": sorted_scene_path[0].get("x"), "y": sorted_scene_path[0].get("y")}} if sorted_scene_path else {}),
        **({"track_end_scene_center": {"x": sorted_scene_path[-1].get("x"), "y": sorted_scene_path[-1].get("y")}} if sorted_scene_path else {}),
        "track_debug": {
            "frame_observations": sorted_observations,
            "accepted_track_fragments": _track_frame_fragments(sorted_observations),
            "class_votes": {key: round(value, 4) for key, value in class_votes.items()},
            "raw_track_ids": source_track_ids,
            "stabilized_track_path": sorted_scene_path,
            "representative_quality": _bbox_quality(representative.get("best_bbox_norm")),
            "appearance_fingerprints": _bounded_representative_fingerprints(appearance_fingerprints),
        },
        **evaluate_material(weighted_category, max_confidence),
    }
    return merged


def merge_track_fragments(raw_tracks: list[dict], upload_id: str) -> list[dict]:
    logical = sorted(raw_tracks, key=lambda item: int(_coerce_float(item.get("track_first_frame"))))
    changed = True
    merge_reasons = []
    while changed:
        changed = False
        for i, first in enumerate(logical):
            for j in range(i + 1, len(logical)):
                second = logical[j]
                should_merge, reason = _track_merge_score(first, second)
                if not should_merge:
                    continue
                merged = _merge_two_tracks(first, second)
                merge_reasons.append({
                    "source_track_ids": merged.get("source_track_ids"),
                    "reason": reason,
                })
                logical = [item for index, item in enumerate(logical) if index not in {i, j}] + [merged]
                logical.sort(key=lambda item: int(_coerce_float(item.get("track_first_frame"))))
                changed = True
                break
            if changed:
                break
    for index, item in enumerate(logical, start=1):
        source_track_ids = item.get("source_track_ids") or [item.get("track_id") or index]
        object_uid = f"{upload_id}-object-{index:04d}"
        item["object_uid"] = object_uid
        item["stable_object_id"] = object_uid
        item["source_track_ids"] = source_track_ids
        item.setdefault("track_debug", {})["merge_reasons"] = [
            reason for reason in merge_reasons if set(reason.get("source_track_ids") or []).issubset(set(source_track_ids))
        ]
        _video_debug(
            "logical_object_finalized",
            scan_id=upload_id,
            logical_object_id=object_uid,
            source_track_ids=source_track_ids,
            first_frame=item.get("track_first_frame"),
            last_frame=item.get("track_last_frame"),
            observations=item.get("track_frame_count"),
            final_class=item.get("category"),
        )
    return logical


def _persist_tracked_video_objects(
    *,
    tracked_objects: list[dict],
    source_name: str,
    file_id: str,
    job: dict,
    principal: Principal | None,
    database: SupabaseExecutor,
    existing_drive_metadata: dict,
    annotated_video_metadata: dict | None = None,
    annotated_video_path: str | Path | None = None,
    annotated_observations_by_track: dict[str, list[dict]] | None = None,
) -> list[str]:
    scan_ids: list[str] = []
    object_ids = [str(item.get("stable_object_id") or item.get("object_uid") or "") for item in tracked_objects]
    if len(object_ids) != len(set(object_ids)):
        raise RuntimeError("Tracked video persistence received duplicate final physical-object IDs.")
    unfinished = [
        object_id
        for object_id, item in zip(object_ids, tracked_objects)
        if not (item.get("track_debug") or {}).get("physical_reconciliation_completed")
    ]
    if unfinished:
        raise RuntimeError(f"Tracked video persistence received non-final physical-object clusters: {unfinished[:5]}")
    _video_processing_log(
        "physical_persistence_invariant_verified",
        job_id=str(job["id"]),
        final_cluster_count=len(tracked_objects),
        logical_object_ids=object_ids,
    )
    namespace = UUID(str(job["id"]))
    for item in tracked_objects:
        stable_object_id = str(item["stable_object_id"])
        scan_uuid = uuid5(namespace, stable_object_id)
        public_material = {key: value for key, value in item.items() if not key.startswith("_")}
        public_material["material_name"] = (
            public_material.get("material_name")
            or public_material.get("category")
            or "Detected object"
        )
        public_material["confidence"] = float(
            public_material.get("confidence")
            or public_material.get("track_max_confidence")
            or public_material.get("track_avg_confidence")
            or 0
        )
        preview_bytes = None
        preview_metadata = {
            "format": "unavailable",
            "best_box": public_material.get("best_box"),
            "bbox_x": public_material.get("bbox_x"),
            "bbox_y": public_material.get("bbox_y"),
            "bbox_width": public_material.get("bbox_width"),
            "bbox_height": public_material.get("bbox_height"),
        }
        filename = f"{Path(source_name).stem}_{stable_object_id}.jpg"
        selected_observation = _select_annotated_preview_observation(public_material, annotated_observations_by_track)
        if not selected_observation:
            _video_processing_log(
                "tracked_preview_observation_unavailable",
                job_id=str(job["id"]),
                scan_id=str(job["id"]),
                object_scan_id=str(scan_uuid),
                logical_object_id=stable_object_id,
                track_id=public_material.get("track_id"),
                source_track_ids=public_material.get("source_track_ids"),
                annotated_track_ids=sorted((annotated_observations_by_track or {}).keys()),
                worker_build_revision=_worker_build_revision(),
            )
        if selected_observation and annotated_video_metadata and annotated_video_metadata.get("annotated_video_status") == "ready":
            try:
                extracted = _extract_annotated_video_object_preview(annotated_video_path, public_material, selected_observation)
                if extracted:
                    preview_bytes, preview_metadata = extracted
            except Exception as exc:
                best_box = public_material.get("best_box") if isinstance(public_material.get("best_box"), dict) else {}
                _video_processing_log(
                    "tracked_preview_annotated_video_extraction_failed",
                    scan_id=str(job["id"]),
                    object_scan_id=str(scan_uuid),
                    logical_object_id=stable_object_id,
                    track_id=public_material.get("track_id"),
                    annotated_video_path=str(annotated_video_path) if annotated_video_path else None,
                    frame=best_box.get("frame"),
                    timestamp=best_box.get("timestamp"),
                    bbox=best_box.get("xyxy") or public_material.get("best_bbox_norm"),
                    bbox_input=preview_metadata,
                    stage="annotated_video_extraction",
                    selected_observation=selected_observation,
                    worker_build_revision=_worker_build_revision(),
                    error_type=type(exc).__name__,
                    error=safe_error_message(exc),
                )
                traceback.print_exc()
        if preview_bytes is None:
            best_box = public_material.get("best_box") if isinstance(public_material.get("best_box"), dict) else {}
            _video_processing_log(
                "tracked_preview_unavailable",
                job_id=str(job["id"]),
                scan_id=str(job["id"]),
                object_scan_id=str(scan_uuid),
                logical_object_id=stable_object_id,
                track_id=public_material.get("track_id"),
                annotated_video_status=(annotated_video_metadata or {}).get("annotated_video_status"),
                annotated_video_path=str(annotated_video_path) if annotated_video_path else None,
                frame=best_box.get("frame"),
                timestamp=best_box.get("timestamp"),
                bbox=best_box.get("xyxy") or public_material.get("best_bbox_norm"),
                selected_observation=selected_observation,
                preview_source=preview_metadata.get("format"),
                worker_build_revision=_worker_build_revision(),
            )
            raise RuntimeError(f"Tracked-object preview unavailable for final physical cluster {stable_object_id}.")
        if not isinstance(public_material.get("track_debug"), dict):
            public_material["track_debug"] = {}
        public_material["track_debug"]["annotated_frame_observation"] = selected_observation
        public_material["track_debug"]["preview_bbox"] = preview_metadata
        public_material["track_debug"]["preview_annotation_status"] = preview_metadata.get("format", "unavailable")
        public_material["result_type"] = "video_track_object"
        _video_processing_log(
            "tracked_preview_annotation_completed",
            job_id=str(job["id"]),
            scan_id=str(scan_uuid),
            logical_object_id=stable_object_id,
            track_id=public_material.get("track_id"),
            representative_frame_index=preview_metadata.get("frame"),
            source_width=preview_metadata.get("source_width"),
            source_height=preview_metadata.get("source_height"),
            bbox_input=preview_metadata.get("bbox_input"),
            bbox_format=preview_metadata.get("bbox_format"),
            converted_pixel_bbox=preview_metadata.get("box_xyxy"),
            crop_origin={"x": preview_metadata.get("crop_x"), "y": preview_metadata.get("crop_y")},
            crop_dimensions={"width": preview_metadata.get("crop_width"), "height": preview_metadata.get("crop_height")},
            preview_source=preview_metadata.get("format"),
            annotation_status=preview_metadata.get("format"),
            worker_build_revision=_worker_build_revision(),
        )
        object_summary = summarize([public_material])
        object_summary.update({
            "result_kind": "video_track_object",
            "legacy_result": False,
            "total_unique_objects": 1,
            "video_tracking_summary": {
                "total_unique_objects": 1,
                "counts_by_class": {public_material["category"]: 1},
                "hazards": [public_material] if public_material.get("track_hazard_status") == "hazard" else [],
                "tracked_objects": [public_material],
                **(annotated_video_metadata or {}),
            },
        })
        _video_debug(
            "database_write",
            scan_id=str(scan_uuid),
            logical_object_id=stable_object_id,
            source_track_ids=public_material.get("source_track_ids"),
            class_name=public_material.get("category"),
            confidence=public_material.get("confidence"),
        )
        result = persist_scan(
            preview_bytes,
            filename,
            "tracked_video",
            [public_material],
            object_summary,
            source_ref=file_id,
            batch_id=str(job["id"]),
            principal=principal,
            content_type="image/jpeg",
            existing_drive_metadata=existing_drive_metadata,
            database=database,
            scan_result_id=scan_uuid,
            model_version=os.getenv("MODEL_VERSION", "yolov8m-seg-bytetrack"),
        )
        scan_ids.append(str(result["scan_result_id"]))
    if len(scan_ids) != len(tracked_objects):
        raise RuntimeError(
            f"Tracked video persistence count mismatch: {len(scan_ids)} rows for {len(tracked_objects)} final physical clusters."
        )
    return scan_ids


def _process_video_drive_file(file_id: str, job: dict, principal: Principal | None, database: SupabaseExecutor, existing: dict, payload: bytes, name: str) -> list[str]:
    import cv2
    scan_id = str(job["id"])
    job_dir = _video_job_dir(scan_id)
    tmp_path = _source_video_path(job_dir, name)
    annotated_tmp_path = job_dir / "annotated-intermediate.mp4"
    encoded_tmp_path = job_dir / "result.mp4"
    annotated_writer = None
    capture = None
    annotated_video_metadata = {
        "annotated_video_url": None,
        "annotated_video_storage_path": None,
        "annotated_video_status": "unavailable",
        "annotated_video_error": None,
        "annotated_video_probe": None,
    }
    scan_ids: list[str] = [str(scan_id) for scan_id in job.get("scan_ids") or []]
    options = job.get("options") or {}
    aggregator = VideoTrackAggregator(str(job["id"]), counting_line=_parse_counting_line(options))
    annotated_observations_by_track: dict[str, list[dict]] = {}
    last_checkpoint_at = time.monotonic()
    annotated_frames_written = 0
    frame_total = 0
    fps = DEFAULT_VIDEO_FPS
    try:
        _video_processing_log(
            "mp4_job_worker_identity",
            job_id=scan_id,
            scan_id=scan_id,
            file_id=file_id,
            filename=name,
            worker_build_revision=_worker_build_revision(),
            preview_generation_impl="annotated_video_frame_extraction",
            **deployment_identity(),
        )
        payload_size = len(payload or b"")
        _video_processing_log(
            "upload_received",
            scan_id=scan_id,
            file_id=file_id,
            filename=name,
            mime_type=existing.get("drive_mime_type") or "video/mp4",
            upload_size=payload_size,
        )
        if payload_size <= 0:
            raise VideoDecodeError("Downloaded MP4 source is empty.")
        tmp_path.write_bytes(payload)
        source_size = tmp_path.stat().st_size if tmp_path.exists() else 0
        _video_processing_log("source_saved", scan_id=scan_id, input_path=tmp_path, input_exists=tmp_path.exists(), input_size=source_size)
        if source_size <= 0:
            raise VideoDecodeError("Saved MP4 source is empty.")
        capture = cv2.VideoCapture(str(tmp_path))
        capture_opened = bool(capture.isOpened())
        raw_width = int(capture.get(cv2.CAP_PROP_FRAME_WIDTH) or 0)
        raw_height = int(capture.get(cv2.CAP_PROP_FRAME_HEIGHT) or 0)
        raw_fps = float(capture.get(cv2.CAP_PROP_FPS) or 0)
        frame_total = int(capture.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
        fps = raw_fps if math.isfinite(raw_fps) and raw_fps > 0 else DEFAULT_VIDEO_FPS
        fps_fallback_used = fps != raw_fps
        _video_processing_log(
            "capture_opened",
            scan_id=scan_id,
            is_opened=capture_opened,
            metadata_width=raw_width,
            metadata_height=raw_height,
            metadata_fps=raw_fps,
            fps=fps,
            fps_fallback_used=fps_fallback_used,
            frame_count=frame_total,
        )
        if not capture_opened:
            raise VideoDecodeError("OpenCV could not open the MP4 source.")
        first_ok, first_frame = capture.read()
        _video_processing_log("first_frame_decoded", scan_id=scan_id, ok=bool(first_ok), has_frame=first_frame is not None)
        if not first_ok or first_frame is None:
            raise VideoDecodeError("Unable to decode the first video frame.")
        frame_height, frame_width = first_frame.shape[:2]
        width, height = _even_video_dimensions(frame_width, frame_height)
        if width <= 0 or height <= 0:
            raise VideoDecodeError(f"Invalid decoded video dimensions: {frame_width}x{frame_height}")
        camera_motion = VideoCameraMotionState(width, height)
        annotated_tmp_path = job_dir / "annotated-intermediate.mp4"
        encoded_tmp_path = job_dir / "result.mp4"
        video_model = get_model()
        tracker_path = str(APP_ROOT / VIDEO_TRACKER_CONFIG)
        if not Path(tracker_path).exists():
            tracker_path = VIDEO_TRACKER_CONFIG
        _video_debug("video_tracking_started", scan_id=scan_id, source_name=name, frame_total=frame_total, tracker=tracker_path, stride=1)
        _update_job(job["id"], database, total_count=None, result_summary={"mode": "tracked_video", "frame_total": frame_total})
        frame_index = 0
        pending_frame = first_frame
        while True:
            if pending_frame is None:
                ok, frame = capture.read()
                if not ok:
                    break
            else:
                frame = pending_frame
                pending_frame = None
            frame = _normalize_video_frame(frame, width, height)
            timestamp = frame_index / fps if fps else 0.0
            motion_meta = camera_motion.observe(frame, frame_index)
            results = video_model.track(frame, persist=True, tracker=tracker_path, verbose=False, device=MODEL_DEVICE)
            result = results[0] if results else None
            detections = _result_track_observations(result, frame, frame_index, timestamp) if result else []
            detections = attach_camera_motion_to_detections(detections, camera_motion, motion_meta)
            aggregator.observe(frame_index, timestamp, detections)
            if annotated_video_metadata.get("annotated_video_status") != "failed":
                try:
                    if annotated_writer is None:
                        fourcc = cv2.VideoWriter_fourcc(*"mp4v")
                        annotated_writer = cv2.VideoWriter(str(annotated_tmp_path), fourcc, fps, (width, height))
                        writer_opened = bool(annotated_writer.isOpened())
                        _video_processing_log(
                            "writer_opened",
                            scan_id=scan_id,
                            writer_path=annotated_tmp_path,
                            codec="mp4v",
                            fps=fps,
                            width=width,
                            height=height,
                            is_opened=writer_opened,
                        )
                        if not writer_opened:
                            raise RuntimeError("OpenCV VideoWriter could not open annotated MP4 output")
                        annotated_video_metadata["annotated_video_status"] = "processing"
                    output_frame = _annotate_video_frame(frame, detections) if detections else frame
                    output_frame = _normalize_video_frame(output_frame, width, height)
                    annotated_writer.write(output_frame)
                    annotated_frames_written += 1
                    _record_annotated_frame_observations(annotated_observations_by_track, detections, frame_index, timestamp, width, height)
                except Exception as exc:
                    annotated_video_metadata.update({
                        "annotated_video_status": "failed",
                        "annotated_video_error": safe_error_message(exc),
                    })
                    _video_processing_log("annotation_failed", scan_id=scan_id, error_type=type(exc).__name__, error=safe_error_message(exc))
                    traceback.print_exc()
                    if annotated_writer is not None:
                        annotated_writer.release()
                        annotated_writer = None
            now = time.monotonic()
            if now - last_checkpoint_at >= 8:
                summary = _video_tracking_summary(aggregator.finalized)
                summary["frame_detections"] = sum(len(track.get("track_debug", {}).get("frame_observations", [])) for track in aggregator.finalized)
                summary["raw_track_count"] = len(aggregator.finalized) + len(aggregator.active)
                summary.update(annotated_video_metadata)
                _update_job(job["id"], database, processed_count=summary["total_unique_objects"], scan_ids=scan_ids, result_summary=summary)
                last_checkpoint_at = now
            frame_index += 1
        capture.release()
        capture = None
        if annotated_writer is not None:
            annotated_writer.release()
            annotated_writer = None
        _video_processing_log(
            "annotation_completed",
            scan_id=scan_id,
            annotated_status=annotated_video_metadata.get("annotated_video_status"),
            annotated_frames_written=annotated_frames_written,
            intermediate_path=annotated_tmp_path,
            intermediate_exists=Path(annotated_tmp_path).exists() if annotated_tmp_path else False,
            intermediate_size=Path(annotated_tmp_path).stat().st_size if annotated_tmp_path and Path(annotated_tmp_path).exists() else 0,
        )
        if annotated_tmp_path and annotated_video_metadata.get("annotated_video_status") != "failed":
            try:
                if annotated_frames_written <= 0:
                    raise RuntimeError("Annotated writer did not write any frames.")
                _encode_browser_mp4(annotated_tmp_path, encoded_tmp_path)
                probe = _ffprobe_mp4(encoded_tmp_path)
                storage_path = f"annotated-videos/{job['id']}/result.mp4"
                upload = upload_file_to_supabase_storage(
                    encoded_tmp_path,
                    storage_path,
                    "video/mp4",
                    database,
                )
                annotated_video_metadata.update({
                    "annotated_video_url": upload["public_url"],
                    "annotated_video_storage_path": storage_path,
                    "annotated_video_status": "ready",
                    "annotated_video_error": None,
                    "annotated_video_probe": probe,
                })
                _video_processing_log("supabase_upload_completed", scan_id=scan_id, storage_path=storage_path, content_type="video/mp4", has_url=bool(upload.get("public_url")))
            except Exception as exc:
                annotated_video_metadata.update({
                    "annotated_video_url": None,
                    "annotated_video_status": "failed",
                    "annotated_video_error": safe_error_message(exc),
                })
                _video_processing_log("annotated_mp4_failed", scan_id=scan_id, error_type=type(exc).__name__, error=safe_error_message(exc))
                traceback.print_exc()
        aggregator.finish(frame_index)
        raw_tracks = aggregator.finalized
        logical_objects = merge_track_fragments(raw_tracks, str(job["id"]))
        canonical_objects, duplicate_report = reconcile_duplicate_tracked_objects(logical_objects, str(job["id"]))
        summary = _video_tracking_summary(canonical_objects)
        if summary["total_unique_objects"] != len(canonical_objects):
            raise RuntimeError("Final video summary count does not match physical cluster count.")
        scan_ids = _persist_tracked_video_objects(
            tracked_objects=canonical_objects,
            source_name=name,
            file_id=file_id,
            job=job,
            principal=principal,
            database=database,
            existing_drive_metadata=existing,
            annotated_video_metadata=annotated_video_metadata,
            annotated_video_path=encoded_tmp_path if encoded_tmp_path.exists() else annotated_tmp_path,
            annotated_observations_by_track=annotated_observations_by_track,
        )
        if len(scan_ids) != summary["total_unique_objects"]:
            raise RuntimeError("Persisted tracked-object row count does not match final physical cluster count.")
        summary.update({
            "scan_id": str(job["id"]),
            "result_type": "video_tracking",
            "frame_total": frame_total,
            "frame_detections": sum(len(track.get("track_debug", {}).get("frame_observations", [])) for track in raw_tracks),
            "raw_track_count": len(raw_tracks),
            "logical_track_count": len(logical_objects),
            "duplicate_groups_confirmed": len(duplicate_report.get("confirmed_groups") or []),
            "physical_reconciliation": duplicate_report,
            "camera_motion": camera_motion.summary(),
            "filtered_tracks": max(0, len(raw_tracks) - len(canonical_objects)),
            "database_rows_written": len(scan_ids),
            **annotated_video_metadata,
        })
        _video_debug(
            "video_tracking_completed",
            scan_id=str(job["id"]),
            frame_total=frame_total,
            frame_detections=summary["frame_detections"],
            raw_track_count=len(raw_tracks),
            final_logical_objects=len(canonical_objects),
            database_rows_written=len(scan_ids),
        )
        _update_job(job["id"], database, processed_count=len(scan_ids), total_count=summary["total_unique_objects"], scan_ids=scan_ids, result_summary=summary)
        _video_processing_log("scan_completed", scan_id=scan_id, final_scan_ids=scan_ids, annotated_video_status=annotated_video_metadata.get("annotated_video_status"))
        return scan_ids
    except Exception:
        _video_processing_log("scan_failed", scan_id=scan_id, final_status="failed")
        traceback.print_exc()
        raise
    finally:
        try:
            if annotated_writer is not None:
                annotated_writer.release()
        except Exception:
            pass
        try:
            if capture is not None:
                capture.release()
        except Exception:
            pass
        shutil.rmtree(job_dir, ignore_errors=True)
        _video_processing_log("temp_cleanup_completed", scan_id=scan_id, job_dir=job_dir, exists=job_dir.exists())


def _process_drive_file(file_id: str, job: dict, principal: Principal | None, database: SupabaseExecutor) -> list[str]:
    info = _drive_file_info(file_id)
    if info.get("trashed"):
        return []
    existing = {
        "storage_provider": "google_drive_and_supabase_storage",
        "drive_file_id": info.get("id"),
        "drive_file_name": info.get("name"),
        "drive_mime_type": info.get("mimeType"),
        "drive_web_url": info.get("webViewLink"),
        "image_url": info.get("webViewLink"),
    }
    payload = _download_drive_file(file_id)
    mime = str(info.get("mimeType") or "")
    name = info.get("name") or file_id
    if mime == "video/mp4" or name.lower().endswith(".mp4"):
        return _process_video_drive_file(file_id, job, principal, database, existing, payload, name)
    result = run_scan(
        payload,
        name,
        "image",
        source_ref=file_id,
        batch_id=str(job["id"]),
        principal=principal,
        content_type=mime or guess_type(name)[0] or "image/jpeg",
        existing_drive_metadata=existing,
        database=database,
    )
    return [str(result["scan_result_id"])]


def _update_job(job_id: str, database: SupabaseExecutor, **fields) -> None:
    payload = {**fields, "updated_at": datetime.now(timezone.utc).isoformat()}
    try:
        database.execute(
            lambda client: client.table(JOBS_TABLE).update(payload).eq("id", job_id).execute()
        )
    except Exception as exc:
        if "result_summary" in payload and getattr(exc, "code", "") in {"PGRST204", "PGRST205", "42703"}:
            payload.pop("result_summary", None)
            database.execute(
                lambda client: client.table(JOBS_TABLE).update(payload).eq("id", job_id).execute()
            )
            return
        raise


def _complete_processing_job(job_id: str, database: SupabaseExecutor, scan_ids: list[str]) -> None:
    completion = {
        "status": "completed",
        "scan_ids": scan_ids,
        "processed_count": len(scan_ids),
        "total_count": len(scan_ids),
        "completed_at": datetime.now(timezone.utc).isoformat(),
    }
    try:
        _update_job(job_id, database, **completion, lease_expires_at=None, worker_id=None)
    except Exception as exc:
        if getattr(exc, "code", "") not in {"PGRST204", "PGRST205", "42703"}:
            raise
        print(f"[worker] optional completion cleanup skipped: {type(exc).__name__}: {safe_worker_error_message(exc)}")
        _update_job(job_id, database, **completion)
    try:
        _finalize_false_positive_reprocess(database, job_id, scan_ids)
    except Exception as exc:
        print(f"[false-positive] reprocess completion linkage failed: {type(exc).__name__}: {safe_worker_error_message(exc)}")


def _process_job(job: dict, database: SupabaseExecutor) -> list[str]:
    principal = Principal(
        str(job.get("created_by_type") or "api_key"),
        str(job.get("created_by") or "worker"),
        frozenset({"scan:read", "scan:write", "job:read", "review:write"}),
    )
    source = job.get("source")
    if source == "drive_file":
        file_ids = [job["source_ref"]]
    elif source == "drive_folder":
        service = _drive_service()
        response = service.files().list(
            q=f"'{job['source_ref']}' in parents and trashed = false",
            fields="files(id,mimeType,name)",
            pageSize=1000,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        file_ids = [item["id"] for item in response.get("files", []) if item.get("mimeType", "").startswith(("image/", "video/"))]
    else:
        raise ValueError("Only Drive file and Drive folder ingestion is enabled; URL ingestion is disabled for SSRF safety.")

    scan_ids: list[str] = []
    is_false_positive_reprocess = isinstance(job.get("options"), dict) and job.get("options", {}).get("job_type") == "false_positive_reprocess"
    for file_id in file_ids:
        existing_response = database.execute(
            lambda client: client.table(PROCESSED_DRIVE_FILES_TABLE).select("drive_file_id").eq("drive_file_id", file_id).maybe_single().execute()
        )
        existing = existing_response.data if existing_response else None
        if existing and not is_false_positive_reprocess:
            saved = database.execute(
                lambda client: client.table(SCAN_RESULTS_TABLE).select("id").eq("batch_id", str(job["id"])).execute()
            ).data or []
            scan_ids.extend(str(row["id"]) for row in saved if row.get("id"))
            continue
        scan_ids.extend(_process_drive_file(file_id, job, principal, database))
        def recover_processed_file(client):
            response = client.table(PROCESSED_DRIVE_FILES_TABLE).select("drive_file_id").eq("drive_file_id", file_id).maybe_single().execute()
            return response.data if response and response.data else None

        if not is_false_positive_reprocess:
            database.execute(
                lambda client: client.table(PROCESSED_DRIVE_FILES_TABLE).insert({"drive_file_id": file_id, "scan_result_id": scan_ids[-1] if scan_ids else None}).execute(),
                recover=recover_processed_file,
            )
    return scan_ids


def _worker_loop() -> None:
    database = SupabaseExecutor()
    active_job = None
    while True:
        job_id = None
        try:
            if not database.client:
                time.sleep(15)
                continue
            if active_job is None:
                job_response = database.execute(
                    lambda client: client.table(JOBS_TABLE).select("*").eq("status", "queued").order("created_at").limit(1).maybe_single().execute()
                )
                active_job = job_response.data if job_response else None
                if not active_job:
                    time.sleep(5)
                    continue
                job_id = str(active_job["id"])
                _update_job(job_id, database, status="processing", started_at=datetime.now(timezone.utc).isoformat(), attempts=int(active_job.get("attempts") or 0) + 1)
            job_id = str(active_job["id"])
            scan_ids = _process_job(active_job, database)
            _update_job(job_id, database, status="completed", scan_ids=scan_ids, processed_count=len(scan_ids), total_count=len(scan_ids), completed_at=datetime.now(timezone.utc).isoformat())
            try:
                _finalize_false_positive_reprocess(database, job_id, scan_ids)
            except Exception as exc:
                print(f"[false-positive] reprocess completion linkage failed: {type(exc).__name__}: {safe_worker_error_message(exc)}")
            active_job = None
        except SupabaseTemporarilyUnavailable:
            # Keep this in-memory job alive. A later retry resumes via persisted frame names.
            print("[worker] Supabase temporarily unavailable; keeping job active for retry")
            time.sleep(2)
        except Exception as exc:
            print(f"[worker] job failed: {type(exc).__name__}: {safe_error_message(exc)}")
            if job_id:
                try:
                    _update_job(job_id, database, status="failed", error=safe_error_message(exc))
                except SupabaseTemporarilyUnavailable:
                    print("[worker] unable to persist failure; keeping job active for retry")
                    time.sleep(2)
                    continue
            active_job = None
            time.sleep(5)


@app.on_event("startup")
def start_worker() -> None:
    print(f"[startup] diagnostics {json.dumps(safe_startup_diagnostics(), sort_keys=True)}")
    if (
        SERVICE_MODE == "api"
        and PROCESSING_BACKEND == "local-thread"
        and not is_production()
        and SUPABASE_URL
        and SUPABASE_SERVICE_ROLE_KEY
        and os.getenv("DISABLE_WORKER", "false").lower() != "true"
    ):
        try:
            startup_database = SupabaseExecutor()
            startup_database.execute(
                lambda client: client.table(JOBS_TABLE).update({"status": "queued", "error": "Worker restarted before completion."}).eq("status", "processing").execute()
            )
        except Exception as exc:
            print(f"[worker] restart recovery skipped: {type(exc).__name__}: {safe_error_message(exc)}")
        threading.Thread(target=_worker_loop, name="purityloop-worker", daemon=True).start()


@app.post("/api/uploads/start")
def start_upload(payload: UploadStartInput, principal: Principal = Depends(require_scope("scan:write"))):
    upload_id = None
    try:
        filename, size_bytes, mime = _validate_upload_start_payload(payload)
        name = safe_drive_filename(filename)
        _upload_start_log(
            "request_validated",
            filename=filename,
            safe_filename=name,
            mime_type=mime,
            size_bytes=size_bytes,
            upload_type="video/mp4",
            principal_kind=getattr(principal, "kind", None),
            principal_id=getattr(principal, "id", None),
            drive_folder_configured=bool(GOOGLE_DRIVE_UPLOADED_IMAGES_FOLDER_ID),
        )
        _upload_start_log("drive_resumable_init_started", filename=name, mime_type=mime, size_bytes=size_bytes, storage_operation="google_drive_resumable_upload")
        upload_url = _create_drive_resumable_upload(name, size_bytes, mime)
        upload_id = str(uuid4())
        database = SupabaseExecutor()
        session = _insert_upload_session(database, principal, upload_id=upload_id, filename=name, mime=mime, size_bytes=size_bytes, upload_url=upload_url)
        _upload_start_log("drive_resumable_init_completed", upload_id=upload_id, filename=name, chunk_size=UPLOAD_CHUNK_SIZE_BYTES, storage_operation="google_drive_resumable_upload")
        return _upload_session_response(session)
    except UploadStartFailure as exc:
        _upload_start_log("failed", upload_id=upload_id, code=exc.code, stage=exc.stage, status_code=exc.status_code, error_type=type(exc.__cause__ or exc).__name__, error=safe_error_message(exc))
        traceback.print_exc()
        raise _upload_start_error_response(exc) from exc
    except HTTPException:
        raise
    except Exception as exc:
        _upload_start_log("failed", upload_id=upload_id, code="UPLOAD_START_INTERNAL_ERROR", stage="unexpected", status_code=500, error_type=type(exc).__name__, error=safe_error_message(exc))
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail={"code": "UPLOAD_START_INTERNAL_ERROR", "message": "Unexpected upload-start failure.", "stage": "unexpected"},
        ) from exc


@app.put("/api/uploads/{upload_id}")
async def upload_chunk(upload_id: str, request: Request, principal: Principal = Depends(require_scope("scan:write"))):
    database = SupabaseExecutor()
    session = _load_upload_session(database, upload_id, principal)
    content_range = request.headers.get("content-range")
    if not session:
        raise HTTPException(status_code=404, detail="MP4 upload session was not found.")
    if session.get("status") == "failed":
        raise HTTPException(status_code=409, detail="MP4 upload session previously failed.")
    if session.get("status") == "completed":
        return {"complete": True, "drive_file": {"id": session.get("drive_file_id")}}
    expires_at = session.get("expires_at")
    if expires_at and datetime.fromisoformat(str(expires_at).replace("Z", "+00:00")) <= datetime.now(timezone.utc):
        _update_upload_session(database, upload_id, {"status": "expired", "error": "Upload session expired."})
        raise HTTPException(status_code=410, detail="MP4 upload session expired.")
    if not content_range:
        raise HTTPException(status_code=400, detail="MP4 chunk is missing Content-Range.")
    try:
        from google.auth.transport.requests import AuthorizedSession
        chunk = await request.body()
        received_size = _content_range_end(content_range)
        response = AuthorizedSession(oauth_drive_credentials()).put(
            session["drive_resumable_url"],
            headers={"Content-Range": content_range, "Content-Length": str(len(chunk))},
            data=chunk,
            timeout=120,
        )
        if response.status_code == 308:
            if received_size is not None:
                _update_upload_session(database, upload_id, {"received_size": max(int(session.get("received_size") or 0), received_size), "status": "upload_pending"})
            return {"complete": False}
        response.raise_for_status()
        drive_file = response.json()
        _update_upload_session(
            database,
            upload_id,
            {
                "received_size": int(session.get("total_size") or received_size or 0),
                "drive_file_id": drive_file.get("id"),
                "status": "completed",
                "error": None,
            },
        )
        return {"complete": True, "drive_file": drive_file}
    except HTTPException:
        raise
    except Exception as exc:
        _update_upload_session(database, upload_id, {"status": "failed", "error": safe_error_message(exc)})
        raise HTTPException(status_code=502, detail="Unable to upload MP4 chunk to Google Drive.") from exc


@app.post("/api/ingest")
def ingest(payload: IngestInput, principal: Principal = Depends(require_scope("scan:write"))):
    if payload.source not in {"drive_file", "drive_folder"} or not payload.ref.strip():
        raise HTTPException(status_code=400, detail="Use source=drive_file or source=drive_folder with a Drive id.")
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    database = SupabaseExecutor(supabase)
    source_ref = payload.ref.strip()
    if payload.source == "drive_file":
        _verify_completed_upload_for_drive_file(database, source_ref, principal)
    row = {
        "source": payload.source,
        "source_ref": source_ref,
        "options": payload.options or {},
        "created_by": principal.id,
        "created_by_type": principal.kind,
    }
    inserted = database.execute(lambda client: client.table(JOBS_TABLE).insert(row).execute().data or [])
    if not inserted:
        raise HTTPException(status_code=500, detail="Unable to create processing job.")
    job_id = str(inserted[0]["id"])
    if PROCESSING_BACKEND != "cloud-tasks":
        if is_production():
            raise HTTPException(status_code=500, detail="Production processing backend must be cloud-tasks.")
        return {"job_id": job_id, "status": inserted[0].get("status", "queued")}
    try:
        task = enqueue_processing_task(job_id)
        database.execute(
            lambda client: client.table(JOBS_TABLE).update(
                {"status": "queued", "dispatched_at": datetime.now(timezone.utc).isoformat(), "dispatch_error": None}
            ).eq("id", job_id).execute()
        )
    except Exception as exc:
        database.execute(
            lambda client: client.table(JOBS_TABLE).update(
                {"status": "queued", "dispatch_error": safe_error_message(exc), "updated_at": datetime.now(timezone.utc).isoformat()}
            ).eq("id", job_id).execute()
        )
        raise HTTPException(status_code=503, detail="Unable to dispatch processing task.") from exc
    return {"job_id": job_id, "status": "queued", "task_dispatched": True, "task_duplicate": task.get("duplicate", False)}


@app.get("/api/jobs/{job_id}")
def get_job(job_id: str, principal: Principal = Depends(require_scope("job:read"))):
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    database = SupabaseExecutor()
    try:
        def get_query(client):
            query = client.table(JOBS_TABLE).select("*").eq("id", job_id)
            if principal.kind == "user":
                query = query.eq("created_by", principal.id).eq("created_by_type", "user")
            return query.maybe_single().execute()
        response = database.execute(get_query)
    except SupabaseTemporarilyUnavailable:
        return JSONResponse(
            status_code=503,
            content={"detail": "Job status temporarily unavailable. Retry shortly.", "retryable": True},
            headers={"Retry-After": "3"},
        )
    row = response.data if response else None
    if not row:
        raise HTTPException(status_code=404, detail="Processing job was not found.")
    return row


@app.post("/internal/jobs/process")
def process_internal_job(payload: WorkerJobInput):
    if SERVICE_MODE != "worker":
        raise HTTPException(status_code=404, detail="Not found.")
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise HTTPException(status_code=500, detail="Worker is not configured.")
    if MODEL_DEVICE.startswith("cuda"):
        get_model()
    job_id = str(payload.job_id)
    database = SupabaseExecutor()
    try:
        claimed = _claim_processing_job(database, job_id)
        status = claimed.get("status")
        if status in {"complete", "completed"}:
            return {"job_id": job_id, "status": status, "already_completed": True}
        if not claimed.get("claimed"):
            return JSONResponse(status_code=409, content={"detail": "Job is not available for processing.", "retryable": True})
        job_response = database.execute(lambda client: client.table(JOBS_TABLE).select("*").eq("id", job_id).maybe_single().execute())
        job = job_response.data if job_response else None
        if not job:
            raise HTTPException(status_code=404, detail="Job was not found.")
        scan_ids = _process_job(job, database)
        _complete_processing_job(job_id, database, scan_ids)
        return {"job_id": job_id, "status": "completed", "scan_ids": scan_ids}
    except HTTPException:
        raise
    except SupabaseTemporarilyUnavailable as exc:
        try:
            _update_job(job_id, database, status="queued", error="Transient database failure.", lease_expires_at=None)
        except Exception:
            pass
        return JSONResponse(status_code=503, content={"detail": "Processing temporarily unavailable.", "retryable": True})
    except Exception as exc:
        try:
            current = database.execute(lambda client: client.table(JOBS_TABLE).select("status").eq("id", job_id).maybe_single().execute()).data
            if current and current.get("status") in {"complete", "completed"}:
                print(f"[worker] post-completion error acknowledged: {type(exc).__name__}: {safe_worker_error_message(exc)}")
                return {"job_id": job_id, "status": current.get("status"), "already_completed": True, "post_completion_error": True}
            _update_job(job_id, database, status="failed", error=safe_worker_error_message(exc), lease_expires_at=None, worker_id=None)
        except Exception:
            pass
        return JSONResponse(status_code=500, content={"detail": "Processing failed.", "retryable": True})


@app.post("/api/scans/{scan_id}/false-positive")
def create_false_positive_report(scan_id: str, payload: FalsePositiveReportInput, principal: Principal = Depends(require_scope("review:write"))):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    database = SupabaseExecutor(supabase)
    _require_false_positive_principal(database, principal)
    expected_category = canonical_category_key(payload.expected_category)
    if expected_category == "unknown":
        raise HTTPException(status_code=400, detail="Expected category is not supported.")
    reason = str(payload.reason or "").strip().lower()
    if reason not in FALSE_POSITIVE_REASONS:
        raise HTTPException(status_code=400, detail="False-positive reason is not supported.")
    note = str(payload.note or "").strip() if payload.note else None
    if note and len(note) > 1000:
        raise HTTPException(status_code=400, detail="False-positive note must be 1000 characters or fewer.")
    source = resolve_original_upload(scan_id, database, principal, str(payload.detected_material_id) if payload.detected_material_id else None)
    predicted_category = source["original_category"]
    if expected_category == predicted_category:
        raise HTTPException(status_code=400, detail="Expected category must differ from the original prediction.")
    material_id = str(source["material"].get("id") or "")
    duplicate = _active_false_positive_for_material(database, str(source["scan"]["id"]), material_id, expected_category)
    if duplicate:
        return JSONResponse(status_code=409, content={"ok": False, "duplicate": True, "report": _safe_false_positive_report(duplicate)})
    row = {
        "original_scan_id": str(source["scan"]["id"]),
        "original_detected_material_id": material_id or None,
        "batch_id": source["batch_id"] if _looks_like_uuid(source["batch_id"]) else None,
        "processing_job_id": source["processing_job_id"] if _looks_like_uuid(source["processing_job_id"]) else None,
        "reported_by": principal.id,
        "predicted_category": predicted_category,
        "predicted_confidence": source["original_confidence"],
        "expected_category": expected_category,
        "reason": reason,
        "note": note,
        "source_type": source["source_type"],
        "source_name": source["source_name"],
        "source_storage_path": source["source_storage_path"],
        "source_drive_file_id": source["source_drive_file_id"],
        "original_model_version": source["original_model_version"],
        "original_model_hash": source["original_model_hash"],
        "status": "reported",
    }
    inserted = database.execute(lambda client: client.table(FALSE_POSITIVE_REPORTS_TABLE).insert(row).execute().data or [])
    if not inserted:
        raise HTTPException(status_code=500, detail="Unable to create false-positive report.")
    report = inserted[0]
    false_positive_log(
        "false_positive_report_created",
        report_id=report.get("id"),
        original_scan_id=row["original_scan_id"],
        original_detected_material_id=row["original_detected_material_id"],
        reporter_id=principal.id,
        original_category=predicted_category,
        original_confidence=source["original_confidence"],
        expected_category=expected_category,
        source_type=source["source_type"],
        source_filename=source["source_name"],
        original_model_version=source["original_model_version"],
        original_model_hash=source["original_model_hash"],
    )
    return {
        "ok": True,
        "report_id": report["id"],
        "status": report.get("status", "reported"),
        "original_scan_id": row["original_scan_id"],
        "predicted_category": predicted_category,
        "predicted_confidence": source["original_confidence"],
        "expected_category": expected_category,
    }


@app.get("/api/false-positives/{report_id}")
def get_false_positive_report(report_id: str, principal: Principal = Depends(require_scope("scan:read"))):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    database = SupabaseExecutor(supabase)
    _require_false_positive_principal(database, principal)
    report = _load_false_positive_report(database, report_id)
    scan = _load_scan_for_false_positive(str(report["original_scan_id"]), database, principal)
    material = _first_material(scan, str(report["original_detected_material_id"]) if report.get("original_detected_material_id") else None, database)
    reprocessed_scan = None
    if report.get("reprocessed_scan_id"):
        try:
            reprocessed_scan = _load_scan_for_false_positive(str(report["reprocessed_scan_id"]), database, principal)
        except HTTPException:
            reprocessed_scan = None
    return {
        "report": _safe_false_positive_report(report),
        "original_scan": {
            "id": scan.get("id"),
            "source_name": scan.get("source_name") or scan.get("drive_file_name"),
            "source_type": scan.get("source_type"),
            "model_version": scan.get("model_version"),
            "created_at": scan.get("created_at"),
        },
        "original_prediction": {
            "detected_material_id": material.get("id"),
            "category": report.get("predicted_category"),
            "confidence": normalized_confidence(report.get("predicted_confidence")),
        },
        "reprocessed_result": reprocessed_scan,
        "decision_threshold": DECISION_CONFIDENCE_THRESHOLD,
    }


@app.get("/api/false-positives")
def list_false_positive_reports(
    status: str | None = None,
    predicted_category: str | None = None,
    expected_category: str | None = None,
    scan_id: str | None = None,
    limit: int = 50,
    offset: int = 0,
    principal: Principal = Depends(require_scope("scan:read")),
):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    database = SupabaseExecutor(supabase)
    _require_false_positive_principal(database, principal)
    limit = max(1, min(int(limit), 100))
    offset = max(0, int(offset))
    def run(client):
        query = client.table(FALSE_POSITIVE_REPORTS_TABLE).select("*", count="exact")
        if status:
            query = query.eq("status", str(status).strip().lower())
        if predicted_category:
            query = query.eq("predicted_category", canonical_category_key(predicted_category))
        if expected_category:
            query = query.eq("expected_category", canonical_category_key(expected_category))
        if scan_id:
            query = query.eq("original_scan_id", scan_id)
        return query.order("created_at", desc=True).range(offset, offset + limit - 1).execute()
    response = database.execute(run)
    return {"items": [_safe_false_positive_report(row) for row in (response.data or [])], "total": int(getattr(response, "count", 0) or 0), "limit": limit, "offset": offset}


@app.post("/api/false-positives/{report_id}/reprocess")
def reprocess_false_positive(report_id: str, principal: Principal = Depends(require_scope("scan:write"))):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    database = SupabaseExecutor(supabase)
    _require_false_positive_principal(database, principal)
    report = _load_false_positive_report(database, report_id)
    if report.get("status") in FALSE_POSITIVE_TERMINAL_STATUSES:
        raise HTTPException(status_code=409, detail="Resolved or dismissed reports cannot be reprocessed.")
    result = _create_false_positive_reprocess_job(database, report, principal)
    return {"ok": True, "report": _safe_false_positive_report(result["report"]), "job_id": result["job"]["id"], "status": result["report"].get("status", "queued")}


@app.post("/api/false-positives/{report_id}/retry")
def retry_false_positive(report_id: str, principal: Principal = Depends(require_scope("scan:write"))):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    database = SupabaseExecutor(supabase)
    _require_false_positive_principal(database, principal)
    report = _load_false_positive_report(database, report_id)
    if report.get("status") != "failed":
        raise HTTPException(status_code=409, detail="Only failed reports can be retried.")
    result = _create_false_positive_reprocess_job(database, report, principal)
    return {"ok": True, "report": _safe_false_positive_report(result["report"]), "job_id": result["job"]["id"], "status": result["report"].get("status", "queued")}


@app.post("/api/false-positives/{report_id}/dismiss")
def dismiss_false_positive(report_id: str, payload: FalsePositiveDismissInput | None = None, principal: Principal = Depends(require_scope("review:write"))):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    database = SupabaseExecutor(supabase)
    _require_false_positive_reviewer(database, principal)
    report = _load_false_positive_report(database, report_id)
    if report.get("status") == "dismissed":
        return {"ok": True, "report": _safe_false_positive_report(report)}
    note = str(payload.reason or "").strip() if payload and payload.reason else ""
    failure_reason = f"DISMISSED: {note[:500]}" if note else "DISMISSED"
    updated = _update_false_positive_report(database, report_id, status="dismissed", dismissed_at=datetime.now(timezone.utc).isoformat(), failure_reason=failure_reason)
    false_positive_log("false_positive_report_dismissed", report_id=report_id, original_scan_id=report.get("original_scan_id"), reporter_id=report.get("reported_by"))
    return {"ok": True, "report": _safe_false_positive_report(updated)}


@app.get("/api/analytics")
def analytics(days: int = 7, principal: Principal = Depends(require_scope("scan:read"))):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    days = max(1, min(int(days), 90))
    since = datetime.now(timezone.utc).timestamp() - days * 86400
    since_iso = datetime.fromtimestamp(since, timezone.utc).isoformat()
    rows = scoped_query(supabase.table(SCAN_RESULTS_TABLE).select("overall_status,contamination_risk,overall_confidence,human_review_required,created_at"), principal).gte("created_at", since_iso).execute().data or []
    return {
        "days": days,
        "total_scans": len(rows),
        "review_required": sum(1 for row in rows if row.get("human_review_required") or row.get("overall_status") == "review_required"),
        "accepted": sum(1 for row in rows if row.get("overall_status") == "accepted"),
        "average_confidence": round(sum(float(row.get("overall_confidence") or 0) for row in rows) / len(rows), 4) if rows else 0,
    }


def analytics_category(value: Any) -> str:
    key = re.sub(r"[_-]+", " ", str(value or "").strip().lower())
    key = re.sub(r"\s+", " ", key)
    if "food" in key or "organic" in key:
        return "food organic"
    if any(word in key for word in ("general", "trash", "landfill", "unknown")):
        return "general trash"
    return next((category for category in ANALYTICS_MATERIAL_ESTIMATES if category in key), "general trash")


def analytics_timestamp(value: Any) -> datetime | None:
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    except (TypeError, ValueError):
        return None


def analytics_material_final_status(material: dict, decision: dict | None, scan: dict | None = None) -> str:
    return derive_final_status(confidence=material.get("confidence"), decision=decision, scan=scan)


def analytics_page_rows(database: SupabaseExecutor, table: str, principal: Principal, build_query: Callable[[Any], Any], page_size: int) -> list[dict]:
    rows: list[dict] = []
    offset = 0
    while True:
        response = database.execute(lambda client: scoped_query(build_query(client), principal).range(offset, offset + page_size - 1).execute())
        page = response.data or []
        rows.extend(page)
        if len(page) < page_size:
            return rows
        offset += page_size


ANALYTICS_SUMMARY_CACHE: dict[str, tuple[float, dict]] = {}
ANALYTICS_CACHE_TTL = 60.0  # 60 seconds


def clear_analytics_summary_cache():
    ANALYTICS_SUMMARY_CACHE.clear()


def analytics_child_rows(database: SupabaseExecutor, table: str, scan_ids: list[str], principal: Principal) -> list[dict]:
    if not scan_ids:
        return []
    rows: list[dict] = []
    chunk_size = 100
    chunks = [scan_ids[i:i + chunk_size] for i in range(0, len(scan_ids), chunk_size)]

    def fetch_chunk(ids):
        return analytics_page_rows(
            database,
            table,
            principal,
            lambda client, target_ids=ids: client.table(table).select("*").in_("scan_result_id", target_ids),
            ANALYTICS_CHILD_PAGE_SIZE,
        )

    if len(chunks) == 1:
        return fetch_chunk(chunks[0])

    with ThreadPoolExecutor(max_workers=min(10, len(chunks))) as executor:
        futures = [executor.submit(fetch_chunk, chunk) for chunk in chunks]
        for future in as_completed(futures):
            rows.extend(future.result())
    return rows


@app.get("/api/analytics/summary")
def analytics_summary(
    start_date: str | None = None,
    end_date: str | None = None,
    response: Response = None,
    principal: Principal = Depends(require_scope("scan:read")),
):
    """Aggregate every matching scan server-side; never expose a paginated scan page as analytics."""
    if response is not None:
        response.headers["Cache-Control"] = "no-store"
        response.headers["Pragma"] = "no-cache"
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    if bool(start_date) != bool(end_date):
        raise HTTPException(status_code=400, detail="start_date and end_date must be supplied together.")

    cache_key = f"{principal.id}:{principal.role}:{start_date or ''}:{end_date or ''}"
    now_ts = time.time()
    if cache_key in ANALYTICS_SUMMARY_CACHE:
        cached_time, cached_payload = ANALYTICS_SUMMARY_CACHE[cache_key]
        if now_ts - cached_time < ANALYTICS_CACHE_TTL:
            return cached_payload

    try:
        database = SupabaseExecutor(supabase)

        def build_scans(client):
            query = client.table(SCAN_RESULTS_TABLE).select("id,source_name,source_type,batch_id,overall_status,human_review_required,overall_confidence,created_at,reviewed_at,contamination_risk")
            if start_date:
                query = query.gte("created_at", start_date).lt("created_at", end_date)
            return query.order("created_at", desc=True)

        def build_jobs(client):
            query = client.table(JOBS_TABLE).select("id,source,status,processed_count,total_count,attempts,failed_count,created_at,started_at,completed_at,error")
            if start_date:
                query = query.gte("created_at", start_date).lt("created_at", end_date)
            return query.order("created_at", desc=True)

        with ThreadPoolExecutor(max_workers=2) as executor:
            fut_scans = executor.submit(analytics_page_rows, database, SCAN_RESULTS_TABLE, principal, build_scans, ANALYTICS_PAGE_SIZE)
            fut_jobs = executor.submit(analytics_page_rows, database, JOBS_TABLE, principal, build_jobs, ANALYTICS_PAGE_SIZE)
            scans = fut_scans.result()
            jobs = fut_jobs.result()

        scan_ids = [str(scan["id"]) for scan in scans if scan.get("id")]

        if start_date:
            with ThreadPoolExecutor(max_workers=2) as executor:
                fut_materials = executor.submit(analytics_child_rows, database, DETECTED_MATERIALS_TABLE, scan_ids, principal)
                fut_decisions = executor.submit(analytics_child_rows, database, REVIEW_DECISIONS_TABLE, scan_ids, principal)
                materials = fut_materials.result()
                decisions = fut_decisions.result()
        else:
            with ThreadPoolExecutor(max_workers=2) as executor:
                fut_materials = executor.submit(
                    analytics_page_rows,
                    database,
                    DETECTED_MATERIALS_TABLE,
                    principal,
                    lambda client: client.table(DETECTED_MATERIALS_TABLE).select("*"),
                    ANALYTICS_CHILD_PAGE_SIZE,
                )
                fut_decisions = executor.submit(
                    analytics_page_rows,
                    database,
                    REVIEW_DECISIONS_TABLE,
                    principal,
                    lambda client: client.table(REVIEW_DECISIONS_TABLE).select("*"),
                    ANALYTICS_CHILD_PAGE_SIZE,
                )
                materials = fut_materials.result()
                decisions = fut_decisions.result()
        latest_decisions: dict[str, dict] = {}
        for decision in sorted(decisions, key=lambda item: str(item.get("created_at") or "")):
            latest_decisions[str(decision.get("detected_material_id") or "")] = decision

        scan_by_id = {str(scan["id"]): scan for scan in scans if scan.get("id")}
        object_metrics = object_metrics_from_rows(scans, materials, decisions)
        rejected_count = object_metrics["rejected_objects"]
        review_count = object_metrics["needs_review_objects"]
        confirmed_count = object_metrics["confirmed_objects"]
        category_data: dict[str, dict] = {}
        recyclable_counts: dict[str, int] = {}
        contaminant_counts: dict[str, int] = {}
        confidence_values: list[float] = []
        review_durations: list[float] = []
        high_risk_count = 0
        recovery_opportunity_count = 0
        reviewer_stats: dict[str, dict] = {}
        accuracy_by_category: dict[str, dict] = {}

        for material in materials:
            category = analytics_category(material.get("category") or material.get("material_name"))
            estimate = ANALYTICS_MATERIAL_ESTIMATES[category]
            decision = latest_decisions.get(str(material.get("id") or ""))
            scan = scan_by_id.get(str(material.get("scan_result_id") or ""))
            final_status = analytics_material_final_status(material, decision, scan)
            row = category_data.setdefault(category, {"category": category, "label": estimate["label"], "count": 0, "estimatedWeightKg": 0.0, "pricePerKg": estimate["price_per_kg_rm"], "estimatedResaleValueRm": 0.0, "confidence_total": 0.0, "confidence_count": 0, "recyclable_count": 0, "contaminant_count": 0})
            row["count"] += 1
            row["estimatedWeightKg"] += estimate["average_weight_kg"]
            row["estimatedResaleValueRm"] += estimate["average_weight_kg"] * estimate["price_per_kg_rm"]
            raw_confidence = material.get("confidence")
            confidence = None
            if raw_confidence not in (None, ""):
                try:
                    numeric_confidence = float(raw_confidence)
                    if math.isfinite(numeric_confidence):
                        confidence = max(0.0, min(100.0, numeric_confidence * 100 if numeric_confidence <= 1 else numeric_confidence))
                except (TypeError, ValueError):
                    confidence = None
            if confidence is not None:
                confidence_values.append(confidence)
                row["confidence_total"] += confidence
                row["confidence_count"] += 1
            material_class = str((decision or {}).get("disposition") or material.get("material_class") or estimate["material_class"]).lower()
            if final_status == "confirmed" and material_class == "recyclable":
                recyclable_counts[estimate["label"]] = recyclable_counts.get(estimate["label"], 0) + 1
                row["recyclable_count"] += 1
                if final_status == "confirmed" and estimate["price_per_kg_rm"] > 0:
                    recovery_opportunity_count += 1
            elif final_status == "confirmed" and material_class == "contaminant":
                contaminant_counts[estimate["label"]] = contaminant_counts.get(estimate["label"], 0) + 1
                row["contaminant_count"] += 1
                if final_status == "confirmed" and category == "battery":
                    high_risk_count += 1
            if decision and scan:
                created_at, reviewed_at = analytics_timestamp(scan.get("created_at")), analytics_timestamp(decision.get("created_at"))
                if created_at and reviewed_at and reviewed_at >= created_at:
                    review_durations.append((reviewed_at - created_at).total_seconds() * 1000)
            if decision:
                reviewer = decision.get("reviewer_email") or "Unknown reviewer"
                r = reviewer_stats.setdefault(reviewer, {"reviewer_email": reviewer, "reviewed_count": 0, "agree_count": 0, "override_count": 0, "confirmed_count": 0, "rejected_count": 0})
                r["reviewed_count"] += 1
                r["confirmed_count" if decision.get("outcome") == "confirmed" else "rejected_count"] += 1
                ai_category = analytics_category(material.get("original_category") or material.get("category") or material.get("material_name"))
                chosen_category = analytics_category(decision.get("chosen_category"))
                agreed = ai_category == chosen_category
                r["agree_count" if agreed else "override_count"] += 1
                accuracy_row = accuracy_by_category.setdefault(ai_category, {"category": ai_category, "label": ANALYTICS_MATERIAL_ESTIMATES[ai_category]["label"], "reviewed_count": 0, "agree_count": 0})
                accuracy_row["reviewed_count"] += 1
                accuracy_row["agree_count"] += int(agreed)

        trend: dict[str, dict] = {}
        batch_counts: dict[str, int] = {}
        risk_counts: dict[str, int] = {}
        for scan in scans:
            created_at = analytics_timestamp(scan.get("created_at"))
            if created_at:
                local_day = created_at.astimezone(ANALYTICS_MALAYSIA_TZ)
                key = local_day.strftime("%Y-%m-%d")
                trend.setdefault(key, {"key": key, "label": local_day.strftime("%b %-d"), "value": 0})["value"] += 1
            if scan.get("batch_id"):
                batch_counts[str(scan["batch_id"])] = batch_counts.get(str(scan["batch_id"]), 0) + 1
            risk_key = str(scan.get("contamination_risk") or "unknown").lower()
            risk_counts[risk_key] = risk_counts.get(risk_key, 0) + 1

        status_counts: dict[str, int] = {}
        durations_ms: list[float] = []
        jobs_with_retries = 0
        failed_items_total = 0
        for job in jobs:
            status = str(job.get("status") or "unknown")
            status_counts[status] = status_counts.get(status, 0) + 1
            if (job.get("attempts") or 0) > 1:
                jobs_with_retries += 1
            failed_items_total += job.get("failed_count") or 0
            job_started, job_completed = analytics_timestamp(job.get("started_at")), analytics_timestamp(job.get("completed_at"))
            if job_started and job_completed and job_completed >= job_started:
                durations_ms.append((job_completed - job_started).total_seconds() * 1000)

        resale_rows = sorted(category_data.values(), key=lambda row: (-row["estimatedResaleValueRm"], -row["count"]))
        for row in resale_rows:
            row["averageConfidence"] = row.pop("confidence_total") / row.pop("confidence_count") if row["confidence_count"] else 0
        material_mix = sorted(resale_rows, key=lambda row: (-row["estimatedWeightKg"], -row["count"]))
        top_recyclable = max(recyclable_counts.items(), key=lambda item: item[1], default=None)
        top_contaminant = max(contaminant_counts.items(), key=lambda item: item[1], default=None)
        latest_scan = scans[0] if scans else None
        recent_events = [{"timestamp": scan.get("created_at"), "source": scan.get("source_name") or scan.get("source_type") or "Web Upload", "event": "Scan Rejected" if str(scan.get("overall_status") or "").lower() in {"rejected", "quarantined"} else "Scan Verified", "status": "Rejected" if str(scan.get("overall_status") or "").lower() in {"rejected", "quarantined"} else "Review Needed" if scan.get("human_review_required") else "Confirmed", "details": "Saved scan record"} for scan in scans[:5]]
        payload = {
            "scope": "selected_date" if start_date else "all_history",
            "start_date": start_date,
            "end_date": end_date,
            "total_scans": len(scans),
            "confirmed_count": confirmed_count,
            "review_count": review_count,
            "rejected_count": rejected_count,
            "object_metrics": object_metrics,
            **object_metrics,
            "detected_materials_count": len(materials),
            "average_detection_confidence": sum(confidence_values) / len(confidence_values) if confidence_values else None,
            "estimated_recovery_value": sum(row["estimatedResaleValueRm"] for row in resale_rows),
            "total_estimated_weight_kg": sum(row["estimatedWeightKg"] for row in material_mix),
            "material_mix": material_mix,
            "recoverable_value_by_category": resale_rows,
            "daily_scan_trend": [trend[key] for key in sorted(trend)],
            "recyclable_rows": sorted(recyclable_counts.items(), key=lambda item: -item[1]),
            "contaminated_rows": sorted(contaminant_counts.items(), key=lambda item: -item[1]),
            "top_contamination_source": {"label": top_contaminant[0], "count": top_contaminant[1]} if top_contaminant else None,
            "top_recyclable_material": {"label": top_recyclable[0], "count": top_recyclable[1]} if top_recyclable else None,
            "average_review_turnaround_ms": sum(review_durations) / len(review_durations) if review_durations else None,
            "highest_value_category": resale_rows[0] if resale_rows and resale_rows[0]["estimatedResaleValueRm"] > 0 else None,
            "last_upload": latest_scan,
            "last_upload_batch_count": batch_counts.get(str(latest_scan.get("batch_id")), 1) if latest_scan else 0,
            "high_risk_count": high_risk_count,
            "recovery_opportunity_count": recovery_opportunity_count,
            "recent_events": recent_events,
            "reviewer_activity": sorted(reviewer_stats.values(), key=lambda row: -row["reviewed_count"]),
            "upload_pipeline_health": {
                "total_jobs": len(jobs),
                "status_counts": status_counts,
                "jobs_with_retries": jobs_with_retries,
                "failed_items_total": failed_items_total,
                "average_processing_duration_ms": sum(durations_ms) / len(durations_ms) if durations_ms else None,
                "recent_jobs": [{"id": job.get("id"), "source": job.get("source"), "status": job.get("status"), "processed_count": job.get("processed_count"), "total_count": job.get("total_count"), "attempts": job.get("attempts"), "created_at": job.get("created_at")} for job in jobs[:5]],
            },
            "risk_severity_breakdown": [{"risk": risk, "count": count} for risk, count in sorted(risk_counts.items(), key=lambda item: -item[1])],
            "ai_accuracy_by_category": sorted(({**row, "accuracy_pct": (row["agree_count"] / row["reviewed_count"] * 100) if row["reviewed_count"] else 0} for row in accuracy_by_category.values()), key=lambda row: -row["reviewed_count"]),
        }
        ANALYTICS_SUMMARY_CACHE[cache_key] = (now_ts, payload)
        return payload
    except SupabaseTemporarilyUnavailable:
        return JSONResponse(status_code=503, content={"detail": "Analytics data is temporarily unavailable.", "retryable": True}, headers={"Retry-After": "2"})
    except Exception as exc:
        print(f"[analytics] summary failed: {safe_error_message(exc)}")
        raise HTTPException(status_code=500, detail="Unable to load analytics summary.") from exc


@app.post("/api/scans/browser-detected")
async def save_browser_detected_scan(
    file: UploadFile = File(...),
    submission_id: UUID = Form(...),
    original_width: int = Form(...),
    original_height: int = Form(...),
    model_name: str = Form(...),
    model_version: str = Form(...),
    inference_engine: str = Form(...),
    confidence_threshold: float = Form(...),
    nms_iou_threshold: float = Form(...),
    detections: str = Form(...),
    principal: Principal = Depends(require_scope("scan:write")),
):
    if model_name != BROWSER_MODEL_NAME or model_version != BROWSER_MODEL_VERSION:
        raise HTTPException(status_code=400, detail="Browser model identity does not match the fixed contract.")
    if inference_engine != BROWSER_INFERENCE_ENGINE:
        raise HTTPException(status_code=400, detail="Inference engine must be browser-onnx.")
    if not math.isclose(confidence_threshold, BROWSER_DECISION_CONFIDENCE_THRESHOLD, abs_tol=1e-9):
        raise HTTPException(status_code=400, detail=BROWSER_CONFIDENCE_CONTRACT_DETAIL)
    if not math.isclose(nms_iou_threshold, BROWSER_NMS_IOU_THRESHOLD, abs_tol=1e-9):
        raise HTTPException(status_code=400, detail=BROWSER_NMS_CONTRACT_DETAIL)
    file_bytes = await file.read()
    if not file_bytes or len(file_bytes) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="Image must be non-empty and no larger than 10 MB.")
    _, normalized_type = _image_content_type(file.filename, file.content_type)
    try:
        with Image.open(BytesIO(file_bytes)) as image:
            actual_width, actual_height = image.size
            image.verify()
    except (UnidentifiedImageError, OSError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="Uploaded image is invalid or corrupted.") from exc
    if (original_width, original_height) != (actual_width, actual_height):
        raise HTTPException(status_code=400, detail="Browser image dimensions do not match the uploaded image.")
    try:
        raw_detections = json.loads(detections)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Detection JSON is invalid.") from exc
    materials = validate_browser_detected_detections(raw_detections, actual_width, actual_height)
    preview_bytes = _encode_annotated_image_preview(file_bytes, file.filename, materials)
    return persist_scan(
        preview_bytes,
        file.filename,
        "image",
        materials,
        summarize(materials),
        source_ref="browser-onnx:best.onnx",
        principal=principal,
        content_type=normalized_type,
        scan_result_id=submission_id,
        model_version=model_version,
    )


@app.post("/api/scans/browser-verified")
async def save_browser_verified_scan(
    file: UploadFile = File(...),
    submission_id: UUID = Form(...),
    original_width: int = Form(...),
    original_height: int = Form(...),
    model_name: str = Form(...),
    model_version: str = Form(...),
    inference_engine: str = Form(...),
    confidence_threshold: float = Form(...),
    nms_iou_threshold: float = Form(...),
    verified_detections: str = Form(...),
    verification_outcome: str = Form(...),
    principal: Principal = Depends(require_scope("scan:write")),
):
    if model_name != BROWSER_MODEL_NAME or model_version != BROWSER_MODEL_VERSION:
        raise HTTPException(status_code=400, detail="Browser model identity does not match the fixed contract.")
    if inference_engine != BROWSER_INFERENCE_ENGINE:
        raise HTTPException(status_code=400, detail="Inference engine must be browser-onnx.")
    if not math.isclose(confidence_threshold, BROWSER_DECISION_CONFIDENCE_THRESHOLD, abs_tol=1e-9):
        raise HTTPException(status_code=400, detail=BROWSER_CONFIDENCE_CONTRACT_DETAIL)
    if not math.isclose(nms_iou_threshold, BROWSER_NMS_IOU_THRESHOLD, abs_tol=1e-9):
        raise HTTPException(status_code=400, detail=BROWSER_NMS_CONTRACT_DETAIL)
    if verification_outcome != "verified":
        raise HTTPException(status_code=400, detail="Browser detections must be human verified before saving.")

    file_bytes = await file.read()
    if not file_bytes or len(file_bytes) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="Image must be non-empty and no larger than 10 MB.")
    _, normalized_type = _image_content_type(file.filename, file.content_type)
    try:
        with Image.open(BytesIO(file_bytes)) as image:
            actual_width, actual_height = image.size
            image.verify()
    except (UnidentifiedImageError, OSError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="Uploaded image is invalid or corrupted.") from exc
    if actual_width <= 0 or actual_height <= 0:
        raise HTTPException(status_code=400, detail="Uploaded image dimensions are invalid.")
    if (original_width, original_height) != (actual_width, actual_height):
        raise HTTPException(status_code=400, detail="Browser image dimensions do not match the uploaded image.")

    try:
        raw_detections = json.loads(verified_detections)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Verified detection JSON is invalid.") from exc
    materials = validate_browser_detections(raw_detections, actual_width, actual_height)
    summary = summarize(materials)
    summary.update({
        "overall_status": "verified",
        "human_review_required": False,
        "recommended_action": "Verified after operator review.",
    })
    print(
        "[PurityLoop inference]\n"
        "engine=browser-onnx\n"
        "model=best.onnx\n"
        "source=upload-page"
    )
    preview_bytes = _encode_annotated_image_preview(file_bytes, file.filename, materials)
    return persist_scan(
        preview_bytes,
        file.filename,
        "image",
        materials,
        summary,
        source_ref="browser-onnx:best.onnx",
        principal=principal,
        content_type=normalized_type,
        scan_result_id=submission_id,
        model_version=model_version,
        verified=True,
    )


@app.post("/api/reviews")
def create_review(decision: ReviewDecisionInput, principal: Principal = Depends(require_scope("review:write"))):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    review_payload = {
        "scan_result_id": str(decision.scan_result_id),
        "detected_material_id": str(decision.detected_material_id),
        "action": decision.action,
        "manual_category": decision.manual_category,
        "principal": principal.id,
    }
    print(f"[review] incoming payload: {review_payload}")
    action = str(decision.action or "").strip().lower()
    if action not in {"verify", "reject"}:
        raise HTTPException(status_code=400, detail="Review action must be 'verify' or 'reject'.")
    try:
        scan_result_id = str(decision.scan_result_id)
        detected_material_id = str(decision.detected_material_id)
        print(f"[review] action={action} scan_result_id={scan_result_id} selected_category={decision.manual_category}")
        print(f"[review] selecting {SCAN_RESULTS_TABLE} where id={scan_result_id}")
        scan_response = scoped_query(supabase.table(SCAN_RESULTS_TABLE).select("*").eq("id", scan_result_id), principal).execute()
        print(f"[review] selecting {DETECTED_MATERIALS_TABLE} where id={detected_material_id}, scan_result_id={scan_result_id}")
        material_response = supabase.table(DETECTED_MATERIALS_TABLE).select("*").eq("id", detected_material_id).eq("scan_result_id", scan_result_id).execute()
        if not scan_response.data or not material_response.data:
            raise HTTPException(status_code=404, detail="Scan result or detected material was not found.")

        material = material_response.data[0]
        category = material_category(decision.manual_category or material.get("category"))
        disposition = CATEGORY_CLASS_MAP.get(category)
        if category == "unknown" or disposition not in {"recyclable", "contaminant"}:
            raise HTTPException(status_code=400, detail="Manual category is not supported.")
        print(f"[review] resolved category={category} disposition={disposition}")

        updated_material = material
        if action == "verify" and category != material.get("category"):
            material_update = {"category": category}
            if not material.get("original_category"):
                material_update["original_category"] = material.get("category")
            print(f"[review] updating {DETECTED_MATERIALS_TABLE}: {material_update}")
            update_response = supabase.table(DETECTED_MATERIALS_TABLE).update(material_update).eq("id", detected_material_id).eq("scan_result_id", scan_result_id).execute()
            updated_material = update_response.data[0] if update_response.data else {**material, **material_update}

        outcome = "confirmed" if action == "verify" else "rejected"
        decision_insert = {
            "scan_result_id": scan_result_id,
            "detected_material_id": detected_material_id,
            "chosen_category": category,
            "disposition": disposition,
            "outcome": outcome,
            "reviewer_email": principal.id if principal.kind == "user" else f"api:{principal.id}",
        }
        print(f"[review] inserting {REVIEW_DECISIONS_TABLE}: {decision_insert}")
        inserted = supabase.table(REVIEW_DECISIONS_TABLE).insert(decision_insert).execute()
        scan_update = {
            "overall_status": "verified" if action == "verify" else "rejected",
            "human_review_required": False,
            "recommended_action": "Verified after operator review." if action == "verify" else "Rejected after operator review.",
            "review_status": "verified" if action == "verify" else "rejected",
            "reviewed_at": datetime.now(timezone.utc).isoformat(),
        }
        if action == "verify":
            scan_update["verified_category"] = category
        print(f"[review] updating {SCAN_RESULTS_TABLE}: {scan_update}")
        updated_scan_response = supabase.table(SCAN_RESULTS_TABLE).update(scan_update).eq("id", scan_result_id).execute()
        if not updated_scan_response.data:
            raise HTTPException(status_code=500, detail="Review was saved, but the scan result status could not be updated.")
        clear_analytics_summary_cache()
        return {
            "decision": inserted.data[0] if inserted.data else None,
            "material": updated_material,
            "scan_result": updated_scan_response.data[0],
            **updated_scan_response.data[0],
        }
    except HTTPException:
        raise
    except Exception as exc:
        traceback.print_exc()
        print(f"[review] Supabase update failed: {safe_error_message(exc)}")
        detail = safe_error_message(exc)
        if getattr(exc, "code", "") in {"PGRST205", "42703"}:
            detail = f"Supabase review schema is missing or outdated: {detail}"
        if os.getenv("ENVIRONMENT", "development").lower() in {"development", "dev", "local"}:
            raise HTTPException(status_code=500, detail=detail) from exc
        raise HTTPException(status_code=500, detail="Unable to save review. Check the backend Supabase configuration.") from exc


@app.get("/api/scans")
def get_scan_history(
    limit: int = SCAN_HISTORY_DEFAULT_LIMIT,
    offset: int = 0,
    start_date: str | None = None,
    end_date: str | None = None,
    search: str | None = None,
    category: str | None = None,
    status: str | None = None,
    sort: str = "timestamp",
    direction: str = "desc",
    principal: Principal = Depends(require_scope("scan:read")),
):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    try:
        limit = max(1, min(int(limit), SCAN_HISTORY_MAX_LIMIT))
        offset = max(0, int(offset))
        filters = scan_history_filters(start_date, end_date, search, status)
        normalized_status = filters["status"] or ""
        descending = str(direction).lower() != "asc"
        category_key = canonical_category_key(category) if category else ""
        if category and category_key == "unknown":
            return {"items": [], "total": 0, "limit": limit, "offset": offset, "start_date": start_date, "end_date": end_date, "search": search, "category": category, "status": status, "sort": "confidence" if sort == "confidence" else "timestamp", "direction": "desc" if descending else "asc", "summary": {"confirmed": 0, "needs_review": 0, "rejected": 0, "total_objects": 0, "confirmed_objects": 0, "needs_review_objects": 0, "rejected_objects": 0}}

        database = SupabaseExecutor(client_factory=_new_supabase_client, attempts=2)
        rpc_params = {
            "p_limit": limit,
            "p_offset": offset,
            "p_start_date": filters["start_date"],
            "p_end_date": filters["end_date"],
            "p_search": filters["search"],
            "p_category_key": category_key or None,
            "p_status": normalized_status or None,
            "p_sort": "confidence" if sort == "confidence" else "timestamp",
            "p_direction": "asc" if not descending else "desc",
        }
        page_response = database.execute(lambda client: scoped_query(client.rpc("scan_history_page", rpc_params), principal).execute())
        scans, total, summary_metrics = scan_history_page_from_rpc(page_response.data or [])
        scan_ids = [str(scan.get("id")) for scan in scans if scan.get("id")]
        if scan_ids:
            materials = database.execute(lambda client: client.table(DETECTED_MATERIALS_TABLE).select("*").in_("scan_result_id", scan_ids).execute()).data or []
            decisions = database.execute(lambda client: client.table(REVIEW_DECISIONS_TABLE).select("*").in_("scan_result_id", scan_ids).execute()).data or []
        else:
            materials = []
            decisions = []
        items = attach_scan_children(scans, materials, decisions)
        return {
            "items": items,
            "total": total,
            "limit": limit,
            "offset": offset,
            "start_date": start_date,
            "end_date": end_date,
            "search": search,
            "category": category,
            "status": status,
            "sort": "confidence" if sort == "confidence" else "timestamp",
            "direction": "desc" if descending else "asc",
            "summary": summary_metrics,
        }
    except Exception as exc:
        if not isinstance(exc, (SupabaseTemporarilyUnavailable, RuntimeError)):
            print(f"[scans] history response failed: {type(exc).__name__}: {safe_error_message(exc)}")
        raise HTTPException(status_code=500, detail="Unable to load scan history.") from exc


@app.get("/api/history/export")
def export_scan_history(
    format: str = "excel",
    scope: str = "audit",
    start_date: str | None = None,
    end_date: str | None = None,
    search: str | None = None,
    category: str | None = None,
    status: str | None = None,
    sort: str = "timestamp",
    direction: str = "desc",
    principal: Principal = Depends(require_scope("scan:read")),
):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    export_format = str(format or "").strip().lower()
    if export_format not in {"pdf", "excel"}:
        raise HTTPException(status_code=400, detail="Export format must be pdf or excel.")
    normalized_scope = "scan" if str(scope or "").strip().lower() == "scan" else "audit"
    filters = {
        "scope": normalized_scope,
        "search": search.strip() if search and search.strip() else None,
        "start_date": start_date,
        "end_date": end_date,
        "category": category,
        "status": status,
        "sort": "confidence" if sort == "confidence" else "timestamp",
        "direction": "asc" if str(direction).lower() == "asc" else "desc",
    }
    try:
        started = time.perf_counter()
        rows = fetch_history_export_rows(
            start_date=start_date,
            end_date=end_date,
            search=search,
            category=category,
            status=status,
            sort=filters["sort"] or "timestamp",
            direction=filters["direction"] or "desc",
            principal=principal,
        )
        print(f"[history-export] records fetched rows={len(rows)}")
        thumbnail_started = time.perf_counter()
        thumbnails, thumbnail_stats = fetch_history_thumbnails(rows)
        print(
            "[history-export] images processed "
            f"requested={thumbnail_stats['requested']} unique={thumbnail_stats['unique']} "
            f"cache_hits={thumbnail_stats['cache_hits']} failed={thumbnail_stats['failed']} "
            f"duration={time.perf_counter() - thumbnail_started:.2f}s"
        )
        stamp = datetime.now(ANALYTICS_MALAYSIA_TZ).strftime("%Y%m%d-%H%M%S")
        prefix = "purityloop-scan-history" if normalized_scope == "scan" else "purityloop-audit-history"
        if export_format == "excel":
            content = build_history_excel(rows, thumbnails)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"{prefix}-{stamp}.xlsx"
        else:
            content = build_history_pdf(rows, filters, thumbnails)
            media_type = "application/pdf"
            filename = f"{prefix}-{stamp}.pdf"
        print(f"[history-export] report generated format={export_format} bytes={len(content)}")
        print(f"[history-export] {export_format} scope={normalized_scope} rows={len(rows)} duration={time.perf_counter() - started:.2f}s")
        return Response(
            content=content,
            media_type=media_type,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except ImportError as exc:
        print(f"[history-export] dependency missing: {safe_error_message(exc)}")
        raise HTTPException(status_code=500, detail="History export dependency is not installed.") from exc
    except Exception as exc:
        print(f"[history-export] failed: {type(exc).__name__}: {safe_error_message(exc)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Unable to export history.") from exc


@app.get("/api/scans/{scan_result_id}")
def get_scan_result(scan_result_id: str, principal: Principal = Depends(require_scope("scan:read"))):
    """Return the persisted material IDs needed to review a previously loaded scan."""
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase backend env is not configured.")
    try:
        UUID(scan_result_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail="Scan result was not found.") from exc
    try:
        scan_response = execute_scan_read(
            "selected scan query",
            lambda client: scoped_query(client.table(SCAN_RESULTS_TABLE).select("*").eq("id", scan_result_id), principal).execute(),
        )
        if not scan_response.data:
            raise HTTPException(status_code=404, detail="Scan result was not found.")
        materials = execute_scan_read(
            "selected scan materials query",
            lambda client: client.table(DETECTED_MATERIALS_TABLE).select("*").eq("scan_result_id", scan_result_id).execute(),
        ).data or []
        decisions = execute_scan_read(
            "selected scan decisions query",
            lambda client: client.table(REVIEW_DECISIONS_TABLE).select("*").eq("scan_result_id", scan_result_id).execute(),
        ).data or []
    except HTTPException:
        raise
    except Exception as exc:
        print(f"[scans] selected scan query failed for {scan_result_id[:8]}…: {type(exc).__name__}: {safe_error_message(exc)}")
        raise HTTPException(status_code=500, detail="Unable to load selected scan.") from exc
    latest_decisions = {}
    for item in sorted(decisions, key=lambda entry: str(entry.get("created_at", ""))):
        latest_decisions[str(item.get("detected_material_id", ""))] = item
    return {
        "scan_result": {
            **scan_response.data[0],
            "detected_materials": [
                {**material, "review_decision": latest_decisions.get(str(material.get("id", "")))}
                for material in materials
            ],
        }
    }


@app.get("/api/google/auth")
def google_auth(credentials: HTTPAuthorizationCredentials | None = Depends(bearer_scheme)):
    if is_production():
        raise HTTPException(status_code=404, detail="Not found.")
    require_scope("scan:write")(require_principal(credentials))
    try:
        flow = oauth_flow()
        authorization_url, state = flow.authorization_url(
            access_type="offline",
            prompt="consent",
            include_granted_scopes="true",
        )
        save_oauth_state(state, flow.code_verifier)
    except Exception as exc:
        print(f"[google-auth] OAuth start failed: {type(exc).__name__}: {safe_error_message(exc)}")
        raise HTTPException(status_code=500, detail="Google OAuth is not configured.") from exc
    return RedirectResponse(authorization_url)


@app.get("/api/google/callback")
def google_callback(code: str | None = None, state: str | None = None, error: str | None = None):
    if is_production():
        raise HTTPException(status_code=404, detail="Not found.")
    if error:
        raise HTTPException(status_code=400, detail="Google OAuth authorization failed.")
    if not code:
        raise HTTPException(status_code=400, detail="Missing Google OAuth code.")

    try:
        saved_state = load_oauth_state(state)
        flow = oauth_flow()
        flow.code_verifier = saved_state["code_verifier"]
        flow.fetch_token(code=code)
        save_oauth_token(flow.credentials)
    except HTTPException:
        raise
    except Exception as exc:
        print(f"[google-callback] OAuth token exchange failed: {type(exc).__name__}: {safe_error_message(exc)}")
        raise HTTPException(status_code=500, detail="Unable to save Google OAuth token.") from exc
    return "Google Drive connected. You can now upload images."


@app.post("/api/predict")
async def predict(file: UploadFile = File(...), principal: Principal = Depends(require_scope("scan:write"))):
    file_bytes = await file.read()
    return run_scan(file_bytes, file.filename, "image", principal=principal, content_type=file.content_type)

@app.get("/health")
def health_check():
    model_error = None
    try:
        get_model()
    except Exception as exc:
        model_error = safe_error_message(exc)
    return {
        "status": "ok",
        "service": "purityloop-backend",
        "backend_build_version": BACKEND_BUILD_VERSION,
        **safe_startup_diagnostics(),
        "model_error": model_error,
    }
