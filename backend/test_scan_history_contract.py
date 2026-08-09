import unittest
from contextlib import ExitStack
from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

import httpx
from fastapi import HTTPException
from openpyxl import load_workbook
from PIL import Image

from backend import main


def fake_principal():
    return main.Principal("user", "11111111-1111-4111-8111-111111111111", frozenset({"scan:read", "review:write"}))


class FakeQuery:
    def __init__(self, table, rows):
        self.table = table
        self.rows = rows
        self.count = None
        self.head = None
        self.columns = None
        self.filters = {}
        self.range_args = None
        self.order_args = None
        self.limit_args = None

    def select(self, *_args, **kwargs):
        self.columns = _args
        self.count = kwargs.get("count")
        self.head = kwargs.get("head")
        return self

    def order(self, *args, **kwargs):
        self.order_args = (args, kwargs)
        return self

    def range(self, *args):
        self.range_args = args
        return self

    def limit(self, *args):
        self.limit_args = args
        return self

    def in_(self, *_args):
        self.filters["in"] = _args
        return self

    def eq(self, field, value):
        self.filters[field] = value
        return self

    def neq(self, field, value):
        self.filters[f"{field}!="] = value
        return self

    def gte(self, field, value):
        self.filters[f"gte:{field}"] = value
        return self

    def lt(self, field, value):
        self.filters[f"lt:{field}"] = value
        return self

    def ilike(self, field, value):
        self.filters[field] = value
        return self

    def execute(self):
        rows = self.rows
        if self.filters.get("id") == "missing":
            rows = []
        elif self.filters.get("id"):
            rows = [row for row in rows if row.get("id") == self.filters["id"]]
        if self.filters.get("in"):
            field, values = self.filters["in"]
            allowed = {str(value) for value in values}
            rows = [row for row in rows if str(row.get(field)) in allowed]
        if "source_type!=" in self.filters:
            rows = [row for row in rows if row.get("source_type") != self.filters["source_type!="]]
        count = len(rows) if self.table == main.SCAN_RESULTS_TABLE and self.count == "exact" else None
        if self.range_args:
            start, end = self.range_args
            rows = rows[start:end + 1]
        if self.limit_args:
            rows = rows[:self.limit_args[0]]
        return SimpleNamespace(data=rows, count=count)


class FakeRpc:
    def __init__(self, fake, params):
        self.fake = fake
        self.params = params

    def execute(self):
        rows = list(self.fake.scan_rows)
        rows = [row for row in rows if row.get("source_type") != "video_frame"]
        search = self.params.get("p_search")
        if search:
            rows = [row for row in rows if search.lower() in str(row.get("source_name") or "").lower()]
        if self.params.get("p_status") == "review_needed":
            rows = [row for row in rows if row.get("human_review_required") is True]
        elif self.params.get("p_status") == "confirmed":
            rows = [row for row in rows if row.get("human_review_required") is False]
        total = len(rows)
        offset = max(0, self.params.get("p_offset") or 0)
        limit = max(1, self.params.get("p_limit") or 10)
        page = rows[offset:offset + limit]
        rows = [
            {
                "scan": row,
                "total_count": total,
                "total_objects": total,
                "confirmed_objects": total,
                "needs_review_objects": 0,
                "rejected_objects": 0,
            }
            for row in page
        ]
        if not rows:
            rows = [{"scan": None, "total_count": total, "total_objects": total, "confirmed_objects": total, "needs_review_objects": 0, "rejected_objects": 0}]
        return SimpleNamespace(data=rows, count=None)


class FakeSupabase:
    def __init__(self, scan_count=25):
        self.queries = []
        self.rpc_calls = []
        self.scan_count = scan_count
        self.scan_rows = [
            {
                "id": f"scan-{index}",
                "source_name": f"bottle-{index}.jpg",
                "source_type": "image",
                "created_at": "2026-07-19T00:00:00Z",
                "overall_confidence": 0.91,
                "human_review_required": False,
                "preview_image_url": f"https://example.test/{index}.jpg",
            }
            for index in range(self.scan_count)
        ]

    def table(self, table):
        if table == main.SCAN_RESULTS_TABLE:
            rows = self.scan_rows
        elif table == main.DETECTED_MATERIALS_TABLE:
            rows = [
                {
                    "id": f"material-{index}",
                    "scan_result_id": f"scan-{index}",
                    "category": "Plastic",
                    "confidence": 0.91,
                    "material_class": "recyclable",
                    "estimated_weight_kg": 0.032,
                    "created_at": "2026-07-19T00:00:00Z",
                }
                for index in range(self.scan_count)
            ]
        elif table == main.REVIEW_DECISIONS_TABLE:
            rows = [
                {
                    "id": f"decision-{index}",
                    "scan_result_id": f"scan-{index}",
                    "detected_material_id": f"material-{index}",
                    "chosen_category": "Plastic",
                    "outcome": "confirmed",
                    "reviewer_id": "operator",
                    "created_at": "2026-07-20T00:00:00Z",
                }
                for index in range(self.scan_count)
            ]
        else:
            rows = []
        query = FakeQuery(table, rows)
        self.queries.append(query)
        return query

    def rpc(self, name, params):
        self.rpc_calls.append((name, params))
        return FakeRpc(self, params)


class ScanHistoryContractTests(unittest.TestCase):
    def fake_backend(self, fake):
        stack = ExitStack()
        stack.enter_context(patch.object(main, "supabase", fake))
        stack.enter_context(patch.object(main, "_new_supabase_client", return_value=fake))
        return stack

    def test_scan_history_returns_one_page_and_exact_total(self):
        fake = FakeSupabase()
        with self.fake_backend(fake):
            payload = main.get_scan_history(limit=10, offset=0, principal=fake_principal())

        self.assertEqual(payload["total"], 25)
        self.assertEqual(payload["limit"], 10)
        self.assertEqual(payload["offset"], 0)
        self.assertEqual(len(payload["items"]), 10)
        self.assertEqual(fake.rpc_calls[0][0], "scan_history_page")
        self.assertEqual(fake.rpc_calls[0][1]["p_limit"], 10)
        self.assertEqual(fake.rpc_calls[0][1]["p_offset"], 0)
        self.assertEqual(fake.rpc_calls[0][1]["p_category_key"], None)
        self.assertEqual(fake.queries[0].filters["in"][1], [f"scan-{index}" for index in range(10)])
        self.assertEqual(payload["summary"], {
            "confirmed": 25,
            "needs_review": 0,
            "rejected": 0,
            "total_objects": 25,
            "confirmed_objects": 25,
            "needs_review_objects": 0,
            "rejected_objects": 0,
        })
        self.assertNotIn("scans", payload)

    def test_scan_history_filter_excludes_video_frame_rows(self):
        query = FakeQuery(main.SCAN_RESULTS_TABLE, [])

        main.apply_scan_history_filters(query, {"start_date": None, "end_date": None, "search": None, "status": None})

        self.assertEqual(query.filters["source_type!="], "video_frame")

    def test_scan_history_filter_keeps_image_and_tracked_video_rows(self):
        query = FakeQuery(main.SCAN_RESULTS_TABLE, [
            {"id": "image", "source_type": "image"},
            {"id": "tracked", "source_type": "tracked_video"},
            {"id": "frame", "source_type": "video_frame"},
        ])

        main.apply_scan_history_filters(query, {"start_date": None, "end_date": None, "search": None, "status": None})

        self.assertEqual([row["id"] for row in query.execute().data], ["image", "tracked"])

    def test_scan_history_endpoint_excludes_video_frame_rows(self):
        fake = FakeSupabase(scan_count=0)
        fake.scan_rows = [
            {"id": "image", "source_name": "image.jpg", "source_type": "image", "human_review_required": False},
            {"id": "tracked", "source_name": "tracked.mp4", "source_type": "tracked_video", "human_review_required": False},
            {"id": "frame", "source_name": "frame.jpg", "source_type": "video_frame", "human_review_required": False},
        ]
        with self.fake_backend(fake):
            payload = main.get_scan_history(limit=10, offset=0, principal=fake_principal())

        self.assertEqual(payload["total"], 2)
        self.assertEqual([row["id"] for row in payload["items"]], ["image", "tracked"])

    def test_scan_history_review_page_uses_ten_row_ranges(self):
        fake = FakeSupabase()
        with self.fake_backend(fake):
            payload = main.get_scan_history(limit=10, offset=10, principal=fake_principal())

        self.assertEqual(payload["total"], 25)
        self.assertEqual(payload["limit"], 10)
        self.assertEqual(payload["offset"], 10)
        self.assertEqual(len(payload["items"]), 10)
        self.assertEqual(fake.rpc_calls[0][1]["p_limit"], 10)
        self.assertEqual(fake.rpc_calls[0][1]["p_offset"], 10)
        self.assertEqual(fake.queries[0].filters["in"][1], [f"scan-{index}" for index in range(10, 20)])
        self.assertEqual(payload["summary"]["total_objects"], 25)

    def test_scan_history_accepts_review_filters_and_confidence_sort(self):
        fake = FakeSupabase()
        with self.fake_backend(fake):
            payload = main.get_scan_history(
                limit=10, offset=10, search="bottle", status="review_needed", sort="confidence", direction="asc",
                principal=fake_principal(),
            )

        self.assertEqual(payload["search"], "bottle")
        self.assertEqual(payload["status"], "review_needed")
        self.assertEqual(payload["sort"], "confidence")
        self.assertEqual(payload["direction"], "asc")
        self.assertEqual(fake.rpc_calls[0][1]["p_search"], "bottle")
        self.assertEqual(fake.rpc_calls[0][1]["p_status"], "review_needed")
        self.assertEqual(fake.rpc_calls[0][1]["p_sort"], "confidence")
        self.assertEqual(fake.rpc_calls[0][1]["p_direction"], "asc")
        self.assertEqual(payload["total"], 0)
        self.assertEqual(payload["items"], [])
        self.assertEqual(payload["summary"]["total_objects"], 0)
        self.assertEqual(fake.queries, [])

    def test_scan_history_review_page_three_uses_offset_twenty(self):
        fake = FakeSupabase()
        with self.fake_backend(fake):
            payload = main.get_scan_history(limit=10, offset=20, principal=fake_principal())

        self.assertEqual(payload["total"], 25)
        self.assertEqual(len(payload["items"]), 5)
        self.assertEqual(fake.rpc_calls[0][1]["p_limit"], 10)
        self.assertEqual(fake.rpc_calls[0][1]["p_offset"], 20)
        self.assertEqual(fake.queries[0].filters["in"][1], [f"scan-{index}" for index in range(20, 25)])

    def test_scan_history_offset_beyond_total_keeps_exact_total_without_child_queries(self):
        fake = FakeSupabase()
        with self.fake_backend(fake):
            payload = main.get_scan_history(limit=10, offset=30, principal=fake_principal())

        self.assertEqual(payload["total"], 25)
        self.assertEqual(payload["items"], [])
        self.assertEqual(fake.rpc_calls[0][1]["p_offset"], 30)
        self.assertEqual(fake.queries, [])

    def test_final_category_filter_prefers_verified_then_reviewed_category(self):
        scans = [
            {"id": "verified", "verified_category": "Glass"},
            {"id": "reviewed"},
            {"id": "predicted"},
        ]
        materials = [
            {"id": "material-1", "scan_result_id": "verified", "category": "Plastic"},
            {"id": "material-2", "scan_result_id": "reviewed", "category": "Glass"},
            {"id": "material-3", "scan_result_id": "predicted", "category": "PET Bottle"},
        ]
        decisions = [{"detected_material_id": "material-2", "chosen_category": "Plastic", "created_at": "2026-07-28T00:00:00Z"}]

        filtered = main.filter_scans_by_final_category(scans, materials, decisions, "plastic")

        self.assertEqual([scan["id"] for scan in filtered], ["reviewed", "predicted"])

    def test_scan_history_accepts_category_filter(self):
        fake = FakeSupabase()
        with self.fake_backend(fake):
            payload = main.get_scan_history(limit=10, offset=0, category="plastic", principal=fake_principal())

        self.assertEqual(payload["category"], "plastic")
        self.assertEqual(payload["total"], 25)
        self.assertEqual(len(payload["items"]), 10)
        self.assertEqual(payload["summary"]["total_objects"], 25)
        self.assertEqual(fake.rpc_calls[0][1]["p_category_key"], "plastic")

    def test_category_summary_filters_material_rows_not_whole_scans(self):
        scans = [{"id": "scan-1"}]
        materials = [
            {"id": "material-1", "scan_result_id": "scan-1", "category": "Plastic", "confidence": 0.9},
            {"id": "material-2", "scan_result_id": "scan-1", "category": "Glass", "confidence": 0.9},
        ]

        filtered = main.filter_materials_by_final_category(materials, scans, [], "plastic")

        self.assertEqual([material["id"] for material in filtered], ["material-1"])

    def test_scan_history_search_filter_applies_to_list_scope(self):
        fake = FakeSupabase()
        with self.fake_backend(fake):
            main.get_scan_history(limit=10, offset=0, search="missing", principal=fake_principal())

        self.assertEqual(fake.rpc_calls[0][1]["p_search"], "missing")

    def test_scan_lookup_rejects_invalid_uuid_safely(self):
        fake = FakeSupabase()
        with self.fake_backend(fake):
            with self.assertRaises(HTTPException) as raised:
                main.get_scan_result("not-a-uuid", principal=fake_principal())

        self.assertEqual(raised.exception.status_code, 404)
        self.assertFalse(fake.queries)

    def test_scan_read_retries_one_transient_error_with_fresh_client(self):
        class FlakyClient:
            def __init__(self):
                self.calls = 0

        first = FlakyClient()
        second = FlakyClient()
        clients = [first, second]

        def operation(client):
            client.calls += 1
            if client is first:
                raise httpx.RemoteProtocolError("Server disconnected")
            return "ok"

        with patch.object(main, "_new_supabase_client", side_effect=clients):
            self.assertEqual(main.execute_scan_read("test query", operation), "ok")

        self.assertEqual(first.calls, 1)
        self.assertEqual(second.calls, 1)

    def test_history_export_rejects_unknown_format(self):
        fake = FakeSupabase()
        with self.fake_backend(fake):
            with self.assertRaises(HTTPException) as raised:
                main.export_scan_history(format="csv", principal=fake_principal())

        self.assertEqual(raised.exception.status_code, 400)
        self.assertFalse(fake.queries)

    def test_history_export_applies_filters_and_batches_without_n_plus_one(self):
        fake = FakeSupabase(scan_count=1201)
        with self.fake_backend(fake), patch.object(main, "build_history_excel", return_value=b"xlsx"):
            response = main.export_scan_history(
                format="excel",
                start_date="2026-07-19T00:00:00+00:00",
                end_date="2026-07-20T00:00:00+00:00",
                search="bottle",
                category="plastic",
                status="confirmed",
                sort="confidence",
                direction="asc",
                principal=fake_principal(),
            )

        self.assertEqual(response.media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("purityloop-audit-history", response.headers["content-disposition"])
        scan_queries = [query for query in fake.queries if query.table == main.SCAN_RESULTS_TABLE and query.range_args]
        material_queries = [query for query in fake.queries if query.table == main.DETECTED_MATERIALS_TABLE]
        decision_queries = [query for query in fake.queries if query.table == main.REVIEW_DECISIONS_TABLE]
        self.assertEqual([query.range_args for query in scan_queries], [(0, 499), (500, 999), (1000, 1499)])
        self.assertEqual(len(material_queries), 8)
        self.assertEqual(len(decision_queries), 8)
        self.assertEqual(scan_queries[0].filters["source_name"], "%bottle%")
        self.assertIs(scan_queries[0].filters["human_review_required"], False)
        self.assertEqual(scan_queries[0].filters["gte:created_at"], "2026-07-19T00:00:00+00:00")
        self.assertEqual(scan_queries[0].filters["lt:created_at"], "2026-07-20T00:00:00+00:00")
        self.assertEqual(scan_queries[0].order_args, (("overall_confidence",), {"desc": False}))

    def test_history_export_pdf_response_headers(self):
        fake = FakeSupabase(scan_count=1)
        with self.fake_backend(fake), patch.object(main, "build_history_pdf", return_value=b"pdf"):
            response = main.export_scan_history(format="pdf", scope="scan", principal=fake_principal())

        self.assertEqual(response.media_type, "application/pdf")
        self.assertIn("purityloop-scan-history", response.headers["content-disposition"])

    def test_export_weight_prefers_persisted_positive_weight(self):
        self.assertEqual(main.export_weight({}, {"estimated_weight_kg": 1.2, "category": "plastic"}), "1.200 kg")

    def test_export_weight_preserves_zero_weight(self):
        self.assertEqual(main.export_weight({}, {"estimated_weight_kg": 0, "category": "plastic"}), "0.000 kg")

    def test_export_weight_falls_back_to_review_workspace_category_estimate(self):
        self.assertEqual(main.export_weight({}, {"category": "plastic"}), "0.032 kg")

    def test_export_weight_returns_dash_for_null_without_category(self):
        self.assertEqual(main.export_weight({}, {}), "-")

    def test_export_datetime_formats_utc_as_malaysia_time(self):
        self.assertEqual(main.format_malaysia_datetime("2026-07-29T17:00:00+00:00"), "Thu 30 Jul 1:00 AM")
        self.assertEqual(main.format_malaysia_datetime("2026-07-29T15:45:00+00:00"), "Wed 29 Jul 11:45 PM")
        self.assertEqual(main.format_malaysia_datetime("2026-07-30T04:05:00+00:00"), "Thu 30 Jul 12:05 PM")
        self.assertEqual(main.format_malaysia_datetime("2026-07-29T17:00:00Z"), "Thu 30 Jul 1:00 AM")
        self.assertEqual(main.format_malaysia_datetime(datetime(2026, 7, 29, 17, 0, 0, tzinfo=timezone.utc)), "Thu 30 Jul 1:00 AM")
        self.assertEqual(main.format_malaysia_datetime(datetime(2026, 7, 29, 17, 0, 0)), "Thu 30 Jul 1:00 AM")
        self.assertEqual(main.format_malaysia_datetime(None), "-")
        self.assertNotIn(",", main.format_malaysia_datetime("2026-07-29T17:00:00+00:00"))
        self.assertNotIn(" 01:", main.format_malaysia_datetime("2026-07-29T17:00:00+00:00"))

    def test_export_row_uses_current_review_workspace_primary_material_rule(self):
        row = main.export_scan_row({
            "id": "scan-1",
            "created_at": "2026-07-30T00:00:00Z",
            "source_name": "multi.jpg",
            "detected_materials": [
                {"id": "material-1", "category": "plastic", "confidence": 0.9},
                {"id": "material-2", "category": "glass", "confidence": 0.9},
            ],
        })

        self.assertEqual(row["estimated_weight"], "0.032 kg")
        self.assertEqual(row["datetime"], "Thu 30 Jul 8:00 AM")

    def test_history_thumbnail_fetch_deduplicates_urls_and_keeps_failures_nonfatal(self):
        rows = [{"image_url": "https://example.test/a.jpg"}, {"image_url": "https://example.test/a.jpg"}, {"image_url": "https://example.test/missing.jpg"}]

        def fake_load(_client, url):
            if "missing" in url:
                return main.ThumbnailResult(url, error="HTTPStatusError")
            return main.ThumbnailResult(url, data=b"thumb")

        with patch.object(main, "load_thumbnail", side_effect=fake_load):
            thumbnails, stats = main.fetch_history_thumbnails(rows)

        self.assertEqual(thumbnails, {"https://example.test/a.jpg": b"thumb"})
        self.assertEqual(stats["requested"], 3)
        self.assertEqual(stats["unique"], 2)
        self.assertEqual(stats["cache_hits"], 1)
        self.assertEqual(stats["failed"], 1)

    def test_history_thumbnail_uses_supabase_transform_url_for_public_storage(self):
        with patch.object(main, "SUPABASE_URL", "https://project.supabase.co"):
            transformed = main.transformed_storage_url("https://project.supabase.co/storage/v1/object/public/mock_uploaded_images/folder/image.jpg")

        self.assertIn("/storage/v1/render/image/public/mock_uploaded_images/folder/image.jpg", transformed)
        self.assertIn("width=80", transformed)
        self.assertIn("resize=contain", transformed)

    def test_history_excel_embeds_thumbnail_without_visible_raw_url(self):
        image = Image.new("RGB", (8, 8), "red")
        image_bytes = BytesIO()
        image.save(image_bytes, format="PNG")
        rows = [{
            "scan_id": "scan-1",
            "datetime": "2026-07-29T00:00:00Z",
            "file_name": "bottle.jpg",
            "predicted_category": "Plastic",
            "corrected_category": "Plastic",
            "confidence": "91%",
            "status": "Confirmed Recyclable",
            "quantity": "1",
            "estimated_weight": "0.032 kg",
            "recommended_route": "Plastic Sorting Bin",
            "review_status": "confirmed",
            "reviewer": "operator",
            "image_url": "https://example.test/image.jpg",
        }]

        workbook = load_workbook(BytesIO(main.build_history_excel(rows, {"https://example.test/image.jpg": image_bytes.getvalue()})))
        worksheet = workbook.active

        headers = [cell.value for cell in worksheet[1]]
        self.assertEqual(headers, [
            "Scan ID",
            "Date and time",
            "Category",
            "Confidence",
            "Status",
            "Recommended route",
            "Image Preview",
        ])
        self.assertEqual(len(headers), 7)
        self.assertNotIn("Image/file name", headers)
        self.assertNotIn("Estimated weight", headers)
        self.assertIsNone(worksheet.cell(2, 7).value)
        self.assertEqual(len(worksheet._images), 1)
        self.assertNotEqual(worksheet.cell(1, 7).value, "Image URL")

    def test_pdf_export_visible_columns_remove_file_and_weight(self):
        self.assertEqual(main.PDF_EXPORT_COLUMNS, ["Scan ID", "Date/time", "Category", "Confidence", "Status", "Route", "Preview"])
        self.assertNotIn("File", main.PDF_EXPORT_COLUMNS)
        self.assertNotIn("Weight", main.PDF_EXPORT_COLUMNS)

    def test_history_pdf_does_not_print_raw_image_url(self):
        rows = [{
            "scan_id": "scan-1",
            "datetime": "2026-07-29T00:00:00Z",
            "file_name": "bottle.jpg",
            "corrected_category": "Plastic",
            "confidence": "91%",
            "status": "Confirmed Recyclable",
            "estimated_weight": "0.032 kg",
            "recommended_route": "Plastic Sorting Bin",
            "image_url": "https://example.test/image.jpg",
        }]

        pdf = main.build_history_pdf(rows, {"status": "confirmed"}, {})

        self.assertTrue(pdf.startswith(b"%PDF"))
        self.assertNotIn(b"Image URL", pdf)
        self.assertNotIn(b"https://example.test/image.jpg", pdf)


if __name__ == "__main__":
    unittest.main()
